# -*- coding: utf-8 -*-
"""
네이버 스마트스토어 자동 상품등록 프로그램

엑셀(상품등록_양식.xlsx)을 읽어서 스마트스토어 상품등록 페이지의
입력칸들을 자동으로 채워 준다.

실행:  python uploader.py [엑셀파일경로]
       (엑셀 경로를 생략하면 config.EXCEL_PATH 사용)

동작 순서
  1. 크롬을 전용 프로필(chrome_profile 폴더)로 실행
  2. 로그인이 안 되어 있으면 사용자가 직접 로그인할 때까지 대기 (최초 1회만)
  3. 엑셀의 각 행마다 상품등록 페이지를 열고 카테고리/상품명/판매가/재고/
     이미지/상세설명/배송비를 자동 입력
  4. config.AUTO_SUBMIT 이 True면 [저장하기]까지 자동 클릭
  5. 각 행의 결과를 엑셀 '업로드결과' 열에 기록

※ 네이버가 판매자센터 화면을 개편하면 입력칸을 못 찾을 수 있습니다.
   그 경우 아래 SELECTORS 의 해당 항목만 고치면 됩니다. (README 참고)
"""

import os
import sys
import time
from dataclasses import dataclass, field

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

import config

HOME_URL = "https://sell.smartstore.naver.com/#/home/dashboard"
CREATE_URL = "https://sell.smartstore.naver.com/#/products/create"

# ---------------------------------------------------------------
# 입력칸을 찾는 방법(셀렉터) 모음.
# 네이버 화면이 바뀌어 특정 항목 입력에 실패하면 이 부분만 수정하면 된다.
# 각 항목은 후보 목록이며, 위에서부터 순서대로 시도해서 처음 찾은 것을 쓴다.
# ---------------------------------------------------------------
SELECTORS = {
    "카테고리검색": [
        (By.CSS_SELECTOR, "input[placeholder*='카테고리명']"),
        (By.XPATH, "//input[contains(@placeholder,'카테고리')]"),
    ],
    "카테고리첫결과": [
        (By.XPATH, "//ul[contains(@class,'dropdown-menu') or contains(@class,'result')]//li[1]//a"),
        (By.XPATH, "//ul[contains(@class,'dropdown-menu') or contains(@class,'result')]//li[1]"),
    ],
    "상품명": [
        (By.CSS_SELECTOR, "input[placeholder*='상품명']"),
        (By.XPATH, "//*[self::h3 or self::h4 or self::label][contains(.,'상품명')]/following::input[1]"),
    ],
    "판매가": [
        (By.XPATH, "//*[self::h3 or self::h4 or self::label][contains(.,'판매가')]/following::input[contains(@placeholder,'숫자')][1]"),
        (By.XPATH, "//*[self::h3 or self::h4 or self::label][contains(.,'판매가')]/following::input[1]"),
    ],
    "재고수량": [
        (By.XPATH, "//*[self::h3 or self::h4 or self::label][contains(.,'재고수량')]/following::input[contains(@placeholder,'숫자')][1]"),
        (By.XPATH, "//*[self::h3 or self::h4 or self::label][contains(.,'재고수량')]/following::input[1]"),
    ],
    "이미지파일입력": [
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'상품이미지')]/following::input[@type='file']"),
        (By.XPATH, "//input[@type='file']"),
    ],
    "상세설명작성버튼": [
        (By.XPATH, "//button[contains(.,'SmartEditor') and contains(.,'작성')]"),
        (By.XPATH, "//button[contains(.,'직접 작성')]"),
    ],
    "에디터프레임": [
        (By.XPATH, "//iframe[contains(@src,'editor') or contains(@src,'SmartEditor') or contains(@src,'se.')]"),
        (By.CSS_SELECTOR, "iframe[id*='editor'], iframe[class*='editor'], iframe[id*='se-'], iframe[class*='se-']"),
    ],
    "에디터본문": [
        (By.CSS_SELECTOR, "[contenteditable='true']"),
        (By.CSS_SELECTOR, ".se-content"),
        (By.TAG_NAME, "body"),
    ],
    "에디터등록버튼": [
        (By.XPATH, "//button[contains(.,'등록') and contains(.,'완료')]"),
        (By.XPATH, "//button[contains(.,'작성 완료')]"),
    ],
    "배송섹션": [
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'배송')]"),
    ],
    "배송비무료": [
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'배송')]/following::*[self::label or self::button][contains(normalize-space(.),'무료')][1]"),
    ],
    "배송비유료": [
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'배송')]/following::*[self::label or self::button][normalize-space(.)='유료'][1]"),
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'배송')]/following::*[self::label or self::button][contains(normalize-space(.),'유료')][1]"),
    ],
    "배송비입력": [
        (By.XPATH, "//*[self::h3 or self::h4][contains(.,'배송')]/following::input[contains(@placeholder,'숫자')][1]"),
    ],
    "저장하기": [
        (By.XPATH, "//button[contains(.,'저장하기')]"),
    ],
    "팝업확인": [
        (By.XPATH, "//button[contains(.,'확인')]"),
    ],
}

# 엑셀 헤더 (make_template.py 와 일치)
COL_CATEGORY = "카테고리"
COL_NAME = "상품명"
COL_PRICE = "판매가"
COL_STOCK = "재고수량"
COL_DESC = "상세설명"
COL_MAIN_IMG = "대표이미지"
COL_EXTRA_IMG = "추가이미지"
COL_SHIPPING = "배송비"
COL_RESULT = "업로드결과"

REQUIRED_COLS = [COL_NAME, COL_PRICE, COL_STOCK]


# ---------------------------------------------------------------
# 로그 / 사용자 확인 (GUI에서 바꿔 낄 수 있게 콜백 방식)
# ---------------------------------------------------------------
_log_fn = print


def _default_confirm(message):
    input(f"\n{message}\n확인했으면 Enter 를 누르세요... ")


_confirm_fn = _default_confirm


def log(msg):
    _log_fn(msg)


# ---------------------------------------------------------------
# 엑셀 읽기
# ---------------------------------------------------------------
@dataclass
class Product:
    row: int                       # 엑셀 행 번호 (결과 기록용)
    category: str = ""
    name: str = ""
    price: str = ""
    stock: str = ""
    description: str = ""
    main_image: str = ""
    extra_images: list = field(default_factory=list)
    shipping_fee: str = ""


def _cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def load_products(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["상품목록"] if "상품목록" in wb.sheetnames else wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        h = _cell(ws.cell(row=1, column=c).value)
        if h:
            headers[h] = c

    missing = [h for h in REQUIRED_COLS if h not in headers]
    if missing:
        raise ValueError(f"엑셀에 필수 열이 없습니다: {missing} "
                         f"(make_template.py 로 만든 양식을 사용하세요)")

    def get(row, col_name):
        c = headers.get(col_name)
        return _cell(ws.cell(row=row, column=c).value) if c else ""

    products = []
    for r in range(2, ws.max_row + 1):
        name = get(r, COL_NAME)
        if not name:
            continue

        p = Product(
            row=r,
            category=get(r, COL_CATEGORY),
            name=name,
            price=get(r, COL_PRICE).replace(",", ""),
            stock=get(r, COL_STOCK).replace(",", ""),
            description=get(r, COL_DESC),
            main_image=get(r, COL_MAIN_IMG),
            shipping_fee=get(r, COL_SHIPPING).replace(",", ""),
        )
        extra = get(r, COL_EXTRA_IMG)
        if extra:
            p.extra_images = [s.strip() for s in extra.split(",") if s.strip()]

        problems = []
        if not p.price.isdigit():
            problems.append("판매가는 숫자만 입력")
        if not p.stock.isdigit():
            problems.append("재고수량은 숫자만 입력")
        for img in [p.main_image] + p.extra_images:
            if img and not os.path.isfile(img):
                problems.append(f"이미지 파일 없음: {img}")
        if problems:
            log(f"[건너뜀] {r}행 '{name}': " + " / ".join(problems))
            _write_result(excel_path, r, "건너뜀: " + " / ".join(problems))
            continue

        products.append(p)

    return products


def _write_result(excel_path, row, text):
    """엑셀의 '업로드결과' 열에 결과를 기록한다. (열이 없으면 만들어서)"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["상품목록"] if "상품목록" in wb.sheetnames else wb.active
        result_col = None
        for c in range(1, ws.max_column + 1):
            if _cell(ws.cell(row=1, column=c).value) == COL_RESULT:
                result_col = c
                break
        if result_col is None:
            result_col = ws.max_column + 1
            ws.cell(row=1, column=result_col, value=COL_RESULT)
        ws.cell(row=row, column=result_col, value=text)
        wb.save(excel_path)
    except PermissionError:
        log("※ 엑셀이 열려 있어서 결과를 기록하지 못했습니다. 엑셀을 닫아주세요.")
    except Exception as e:
        log(f"※ 결과 기록 실패: {e}")


# ---------------------------------------------------------------
# 브라우저 제어
# ---------------------------------------------------------------
def create_driver():
    opts = Options()
    profile_dir = os.path.abspath(config.CHROME_PROFILE_DIR)
    opts.add_argument(f"--user-data-dir={profile_dir}")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                              options=opts)
    return driver


def ensure_login(driver, stop_flag=None):
    driver.get(HOME_URL)
    time.sleep(3)

    def on_login_page():
        url = driver.current_url
        return "nid.naver.com" in url or "login" in url

    if on_login_page():
        log("=" * 50)
        log("로그인이 필요합니다.")
        log("열린 크롬 창에서 직접 네이버 로그인을 해주세요.")
        log("(한 번 로그인하면 다음부터는 자동으로 유지됩니다)")
        log("=" * 50)
        while on_login_page():
            if stop_flag and stop_flag():
                raise RuntimeError("사용자가 중지했습니다.")
            time.sleep(2)
        time.sleep(3)
    log("로그인 확인 완료")


def find_first(driver, key, timeout=None, visible=True):
    """SELECTORS[key] 의 후보들을 순서대로 시도해 처음 발견한 요소를 반환."""
    candidates = SELECTORS[key]
    timeout = config.WAIT_TIMEOUT if timeout is None else timeout
    end = time.time() + timeout
    while time.time() < end:
        for by, sel in candidates:
            try:
                for el in driver.find_elements(by, sel):
                    if not visible or el.is_displayed():
                        return el
            except Exception:
                pass
        time.sleep(0.5)
    raise TimeoutError(f"'{key}' 입력칸을 찾지 못했습니다. "
                       f"화면이 바뀌었다면 uploader.py 의 SELECTORS['{key}'] 를 수정하세요.")


def fill_text(driver, el, text):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    time.sleep(0.2)
    el.click()
    el.send_keys(Keys.CONTROL, "a")
    el.send_keys(Keys.DELETE)
    el.send_keys(str(text))
    time.sleep(config.ACTION_DELAY)


def close_popups(driver):
    """공지 팝업 등이 떠 있으면 닫는다 (실패해도 무시)."""
    for text in ("오늘 하루 보지 않기", "닫기"):
        try:
            for el in driver.find_elements(By.XPATH, f"//button[contains(.,'{text}')]"):
                if el.is_displayed():
                    el.click()
                    time.sleep(0.3)
        except Exception:
            pass


# ---------------------------------------------------------------
# 항목별 입력
# ---------------------------------------------------------------
def fill_category(driver, category):
    inp = find_first(driver, "카테고리검색")
    fill_text(driver, inp, category)
    time.sleep(1.5)
    try:
        item = find_first(driver, "카테고리첫결과", timeout=8)
        item.click()
    except TimeoutError:
        # 목록 클릭 실패 시 키보드로 첫 항목 선택
        inp.send_keys(Keys.ARROW_DOWN)
        time.sleep(0.3)
        inp.send_keys(Keys.ENTER)
    time.sleep(config.ACTION_DELAY)


def fill_name(driver, name):
    fill_text(driver, find_first(driver, "상품명"), name)


def fill_price(driver, price):
    fill_text(driver, find_first(driver, "판매가"), price)


def fill_stock(driver, stock):
    fill_text(driver, find_first(driver, "재고수량"), stock)


def fill_images(driver, product):
    paths = []
    if product.main_image:
        paths.append(os.path.abspath(product.main_image))
    paths += [os.path.abspath(p) for p in product.extra_images]
    if not paths:
        return

    file_input = find_first(driver, "이미지파일입력", visible=False)
    # 대표이미지 먼저, 이후 추가이미지 순서로 업로드
    for path in paths:
        file_input.send_keys(path)
        time.sleep(2)  # 업로드 처리 대기
        # 업로드마다 file input 이 새로 생기는 경우가 있어 다시 찾는다
        try:
            file_input = find_first(driver, "이미지파일입력", timeout=5, visible=False)
        except TimeoutError:
            break


def fill_description(driver, text):
    if not text:
        return
    btn = find_first(driver, "상세설명작성버튼")
    driver.execute_script("arguments[0].click();", btn)
    time.sleep(3)  # 에디터 로딩 대기

    iframe = find_first(driver, "에디터프레임", visible=False)
    driver.switch_to.frame(iframe)
    try:
        body = find_first(driver, "에디터본문", timeout=10)
        body.click()
        time.sleep(0.3)
        for line in str(text).split("\n"):
            body.send_keys(line)
            body.send_keys(Keys.ENTER)
        time.sleep(0.5)
    finally:
        driver.switch_to.default_content()

    # '등록 완료' 류의 버튼이 있으면 눌러서 에디터를 닫는다
    try:
        done = find_first(driver, "에디터등록버튼", timeout=5)
        driver.execute_script("arguments[0].click();", done)
        time.sleep(1)
    except TimeoutError:
        pass


def fill_shipping(driver, fee):
    section = find_first(driver, "배송섹션", timeout=8)
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", section)
    time.sleep(0.3)

    if fee in ("", "0"):
        el = find_first(driver, "배송비무료", timeout=8)
        driver.execute_script("arguments[0].click();", el)
    else:
        el = find_first(driver, "배송비유료", timeout=8)
        driver.execute_script("arguments[0].click();", el)
        time.sleep(0.5)
        inp = find_first(driver, "배송비입력", timeout=8)
        fill_text(driver, inp, fee)
    time.sleep(config.ACTION_DELAY)


def submit(driver):
    btn = find_first(driver, "저장하기")
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.5)
    driver.execute_script("arguments[0].click();", btn)
    time.sleep(2)
    # 저장 확인 팝업이 뜨면 확인 클릭
    try:
        ok = find_first(driver, "팝업확인", timeout=5)
        ok.click()
    except TimeoutError:
        pass
    time.sleep(2)


# ---------------------------------------------------------------
# 상품 1건 등록
# ---------------------------------------------------------------
def register_product(driver, product, auto_submit):
    driver.get(CREATE_URL)
    time.sleep(4)  # 페이지(SPA) 로딩 대기
    close_popups(driver)

    failed = []

    def step(label, fn, *args, skip_if=False):
        if skip_if:
            return
        try:
            fn(*args)
            log(f"  [OK] {label}")
        except Exception as e:
            failed.append(label)
            log(f"  [실패] {label}: {e}")

    step("카테고리", fill_category, driver, product.category,
         skip_if=not product.category)
    step("상품명", fill_name, driver, product.name)
    step("판매가", fill_price, driver, product.price)
    step("재고수량", fill_stock, driver, product.stock)
    step("이미지", fill_images, driver, product,
         skip_if=not (product.main_image or product.extra_images))
    step("상세설명", fill_description, driver, product.description,
         skip_if=not product.description)
    step("배송비", fill_shipping, driver, product.shipping_fee)

    if failed:
        return False, failed

    if auto_submit:
        try:
            submit(driver)
            log("  [OK] 저장하기 클릭")
        except Exception as e:
            log(f"  [실패] 저장하기: {e}")
            return False, ["저장하기"]
    else:
        _confirm_fn(f"'{product.name}' 입력이 끝났습니다.\n"
                    f"브라우저에서 내용을 확인하고 [저장하기]를 직접 누른 뒤 계속하세요.")

    return True, []


# ---------------------------------------------------------------
# 메인 실행
# ---------------------------------------------------------------
def run(excel_path=None, auto_submit=None, log_fn=None, confirm_fn=None, stop_flag=None):
    """전체 업로드 실행. GUI에서도 이 함수를 그대로 호출한다.

    stop_flag: 인자 없이 호출하면 True/False 를 돌려주는 함수.
               True 가 되면 다음 상품으로 넘어가지 않고 중단한다.
    """
    global _log_fn, _confirm_fn
    if log_fn:
        _log_fn = log_fn
    if confirm_fn:
        _confirm_fn = confirm_fn

    excel_path = os.path.abspath(excel_path or config.EXCEL_PATH)
    if auto_submit is None:
        auto_submit = config.AUTO_SUBMIT

    if not os.path.isfile(excel_path):
        log(f"엑셀 파일이 없습니다: {excel_path}")
        log("먼저  python make_template.py  로 양식을 만들고 채워주세요.")
        return

    log(f"엑셀 읽는 중: {excel_path}")
    products = load_products(excel_path)
    if not products:
        log("등록할 상품이 없습니다. (상품명이 비어 있는 행은 건너뜁니다)")
        return
    log(f"등록할 상품: {len(products)}개 / 자동저장: {'켜짐' if auto_submit else '꺼짐(직접 저장)'}")

    driver = create_driver()
    ok_count = 0
    try:
        ensure_login(driver, stop_flag)

        for i, p in enumerate(products, start=1):
            if stop_flag and stop_flag():
                log("사용자 요청으로 중단합니다.")
                break

            log(f"\n[{i}/{len(products)}] {p.name}")
            try:
                success, failed = register_product(driver, p, auto_submit)
            except Exception as e:
                success, failed = False, [f"오류: {e}"]
                log(f"  [실패] {e}")

            if success:
                ok_count += 1
                _write_result(excel_path, p.row, "성공")
            else:
                _write_result(excel_path, p.row, "실패: " + ", ".join(failed))

            time.sleep(config.BETWEEN_PRODUCTS_DELAY)

        log(f"\n완료: 성공 {ok_count}개 / 전체 {len(products)}개")
        log(f"각 행의 결과는 엑셀 '{COL_RESULT}' 열에 기록했습니다.")
    finally:
        try:
            _confirm_fn("작업이 끝났습니다. 브라우저를 닫으려면 계속하세요.")
        except Exception:
            pass
        driver.quit()


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    run(excel_path=path)
