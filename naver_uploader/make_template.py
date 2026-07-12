# -*- coding: utf-8 -*-
"""
상품등록용 엑셀 양식(상품등록_양식.xlsx)을 만들어 주는 스크립트.

실행:  python make_template.py
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config

# 열 정의: (헤더, 열 너비, 예시값)
COLUMNS = [
    ("카테고리",   30, "패션잡화>모자>버킷햇"),
    ("상품명",     40, "여름용 코튼 버킷햇 (5색상)"),
    ("판매가",     12, 19900),
    ("재고수량",   10, 100),
    ("상세설명",   60, "가볍고 통기성 좋은 여름용 버킷햇입니다.\n데일리룩에 잘 어울립니다."),
    ("대표이미지", 40, r"C:\상품사진\버킷햇_대표.jpg"),
    ("추가이미지", 50, r"C:\상품사진\버킷햇_1.jpg, C:\상품사진\버킷햇_2.jpg"),
    ("배송비",     10, 0),
]

GUIDE = [
    "■ 작성 방법",
    "",
    "1. '상품목록' 시트의 2행(예시)을 참고해서 한 줄에 상품 하나씩 입력하세요.",
    "   예시 줄은 지우고 실제 상품으로 채우면 됩니다.",
    "",
    "2. 각 열 설명",
    "   - 카테고리   : 카테고리 검색어. 전체 경로(패션잡화>모자>버킷햇) 또는",
    "                  마지막 카테고리명(버킷햇)만 적어도 됩니다. 검색 결과의 첫 항목이 선택됩니다.",
    "   - 상품명     : 필수. 스마트스토어에 표시될 상품명.",
    "   - 판매가     : 필수. 숫자만 입력 (예: 19900)",
    "   - 재고수량   : 필수. 숫자만 입력 (예: 100)",
    "   - 상세설명   : 상품 상세페이지에 들어갈 글. 셀 안에서 줄바꿈(Alt+Enter) 가능.",
    "   - 대표이미지 : 내 컴퓨터에 있는 이미지 파일의 전체 경로.",
    "   - 추가이미지 : 여러 장이면 쉼표(,)로 구분. 없으면 비워두세요.",
    "   - 배송비     : 0 또는 빈칸이면 무료배송, 숫자를 넣으면 유료배송(해당 금액).",
    "",
    "3. 작성이 끝나면 저장하고, uploader.py(또는 gui.py)를 실행하세요.",
    "",
    "※ 옵션(색상/사이즈 등) 자동 입력은 지원하지 않습니다.",
    "   옵션이 필요한 상품은 등록 후 스마트스토어에서 직접 추가해 주세요.",
]


def main():
    wb = Workbook()

    ws = wb.active
    ws.title = "상품목록"

    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(color="FFFFFF", bold=True)

    for i, (name, width, example) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width
        ws.cell(row=2, column=i, value=example).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    guide = wb.create_sheet("작성방법")
    guide.column_dimensions["A"].width = 100
    for r, line in enumerate(GUIDE, start=1):
        guide.cell(row=r, column=1, value=line)

    path = os.path.abspath(config.EXCEL_PATH)
    wb.save(path)
    print(f"엑셀 양식을 만들었습니다: {path}")
    print("'상품목록' 시트를 채운 뒤 gui.py 또는 uploader.py 를 실행하세요.")


if __name__ == "__main__":
    main()
