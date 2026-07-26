#!/usr/bin/env python3
"""
바이낸스 선물 백테스터 — 완전 독립 실행형 (GUI)

🔥 이 파일 하나만 있으면 실행됩니다 (봇 파일 불필요, API 키 불필요).
   필요 라이브러리(pandas/numpy/requests)는 실행 시 자동 설치됩니다.

실행:
  python backtest.py            ← GUI 실행 (더블클릭도 가능)
  python backtest.py --compare  ← 콘솔(CLI) 모드

기능:
- 모든 수치(진입금/레버리지/수수료/TP/ADX/UT/EMA/기간/시간봉) GUI에서 수정 가능
  → 기본값은 현재 봇 설정과 동일
- 바이낸스 USDT 선물 전체 코인 목록 자동 로드, 검색/다중선택/전체선택
- 비교 모드: 설정값vs하이브리드 / 볼륨필터OFFvsON / EMA4종 / 단일
- 거래량(볼륨) 필터 ON/OFF + 배수·평균기간 조절 가능
- 결과: 화면 표 + backtest_result.csv + 거래내역 CSV
- 📅 월별 수익/수수료/손실 집계 엑셀(backtest_월별집계.xlsx) 자동 생성

전략 (봇 5_gui.py와 동일 로직):
- 진입: 완성봉 기준 UT Bot 방향 + EMA 크로스 AND 조건
- 익절: 진입 시 ADX >= 기준이면 추세장 TP, 아니면 횡보장 TP
        🔥 ADX는 별도 시간봉(기본 1시간봉)의 완성봉으로 계산
        (거래소 TP 주문 = TP 가격에 정확히 청산되는 것으로 모델링)
- 손절: 고정 손절 없음. 역신호 시 청산 + 즉시 반대 진입(스위칭)
- 하이브리드(선택): 손실 중일 때만 빠른 EMA 크로스로 조기청산
"""

import argparse
import os
import queue
import subprocess
import sys
import threading
import time

# ==================== 라이브러리 자동 설치 ====================
for _mod in ['pandas', 'numpy', 'requests', 'openpyxl']:
    try:
        __import__(_mod)
    except ImportError:
        print(f"📦 {_mod} 설치 중...")
        try:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', _mod],
                                  stdout=subprocess.DEVNULL)
        except Exception:
            pass  # openpyxl 설치 실패해도 CSV는 됨 (엑셀만 생략)

import numpy as np
import pandas as pd
import requests

try:
    import openpyxl  # noqa: F401
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ==================== 기본값 (현재 봇 설정과 동일) ====================
DEFAULTS = {
    'amount': 50.0,        # 진입금 (USDT)
    'leverage': 3,         # 레버리지
    'fee_pct': 0.04,       # 수수료 % (테이커, 진입/청산 각각)
    'tp_trend': 1.2,       # 추세장 TP %
    'tp_sideways': 1.0,    # 횡보장 TP %
    'adx_period': 10,      # ADX 기간
    'adx_th': 21,          # ADX 추세장 기준
    'adx_interval': '1h',  # 🔥 ADX 계산 시간봉 (TP 결정용, 매매봉과 별개)
    'ut_sens': 10.0,       # UT Bot Key Value
    'ut_atr': 2,           # UT Bot ATR Period
    'ema_fast': 34,        # EMA Fast
    'ema_slow': 55,        # EMA Slow
    'days': 365,           # 백테스트 기간 (일)
    'interval': '5m',      # 시간봉
    'hybrid': False,       # 하이브리드 조기청산
    'hybrid_fast': 9,      # 하이브리드 빠른 EMA Fast
    'hybrid_slow': 21,     # 하이브리드 빠른 EMA Slow
    'hybrid_roi_th': -5.0, # 조기청산 발동 ROI % (이하 손실일 때만)
    'vol_filter': False,   # 🔥 거래량 필터 ON/OFF
    'vol_mult': 1.5,       # 거래량 배수 (현재봉 ≥ 평균 × 배수일 때만 진입)
    'vol_ma': 20,          # 거래량 평균 기간 (봉)
    # 🔥 익절 방식: 'adx'=지금(ADX 고정 TP) / 'atr'=변동성 기반 / 'switch'=TP없이 스위칭만
    'tp_mode': 'adx',
    'atr_tp_period': 14,   # ATR TP용 ATR 기간
    'atr_tp_mult': 1.5,    # ATR TP 배수 (TP거리 = ATR × 배수) ← 중간값 기본
}

DEFAULT_SYMBOLS = ['BTCUSDT', 'XRPUSDT', 'DOGEUSDT']
FALLBACK_SYMBOLS = [
    'BTCUSDT', 'ETHUSDT', 'BNBUSDT', 'SOLUSDT', 'XRPUSDT', 'ADAUSDT',
    'DOGEUSDT', 'TRXUSDT', 'TONUSDT', 'LINKUSDT', 'AVAXUSDT', 'DOTUSDT',
    'LTCUSDT', 'BCHUSDT', 'NEARUSDT', 'APTUSDT', 'SUIUSDT', 'ARBUSDT',
    'OPUSDT', 'FILUSDT',
]
INTERVALS = ['1m', '3m', '5m', '15m', '30m', '1h', '2h', '4h']
RESAMPLE_RULE = {'1m': '1min', '3m': '3min', '5m': '5min', '15m': '15min',
                 '30m': '30min', '1h': '1h', '2h': '2h', '4h': '4h'}


# ==================== 쓰기 가능한 작업 폴더 찾기 ====================
_WORKDIR = None


def get_workdir(log=print):
    """결과/캐시를 저장할 쓰기 가능한 폴더를 찾는다.

    실행 폴더가 보호되어 있으면(WinError 5 등) 홈/문서/임시 폴더 순으로
    폴백. 모두 실패하면 None (캐시·저장 생략).
    """
    global _WORKDIR
    if _WORKDIR is not None:
        return _WORKDIR

    candidates = []
    try:
        candidates.append(os.path.dirname(os.path.abspath(__file__)))
    except Exception:
        pass
    candidates.append(os.getcwd())
    home = os.path.expanduser('~')
    candidates.append(os.path.join(home, 'Documents', '바이낸스백테스트'))
    candidates.append(os.path.join(home, '바이낸스백테스트'))
    import tempfile
    candidates.append(os.path.join(tempfile.gettempdir(), '바이낸스백테스트'))

    for base in candidates:
        try:
            os.makedirs(base, exist_ok=True)
            testfile = os.path.join(base, '.write_test')
            with open(testfile, 'w') as f:
                f.write('ok')
            os.remove(testfile)
            _WORKDIR = base
            log(f"💾 저장 폴더: {base}")
            return base
        except Exception:
            continue

    log("⚠️ 쓰기 가능한 폴더를 못 찾음 — 결과는 화면에만 표시(파일 저장 생략)")
    _WORKDIR = ''  # 빈 문자열 = 저장 불가 표시 (재탐색 방지)
    return ''


# ==================== 데이터 ====================
def fetch_symbols(log=print):
    """바이낸스 USDT 선물 무기한 전체 심볼"""
    try:
        r = requests.get('https://fapi.binance.com/fapi/v1/exchangeInfo', timeout=15)
        r.raise_for_status()
        syms = sorted(
            s['symbol'] for s in r.json()['symbols']
            if s.get('quoteAsset') == 'USDT'
            and s.get('contractType') == 'PERPETUAL'
            and s.get('status') == 'TRADING'
        )
        log(f"✅ 바이낸스 선물 심볼 {len(syms)}개 로드 완료")
        return syms
    except Exception as e:
        log(f"⚠️ 심볼 목록 로드 실패({e}) — 기본 20개 사용")
        return list(FALLBACK_SYMBOLS)


def _parse_date_ms(s):
    """'2024-01-01' 또는 '2024/01/01' 등 → epoch ms. 빈값이면 None."""
    if s is None or not str(s).strip():
        return None
    txt = str(s).strip().replace('/', '-').replace('.', '-')
    try:
        return int(pd.Timestamp(txt).timestamp() * 1000)
    except Exception:
        raise ValueError(f"날짜 형식 오류: '{s}' (예: 2024-01-01)")


def fetch_klines(symbol, days, interval='5m', cache_dir=None, log=print,
                 start_date=None, end_date=None):
    """기간 지정 방식 두 가지:
       - days: 최근 N일 (start_date/end_date 없을 때)
       - start_date/end_date: 'YYYY-MM-DD' 구간 (있으면 우선)
    """
    start_ms = _parse_date_ms(start_date)
    end_ms = _parse_date_ms(end_date)
    use_range = start_ms is not None or end_ms is not None

    # 🔥 쓰기 가능한 폴더 안에 캐시 저장 (권한 오류 방지)
    base = get_workdir(log)
    cache = None
    if base:
        cache_dir = os.path.join(base, 'backtest_data')
        try:
            os.makedirs(cache_dir, exist_ok=True)
            if use_range:
                tag = f"{start_date or 'x'}_{end_date or 'now'}".replace(':', '').replace(' ', '')
                cache = os.path.join(cache_dir, f"{symbol}_{interval}_{tag}.csv")
            else:
                cache = os.path.join(cache_dir, f"{symbol}_{interval}_{days}d.csv")
        except Exception as e:
            log(f"  ⚠️ 캐시 폴더 생성 실패({e}) — 캐시 없이 진행")
            cache = None
    if cache and os.path.exists(cache):
        df = pd.read_csv(cache, index_col=0, parse_dates=True)
        log(f"  📂 캐시 사용: {symbol} ({len(df):,}봉)")
        return df

    url = "https://fapi.binance.com/fapi/v1/klines"
    now_ms = int(time.time() * 1000)
    if use_range:
        end = end_ms if end_ms is not None else now_ms
        start = start_ms if start_ms is not None else end - 365 * 86400 * 1000
        if start >= end:
            raise ValueError(f"{symbol}: 시작일이 종료일보다 늦습니다")
    else:
        end = now_ms
        start = end - days * 86400 * 1000
    rows = []
    cur = start
    last_log = 0
    while cur < end:
        data = None
        for attempt in range(5):
            try:
                r = requests.get(url, params={
                    'symbol': symbol, 'interval': interval,
                    'startTime': cur, 'limit': 1500}, timeout=15)
                r.raise_for_status()
                data = r.json()
                break
            except Exception as e:
                log(f"  ⚠️ 재시도 {attempt + 1}/5: {e}")
                time.sleep(2 * (attempt + 1))
        if data is None:
            raise RuntimeError(f"{symbol} 다운로드 실패 - 네트워크 확인 필요")
        if not data:
            break
        rows.extend(data)
        cur = data[-1][6] + 1
        if len(rows) - last_log >= 15000:
            log(f"  ⬇️  {symbol}: {len(rows):,}봉...")
            last_log = len(rows)
        time.sleep(0.15)
        if len(data) < 1500:
            break

    if not rows:
        raise RuntimeError(f"{symbol}: 데이터 없음 (신규 상장 코인일 수 있음)")

    df = pd.DataFrame(rows, columns=[
        'ts', 'open', 'high', 'low', 'close', 'volume',
        'close_time', 'qv', 'n', 'tb', 'tq', 'ig'])
    df = df[['ts', 'open', 'high', 'low', 'close', 'volume']].astype(float)
    df['ts'] = pd.to_datetime(df['ts'], unit='ms')
    df = df.set_index('ts')
    df = df[~df.index.duplicated(keep='first')].sort_index()
    if cache:
        try:
            df.to_csv(cache)
            log(f"  💾 {symbol}: {len(df):,}봉 다운로드 완료 (캐시 저장)")
        except Exception as e:
            log(f"  ⚠️ {symbol}: {len(df):,}봉 다운로드 완료 (캐시 저장 실패: {e})")
    else:
        log(f"  ✅ {symbol}: {len(df):,}봉 다운로드 완료")
    return df


# ==================== 지표 (봇 3_indicators.py와 동일 공식) ====================
def atr_rma(df, period):
    high, low, close = df['high'], df['low'], df['close']
    tr = pd.concat([high - low,
                    (high - close.shift(1)).abs(),
                    (low - close.shift(1)).abs()], axis=1).max(axis=1)
    return tr.ewm(alpha=1 / period, adjust=False).mean()


def ut_bot(df, sensitivity, atr_period):
    atr = atr_rma(df, atr_period)
    close = df['close'].values
    n_loss = (sensitivity * atr).values
    stop = np.zeros(len(close))
    pos = np.zeros(len(close), dtype=int)
    stop[0] = close[0] - n_loss[0] if not np.isnan(n_loss[0]) else close[0]
    pos[0] = 1
    for i in range(1, len(close)):
        if np.isnan(n_loss[i]):
            stop[i] = stop[i - 1]
            pos[i] = pos[i - 1]
            continue
        prev_stop, cur, prev = stop[i - 1], close[i], close[i - 1]
        if cur > prev_stop and prev > prev_stop:
            stop[i] = max(prev_stop, cur - n_loss[i])
        elif cur < prev_stop and prev < prev_stop:
            stop[i] = min(prev_stop, cur + n_loss[i])
        elif cur > prev_stop:
            stop[i] = cur - n_loss[i]
        else:
            stop[i] = cur + n_loss[i]
        if prev < stop[i - 1] and cur > stop[i]:
            pos[i] = 1
        elif prev > stop[i - 1] and cur < stop[i]:
            pos[i] = -1
        else:
            pos[i] = pos[i - 1]
    return pos


def ema_pair(df, fast, slow):
    return (df['close'].ewm(span=fast, adjust=False).mean(),
            df['close'].ewm(span=slow, adjust=False).mean())


def adx_on_interval(df, period, adx_interval, trade_interval):
    """🔥 ADX를 별도 시간봉으로 계산해 매매봉 인덱스에 정렬

    - adx_interval == trade_interval이면 매매봉 그대로 계산
    - 다르면 상위 봉으로 리샘플 후 계산하고, '완성된 상위 봉'만 사용
      (shift(1) → 진행 중인 상위 봉을 참조하는 미래 참조 방지)
    """
    if adx_interval == trade_interval:
        return adx_full_series(df, period)
    rule = RESAMPLE_RULE.get(adx_interval, adx_interval)
    ohlc = df.resample(rule).agg({
        'open': 'first', 'high': 'max', 'low': 'min', 'close': 'last'}).dropna()
    adx_h = adx_full_series(ohlc, period)
    return adx_h.shift(1).reindex(df.index, method='ffill').fillna(20.0)


def adx_full_series(df, period):
    high, low, close = df['high'], df['low'], df['close']
    tr = pd.concat([high - low,
                    (high - close.shift(1)).abs(),
                    (low - close.shift(1)).abs()], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1 / period, adjust=False).mean()
    up = high - high.shift(1)
    down = low.shift(1) - low
    plus_dm = pd.Series(np.where((up > down) & (up > 0), up, 0), index=df.index)
    minus_dm = pd.Series(np.where((down > up) & (down > 0), down, 0), index=df.index)
    plus_di = 100 * plus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr
    minus_di = 100 * minus_dm.ewm(alpha=1 / period, adjust=False).mean() / atr
    dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di)
    return dx.ewm(alpha=1 / period, adjust=False).mean().fillna(20.0)


# ==================== 시뮬레이션 ====================
def run_backtest(df, p):
    """p: 파라미터 dict (DEFAULTS 형식)"""
    o = df['open'].values
    h = df['high'].values
    l = df['low'].values
    c = df['close'].values

    fee_rate = p['fee_pct'] / 100.0
    amount = p['amount']
    lev = p['leverage']
    liq_move = 100.0 / lev
    warmup = max(60, int(p['ema_slow']) + 5)

    ut = ut_bot(df, p['ut_sens'], p['ut_atr'])
    fast, slow = ema_pair(df, p['ema_fast'], p['ema_slow'])
    ema_long = (fast > slow).values
    ema_short = (fast < slow).values
    adx = adx_on_interval(df, p['adx_period'],
                          p.get('adx_interval', p['interval']), p['interval']).values

    long_sig = (ut == 1) & ema_long
    short_sig = (ut == -1) & ema_short

    # 🔥 거래량 필터: 현재봉 거래량 >= 평균(vol_ma봉) × vol_mult 일 때만 진입 허용
    #    (봇 3_indicators.check_volume과 동일 로직, 매매봉 기준)
    if p.get('vol_filter'):
        vma = int(p.get('vol_ma', 20))
        vol_avg = df['volume'].rolling(vma).mean()
        vol_ok = (df['volume'] >= vol_avg * p.get('vol_mult', 1.5))
        vol_ok = vol_ok.fillna(False).values  # 평균 계산 전(NaN) 구간은 진입 차단
        long_sig = long_sig & vol_ok
        short_sig = short_sig & vol_ok

    if p.get('hybrid'):
        hf, hs = ema_pair(df, p['hybrid_fast'], p['hybrid_slow'])
        h_dead = (hf < hs).values
        h_gold = (hf > hs).values
        roi_th = p['hybrid_roi_th']

    # 🔥 익절 방식
    tp_mode = p.get('tp_mode', 'adx')
    if tp_mode == 'atr':
        atr_tp = atr_rma(df, int(p.get('atr_tp_period', 14))).values  # TP거리 계산용 ATR

    trades = []
    pos = None
    equity = 0.0
    peak = 0.0
    max_dd = 0.0

    def tp_pct_at(i):
        return p['tp_trend'] if adx[i] >= p['adx_th'] else p['tp_sideways']

    def open_pos(side, i):
        entry = o[i + 1]
        if tp_mode == 'switch':
            # TP 없음 — 반대신호(스위칭)나 강제청산으로만 나감
            return {'side': side, 'entry': entry, 'qty': amount * lev / entry,
                    'tp_price': None, 'tp_pct': 0.0, 'entry_i': i + 1, 'min_roi': 0.0}
        if tp_mode == 'atr':
            # TP 거리 = ATR × 배수 (신호봉 i의 ATR 사용 = 완성봉)
            atr_val = atr_tp[i]
            if np.isnan(atr_val) or atr_val <= 0:
                dist_pct = p['tp_sideways']  # ATR 아직이면 안전값
            else:
                dist_pct = (atr_val * p.get('atr_tp_mult', 1.5)) / entry * 100
            tp = dist_pct
        else:  # 'adx'
            tp = tp_pct_at(i)
        tp_price = entry * (1 + tp / 100) if side == 'LONG' else entry * (1 - tp / 100)
        return {'side': side, 'entry': entry, 'qty': amount * lev / entry,
                'tp_price': tp_price, 'tp_pct': round(tp, 3), 'entry_i': i + 1, 'min_roi': 0.0}

    def close_pos(pp, exit_price, exit_i, reason):
        nonlocal equity, peak, max_dd
        if pp['side'] == 'LONG':
            gross = pp['qty'] * (exit_price - pp['entry'])
        else:
            gross = pp['qty'] * (pp['entry'] - exit_price)
        fee = fee_rate * pp['qty'] * (pp['entry'] + exit_price)
        net = gross - fee
        equity += net
        peak = max(peak, equity)
        max_dd = min(max_dd, equity - peak)
        trades.append({
            '시각': df.index[exit_i], '포지션': pp['side'], '진입가': pp['entry'],
            '청산가': exit_price, 'TP%': pp['tp_pct'],
            'ROI%': round(gross / amount * 100, 2), '수익': round(gross, 4),
            '수수료': round(fee, 4), '순손익': round(net, 4),
            '최저ROI%': round(pp['min_roi'], 2), '유형': reason,
            '보유(봉)': exit_i - pp['entry_i'],
        })

    i = warmup
    while i < len(df) - 1:
        if pos is None:
            if long_sig[i]:
                pos = open_pos('LONG', i)
            elif short_sig[i]:
                pos = open_pos('SHORT', i)
            i += 1
            continue

        j = max(i, pos['entry_i'])
        if j >= len(df) - 1:
            break

        if pos['side'] == 'LONG':
            liq_price = pos['entry'] * (1 - liq_move / 100)
            roi_low = (l[j] - pos['entry']) / pos['entry'] * 100 * lev
            roi_close = (c[j] - pos['entry']) / pos['entry'] * 100 * lev
            pos['min_roi'] = min(pos['min_roi'], roi_low)
            if l[j] <= liq_price:
                close_pos(pos, liq_price, j, '강제청산')
                pos = None
            elif pos['tp_price'] is not None and h[j] >= pos['tp_price']:
                close_pos(pos, max(pos['tp_price'], o[j]), j, 'TP익절')
                pos = None
            elif short_sig[j]:
                close_pos(pos, o[j + 1], j, '스위칭')
                pos = open_pos('SHORT', j)
            elif p.get('hybrid') and roi_close <= roi_th and h_dead[j]:
                close_pos(pos, o[j + 1], j, '조기청산')
                pos = None
        else:
            liq_price = pos['entry'] * (1 + liq_move / 100)
            roi_low = (pos['entry'] - h[j]) / pos['entry'] * 100 * lev
            roi_close = (pos['entry'] - c[j]) / pos['entry'] * 100 * lev
            pos['min_roi'] = min(pos['min_roi'], roi_low)
            if h[j] >= liq_price:
                close_pos(pos, liq_price, j, '강제청산')
                pos = None
            elif pos['tp_price'] is not None and l[j] <= pos['tp_price']:
                close_pos(pos, min(pos['tp_price'], o[j]), j, 'TP익절')
                pos = None
            elif long_sig[j]:
                close_pos(pos, o[j + 1], j, '스위칭')
                pos = open_pos('LONG', j)
            elif p.get('hybrid') and roi_close <= roi_th and h_gold[j]:
                close_pos(pos, o[j + 1], j, '조기청산')
                pos = None

        i = j + 1

    return pd.DataFrame(trades), max_dd


def summarize(trades, symbol, cfg_name, max_dd, p):
    if trades.empty:
        return None
    total = len(trades)
    wins = int((trades['순손익'] > 0).sum())
    win_rate = wins / total * 100
    return {
        '심볼': symbol, '설정': cfg_name, '거래수': total,
        '승률%': round(win_rate, 1),
        'TP익절': int((trades['유형'] == 'TP익절').sum()),
        '스위칭': int((trades['유형'] == '스위칭').sum()),
        '조기청산': int((trades['유형'] == '조기청산').sum()),
        '강제청산': int((trades['유형'] == '강제청산').sum()),
        '총수익': round(trades['수익'].sum(), 2),
        '수수료': round(trades['수수료'].sum(), 2),
        '순손익': round(trades['순손익'].sum(), 2),
        '최대단일손실': round(trades['순손익'].min(), 2),
        '최저ROI%': round(trades['최저ROI%'].min(), 1),
        '최대낙폭': round(max_dd, 2),
    }


def make_configs(p, compare):
    """실행할 설정 목록.

    compare: 'off'  → 현재 설정 1종
             'vs'   → 설정값 vs 하이브리드 (2종)  ← 기본 추천
             'full' → EMA 4종 비교
             'vol'  → 거래량 필터 OFF vs ON (2종)
             'tp'   → 익절방식 비교: 지금(ADX) vs 스위칭만 vs ATR(여러 배수)
    (하위호환: True=='full', False=='off')
    """
    if compare is True:
        compare = 'full'
    elif compare is False:
        compare = 'off'

    setval_name = f"EMA{int(p['ema_fast'])}-{int(p['ema_slow'])}(설정값)"

    if compare == 'off':
        base_name = f"EMA{int(p['ema_fast'])}-{int(p['ema_slow'])}"
        if p.get('hybrid'):
            base_name += "+하이브리드"
        return [dict(p, name=base_name)]

    if compare == 'vs':
        return [
            dict(p, name=setval_name, hybrid=False),
            dict(p, name='하이브리드', hybrid=True),
        ]

    if compare == 'vol':
        mult = p.get('vol_mult', 1.5)
        return [
            dict(p, name='볼륨필터 OFF', vol_filter=False),
            dict(p, name=f'볼륨필터 ON(x{mult:g})', vol_filter=True),
        ]

    if compare == 'tp':
        # 익절 방식 비교: 지금(ADX 고정) vs 스위칭만(TP없음) vs ATR(배수 4종)
        return [
            dict(p, name='지금(ADX TP)', tp_mode='adx'),
            dict(p, name='스위칭만(TP없음)', tp_mode='switch'),
            dict(p, name='ATR TP x1.0', tp_mode='atr', atr_tp_mult=1.0),
            dict(p, name='ATR TP x1.5', tp_mode='atr', atr_tp_mult=1.5),
            dict(p, name='ATR TP x2.0', tp_mode='atr', atr_tp_mult=2.0),
            dict(p, name='ATR TP x2.5', tp_mode='atr', atr_tp_mult=2.5),
        ]

    return [
        dict(p, name=setval_name, hybrid=False),
        dict(p, name='EMA12-26', ema_fast=12, ema_slow=26, hybrid=False),
        dict(p, name='EMA9-21', ema_fast=9, ema_slow=21, hybrid=False),
        dict(p, name='하이브리드', hybrid=True),
    ]


def monthly_table(trades):
    """거래 내역(trades DF) → 월별 집계 DataFrame

    컬럼: 년-월 | 거래수 | 승 | 패 | 승률% | 수익 | 손실 | 수수료 | 순손익 | 누적순손익
    """
    if trades is None or trades.empty:
        return pd.DataFrame()
    t = trades.copy()
    t['시각'] = pd.to_datetime(t['시각'])
    t['년월'] = t['시각'].dt.strftime('%Y-%m')
    rows = []
    cum = 0.0
    for ym, g in t.groupby('년월', sort=True):
        wins = int((g['순손익'] > 0).sum())
        losses = len(g) - wins
        profit = g.loc[g['순손익'] > 0, '순손익'].sum()
        loss = g.loc[g['순손익'] <= 0, '순손익'].sum()  # 음수
        fee = g['수수료'].sum()
        net = g['순손익'].sum()
        cum += net
        rows.append({
            '년-월': ym, '거래수': len(g), '승': wins, '패': losses,
            '승률%': round(wins / len(g) * 100, 1) if len(g) else 0,
            '수익(USDT)': round(profit, 2), '손실(USDT)': round(loss, 2),
            '수수료(USDT)': round(fee, 2), '순손익(USDT)': round(net, 2),
            '누적순손익': round(cum, 2),
        })
    df = pd.DataFrame(rows)
    # 합계 행
    total = {
        '년-월': '합계', '거래수': int(df['거래수'].sum()),
        '승': int(df['승'].sum()), '패': int(df['패'].sum()),
        '승률%': round(df['승'].sum() / df['거래수'].sum() * 100, 1) if df['거래수'].sum() else 0,
        '수익(USDT)': round(df['수익(USDT)'].sum(), 2),
        '손실(USDT)': round(df['손실(USDT)'].sum(), 2),
        '수수료(USDT)': round(df['수수료(USDT)'].sum(), 2),
        '순손익(USDT)': round(df['순손익(USDT)'].sum(), 2),
        '누적순손익': round(cum, 2),
    }
    return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


def write_monthly_excel(all_trades, path, log=print):
    """(심볼, 설정명, trades) 목록 → 월별 집계 엑셀 생성.

    - 시트 1개당 (심볼 × 설정) 하나, 월별 표
    - 색상: 순손익 양수 초록 / 음수 빨강, 합계행 강조
    - openpyxl 없으면 CSV로 폴백
    """
    tables = [(sym, name, monthly_table(tr)) for sym, name, tr in all_trades]
    tables = [(s, n, t) for s, n, t in tables if not t.empty]
    if not tables:
        return

    if not OPENPYXL_OK:
        # 폴백: 월별 표를 CSV로
        try:
            for sym, name, mt in tables:
                safe = f"{sym}_{name}".replace('/', '-').replace('(', '').replace(')', '')
                mt.to_csv(path.replace('.xlsx', f'_{safe}.csv'),
                          index=False, encoding='utf-8-sig')
            log(f"📅 월별 집계 CSV 저장 (openpyxl 미설치): {os.path.dirname(path)}")
        except Exception as e:
            log(f"  ⚠️ 월별 CSV 저장 실패: {e}")
        return

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.remove(wb.active)
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_fill = PatternFill('solid', fgColor='305496')
        tot_fill = PatternFill('solid', fgColor='FFF2CC')
        pos_font = Font(color='1F7A1F', bold=True)
        neg_font = Font(color='C00000', bold=True)
        center = Alignment(horizontal='center')
        thin = Side(style='thin', color='D9D9D9')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        used_names = set()

        for sym, cfg_name, mt in tables:
            title = f"{sym}_{cfg_name}"[:31].replace('/', '-').replace('*', '')
            base_title = title
            n = 1
            while title in used_names:
                n += 1
                title = f"{base_title[:28]}_{n}"
            used_names.add(title)
            ws = wb.create_sheet(title=title)

            cols = list(mt.columns)
            for c, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=c, value=col)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = center
                cell.border = border

            net_idx = cols.index('순손익(USDT)') + 1
            for r, (_, row) in enumerate(mt.iterrows(), 2):
                is_total = (row['년-월'] == '합계')
                for c, col in enumerate(cols, 1):
                    cell = ws.cell(row=r, column=c, value=row[col])
                    cell.alignment = center
                    cell.border = border
                    if is_total:
                        cell.fill = tot_fill
                        cell.font = Font(bold=True)
                # 순손익 색상
                nv = row['순손익(USDT)']
                ws.cell(row=r, column=net_idx).font = pos_font if nv >= 0 else neg_font

            # 열 너비
            for c, col in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(c)].width = max(10, len(str(col)) + 2)
            ws.freeze_panes = 'A2'

        wb.save(path)
        log(f"📅 월별 집계 엑셀 저장 완료:\n   {path}")
    except Exception as e:
        log(f"  ⚠️ 월별 엑셀 생성 실패({e}) — CSV로 대체 시도")
        try:
            for sym, name, mt in tables:
                safe = f"{sym}_{name}".replace('/', '-').replace('(', '').replace(')', '')
                mt.to_csv(path.replace('.xlsx', f'_{safe}.csv'),
                          index=False, encoding='utf-8-sig')
        except Exception:
            pass


def run_all(symbols, p, compare, log=print, on_row=None):
    """심볼 목록 전체 백테스트. on_row(row)로 결과 행 전달"""
    configs = make_configs(p, compare)
    if p.get('start_date') or p.get('end_date'):
        period = f"{p.get('start_date') or '처음'} ~ {p.get('end_date') or '현재'}"
    else:
        period = f"최근 {p['days']}일"
    log("=" * 50)
    log(f"🔬 백테스트: {len(symbols)}개 심볼 | {period} | 매매 {p['interval']}봉"
        f" | ADX {p.get('adx_interval', p['interval'])}봉")
    log(f"   설정 {len(configs)}종: {', '.join(c['name'] for c in configs)}")
    log("=" * 50)

    base = get_workdir(log)

    def save_csv(df_out, filename):
        """쓰기 가능하면 저장, 아니면 조용히 생략 (경로 반환/None)"""
        if not base:
            return None
        try:
            path = os.path.join(base, filename)
            df_out.to_csv(path, index=False, encoding='utf-8-sig')
            return path
        except Exception as e:
            log(f"  ⚠️ 저장 실패({filename}): {e}")
            return None

    summaries = []
    all_trades = []  # (심볼, 설정명, trades_df) — 월별 엑셀용
    for k, sym in enumerate(symbols, 1):
        log(f"\n[{k}/{len(symbols)}] {sym} 데이터 준비...")
        try:
            df = fetch_klines(sym, p['days'], p['interval'], log=log,
                              start_date=p.get('start_date'),
                              end_date=p.get('end_date'))
        except Exception as e:
            log(f"  ❌ {sym} 건너뜀: {e}")
            continue
        if len(df) < 200:
            log(f"  ⚠️ {sym} 데이터 부족({len(df)}봉) — 건너뜀")
            continue
        for cfg in configs:
            trades, max_dd = run_backtest(df, cfg)
            row = summarize(trades, sym, cfg['name'], max_dd, cfg)
            if row:
                summaries.append(row)
                all_trades.append((sym, cfg['name'], trades))
                if on_row:
                    on_row(row)
                log(f"  ▶ [{cfg['name']}] {row['거래수']}회 | 승률 {row['승률%']}% "
                    f"| 순손익 {row['순손익']:+,.2f} | 수수료 -{row['수수료']:,.2f}")
                safe = cfg['name'].replace('/', '-')
                save_csv(trades, f"backtest_trades_{sym}_{safe}.csv")

    if summaries:
        result = pd.DataFrame(summaries)
        saved = save_csv(result, 'backtest_result.csv')
        log("\n" + "=" * 50)
        if saved:
            log(f"📊 완료! 요약 저장됨:\n   {saved}")
        else:
            log("📊 완료! (파일 저장은 생략 — 아래 결과 참고)")

        # 🏆 설정별 합산 순손익 집계 → 승자 추천
        agg = []
        for name, grp in result.groupby('설정', sort=False):
            net = grp['순손익'].sum()
            fee = grp['수수료'].sum()
            trades = grp['거래수'].sum()
            wr = (grp['승률%'] * grp['거래수']).sum() / trades if trades else 0
            dd = grp['최대낙폭'].min()
            agg.append({'설정': name, '순손익': net, '수수료': fee,
                        '승률': wr, '낙폭': dd})
            log(f"   {name:20s}: 순손익 {net:+10,.2f} | 승률 {wr:4.1f}%"
                f" | 최대낙폭 {dd:+.2f} | 수수료 -{fee:,.2f}")
        log("=" * 50)

        agg.sort(key=lambda x: x['순손익'], reverse=True)
        best = agg[0]
        if best['순손익'] <= 0:
            log(f"\n⚠️ 추천 보류: 1위 [{best['설정']}]도 순손익 {best['순손익']:+,.2f} "
                f"— 이 기간엔 모든 설정이 손실입니다.")
        else:
            log(f"\n🏆 추천: [{best['설정']}]  순손익 {best['순손익']:+,.2f} USDT"
                f" (승률 {best['승률']:.1f}%, 최대낙폭 {best['낙폭']:+.2f})")
            if len(agg) > 1:
                second = agg[1]
                gap = best['순손익'] - second['순손익']
                denom = abs(second['순손익']) or 1
                log(f"   2위 [{second['설정']}] 대비 {gap:+,.2f} USDT"
                    f" ({gap / denom * 100:+.1f}%) 우위")
                if gap / denom < 0.1:
                    log("   ⚠️ 1·2위 차이가 10% 미만 — 우열 크지 않음(우연일 수 있음). "
                        "여러 코인·기간으로 재확인 권장.")

        # 📅 월별 수익/수수료/손실 엑셀 자동 생성
        if base:
            xlsx_path = os.path.join(base, 'backtest_월별집계.xlsx')
            write_monthly_excel(all_trades, xlsx_path, log)
    else:
        log("\n⚠️ 결과 없음")
    return summaries


# ==================== GUI ====================
def launch_gui():
    import tkinter as tk
    from tkinter import ttk, messagebox

    BG, PANEL, FG, ACCENT = '#1e1e1e', '#2d2d2d', '#ffffff', '#00ff88'

    root = tk.Tk()
    root.title("📊 바이낸스 선물 백테스터 (메인넷 실제 데이터)")
    root.geometry("1150x860")
    root.configure(bg=BG)

    tk.Label(root, text="🌐 데이터 출처: 바이낸스 선물 메인넷 실제 시세 (fapi.binance.com)",
             bg='#0a3d2e', fg='#00ff88', font=('Arial', 9, 'bold')).pack(fill='x')

    log_q = queue.Queue()
    running = [False]
    all_symbols = list(FALLBACK_SYMBOLS)

    # ---------- 좌측: 설정 (스크롤 가능) ----------
    left_wrap = tk.Frame(root, bg=PANEL)
    left_wrap.pack(side='left', fill='y', padx=8, pady=8)
    left_canvas = tk.Canvas(left_wrap, bg=PANEL, highlightthickness=0, width=310)
    left_scroll = tk.Scrollbar(left_wrap, orient='vertical',
                               command=left_canvas.yview)
    left_canvas.configure(yscrollcommand=left_scroll.set)
    left_scroll.pack(side='right', fill='y')
    left_canvas.pack(side='left', fill='both', expand=True)
    left = tk.Frame(left_canvas, bg=PANEL)   # 실제 설정 위젯들이 담기는 내부 프레임
    left_canvas.create_window((0, 0), window=left, anchor='nw')

    def _left_configure(_e=None):
        left_canvas.configure(scrollregion=left_canvas.bbox('all'))
    left.bind('<Configure>', _left_configure)

    def _left_wheel(e):
        left_canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')

    def _left_wheel_up(_e):
        left_canvas.yview_scroll(-1, 'units')

    def _left_wheel_down(_e):
        left_canvas.yview_scroll(1, 'units')

    def _left_enter(_e):   # 마우스가 왼쪽 패널 위에 있을 때만 휠 작동
        left_canvas.bind_all('<MouseWheel>', _left_wheel)      # Windows/Mac
        left_canvas.bind_all('<Button-4>', _left_wheel_up)     # Linux
        left_canvas.bind_all('<Button-5>', _left_wheel_down)

    def _left_leave(_e):
        left_canvas.unbind_all('<MouseWheel>')
        left_canvas.unbind_all('<Button-4>')
        left_canvas.unbind_all('<Button-5>')
    left_canvas.bind('<Enter>', _left_enter)
    left_canvas.bind('<Leave>', _left_leave)

    tk.Label(left, text="⚙️ 설정 (모두 수정 가능)", bg=PANEL, fg=ACCENT,
             font=('Arial', 12, 'bold')).pack(pady=(8, 4))

    form = tk.Frame(left, bg=PANEL)
    form.pack(padx=10)

    fields = [
        ('진입금 (USDT)', 'amount'), ('레버리지 (배)', 'leverage'),
        ('수수료 % (편도)', 'fee_pct'), ('추세장 TP %', 'tp_trend'),
        ('횡보장 TP %', 'tp_sideways'), ('ADX 기간', 'adx_period'),
        ('ADX 추세 기준', 'adx_th'), ('UT Key Value', 'ut_sens'),
        ('UT ATR 기간', 'ut_atr'), ('EMA Fast', 'ema_fast'),
        ('EMA Slow', 'ema_slow'),
    ]
    vars_ = {}
    for r, (label, key) in enumerate(fields):
        tk.Label(form, text=label, bg=PANEL, fg=FG, font=('Arial', 10),
                 anchor='w').grid(row=r, column=0, sticky='w', pady=2)
        v = tk.StringVar(value=str(DEFAULTS[key]))
        tk.Entry(form, textvariable=v, width=10, font=('Arial', 10),
                 justify='center').grid(row=r, column=1, padx=6, pady=2)
        vars_[key] = v

    tk.Label(form, text='시간봉 (매매)', bg=PANEL, fg=FG, font=('Arial', 10),
             anchor='w').grid(row=len(fields), column=0, sticky='w', pady=2)
    interval_var = tk.StringVar(value=DEFAULTS['interval'])
    ttk.Combobox(form, textvariable=interval_var, values=INTERVALS, width=8,
                 state='readonly').grid(row=len(fields), column=1, padx=6, pady=2)

    tk.Label(form, text='ADX 시간봉', bg=PANEL, fg=FG, font=('Arial', 10),
             anchor='w').grid(row=len(fields) + 1, column=0, sticky='w', pady=2)
    adx_interval_var = tk.StringVar(value=DEFAULTS['adx_interval'])
    ttk.Combobox(form, textvariable=adx_interval_var, values=INTERVALS, width=8,
                 state='readonly').grid(row=len(fields) + 1, column=1, padx=6, pady=2)

    # ---------- 기간 지정 ----------
    tk.Label(left, text="─" * 30, bg=PANEL, fg='#555555').pack()
    tk.Label(left, text="📅 백테스트 기간", bg=PANEL, fg='#00ffff',
             font=('Arial', 10, 'bold')).pack(anchor='w', padx=10)

    days_var = tk.StringVar(value=str(DEFAULTS['days']))
    start_var = tk.StringVar(value='')
    end_var = tk.StringVar(value='')

    # 빠른 버튼 (최근 N일 → days 설정 + 날짜칸 비움)
    quick = tk.Frame(left, bg=PANEL)
    quick.pack(anchor='w', padx=22, pady=2)

    def set_days(d):
        days_var.set(str(d))
        start_var.set('')
        end_var.set('')

    for label, d in [('1개월', 30), ('3개월', 90), ('6개월', 180),
                     ('1년', 365), ('2년', 730)]:
        tk.Button(quick, text=label, command=lambda d=d: set_days(d),
                  bg='#3d3d3d', fg=FG, font=('Arial', 8), width=5,
                  padx=1, pady=1).pack(side='left', padx=1)

    drow = tk.Frame(left, bg=PANEL)
    drow.pack(anchor='w', padx=22, pady=2)
    tk.Label(drow, text='최근', bg=PANEL, fg=FG, font=('Arial', 9)).pack(side='left')
    tk.Entry(drow, textvariable=days_var, width=6, font=('Arial', 9),
             justify='center').pack(side='left', padx=3)
    tk.Label(drow, text='일', bg=PANEL, fg=FG, font=('Arial', 9)).pack(side='left')

    tk.Label(left, text="또는 날짜 직접 지정 (비우면 위 '최근 N일' 사용)",
             bg=PANEL, fg='#888888', font=('Arial', 8)).pack(anchor='w', padx=22)
    drow2 = tk.Frame(left, bg=PANEL)
    drow2.pack(anchor='w', padx=22, pady=2)
    tk.Label(drow2, text='시작', bg=PANEL, fg=FG, font=('Arial', 9)).pack(side='left')
    tk.Entry(drow2, textvariable=start_var, width=11, font=('Arial', 9),
             justify='center').pack(side='left', padx=3)
    tk.Label(drow2, text='끝', bg=PANEL, fg=FG, font=('Arial', 9)).pack(side='left')
    tk.Entry(drow2, textvariable=end_var, width=11, font=('Arial', 9),
             justify='center').pack(side='left', padx=3)
    tk.Label(left, text="예: 2024-01-01   (끝 비우면 오늘까지)",
             bg=PANEL, fg='#888888', font=('Arial', 8)).pack(anchor='w', padx=22)

    # 하이브리드
    tk.Label(left, text="─" * 30, bg=PANEL, fg='#555555').pack()
    hybrid_var = tk.BooleanVar(value=DEFAULTS['hybrid'])
    tk.Checkbutton(left, text="하이브리드 조기청산 사용", variable=hybrid_var,
                   bg=PANEL, fg='#ffaa00', selectcolor=BG, font=('Arial', 10, 'bold'),
                   activebackground=PANEL).pack(anchor='w', padx=10)
    hform = tk.Frame(left, bg=PANEL)
    hform.pack(padx=10)
    hfields = [('빠른 EMA Fast', 'hybrid_fast'), ('빠른 EMA Slow', 'hybrid_slow'),
               ('발동 ROI % 이하', 'hybrid_roi_th')]
    for r, (label, key) in enumerate(hfields):
        tk.Label(hform, text=label, bg=PANEL, fg='#aaaaaa', font=('Arial', 9),
                 anchor='w').grid(row=r, column=0, sticky='w', pady=1)
        v = tk.StringVar(value=str(DEFAULTS[key]))
        tk.Entry(hform, textvariable=v, width=10, font=('Arial', 9),
                 justify='center').grid(row=r, column=1, padx=6, pady=1)
        vars_[key] = v

    # 거래량 필터
    tk.Label(left, text="─" * 30, bg=PANEL, fg='#555555').pack()
    vol_filter_var = tk.BooleanVar(value=DEFAULTS['vol_filter'])
    tk.Checkbutton(left, text="거래량(볼륨) 필터 사용", variable=vol_filter_var,
                   bg=PANEL, fg='#00ffff', selectcolor=BG, font=('Arial', 10, 'bold'),
                   activebackground=PANEL).pack(anchor='w', padx=10)
    vform = tk.Frame(left, bg=PANEL)
    vform.pack(padx=10)
    vfields = [('거래량 배수 (×평균)', 'vol_mult'), ('평균 기간 (봉)', 'vol_ma')]
    for r, (label, key) in enumerate(vfields):
        tk.Label(vform, text=label, bg=PANEL, fg='#aaaaaa', font=('Arial', 9),
                 anchor='w').grid(row=r, column=0, sticky='w', pady=1)
        v = tk.StringVar(value=str(DEFAULTS[key]))
        tk.Entry(vform, textvariable=v, width=10, font=('Arial', 9),
                 justify='center').grid(row=r, column=1, padx=6, pady=1)
        vars_[key] = v
    tk.Label(left, text="(현재봉 거래량 ≥ 평균 × 배수일 때만 진입)", bg=PANEL,
             fg='#888888', font=('Arial', 8)).pack(anchor='w', padx=22)

    # 익절(TP) 방식
    tk.Label(left, text="─" * 30, bg=PANEL, fg='#555555').pack()
    tk.Label(left, text="🎯 익절(TP) 방식", bg=PANEL, fg='#00ffff',
             font=('Arial', 10, 'bold')).pack(anchor='w', padx=10)
    TP_MODE_OPTS = {
        '지금 방식 (ADX 고정 1.2/1.0%)': 'adx',
        'ATR (변동성 기반)': 'atr',
        '스위칭만 (TP 없음)': 'switch',
    }
    tp_mode_label_var = tk.StringVar(value='지금 방식 (ADX 고정 1.2/1.0%)')
    ttk.Combobox(left, textvariable=tp_mode_label_var,
                 values=list(TP_MODE_OPTS.keys()), width=26,
                 state='readonly').pack(anchor='w', padx=22, pady=2)

    aform = tk.Frame(left, bg=PANEL)
    aform.pack(padx=10)
    afields = [('ATR 기간', 'atr_tp_period'), ('ATR 배수 (×ATR)', 'atr_tp_mult')]
    for r, (label, key) in enumerate(afields):
        tk.Label(aform, text=label, bg=PANEL, fg='#aaaaaa', font=('Arial', 9),
                 anchor='w').grid(row=r, column=0, sticky='w', pady=1)
        v = tk.StringVar(value=str(DEFAULTS[key]))
        tk.Entry(aform, textvariable=v, width=10, font=('Arial', 9),
                 justify='center').grid(row=r, column=1, padx=6, pady=1)
        vars_[key] = v
    # 예시 가이드
    tk.Label(left, text="📌 ATR 배수 예시 (TP거리 = ATR × 배수):", bg=PANEL,
             fg='#ffaa00', font=('Arial', 8, 'bold')).pack(anchor='w', padx=22)
    for ex in ["  1.0 = 짧게·자주 익절 (박리다매)",
               "  1.5 = 중간 (기본값·추천 시작점)",
               "  2.0~2.5 = 길게·크게 (큰 추세 노림)"]:
        tk.Label(left, text=ex, bg=PANEL, fg='#888888',
                 font=('Arial', 8)).pack(anchor='w', padx=22)

    tk.Label(left, text="─" * 30, bg=PANEL, fg='#555555').pack()
    tk.Label(left, text="🔬 비교 모드", bg=PANEL, fg='#00ffff',
             font=('Arial', 10, 'bold')).pack(anchor='w', padx=10)
    COMPARE_OPTS = {
        '설정값 vs 하이브리드': 'vs',
        '익절방식 비교(ADX/스위칭/ATR)': 'tp',
        '볼륨필터 OFF vs ON': 'vol',
        'EMA 4종 비교': 'full',
        '단일 (현재 설정만)': 'off',
    }
    compare_label_var = tk.StringVar(value='설정값 vs 하이브리드')
    ttk.Combobox(left, textvariable=compare_label_var,
                 values=list(COMPARE_OPTS.keys()), width=20,
                 state='readonly').pack(anchor='w', padx=26, pady=2)
    tk.Label(left, text="→ 끝나면 수익 1위를 자동 추천", bg=PANEL,
             fg='#888888', font=('Arial', 8)).pack(anchor='w', padx=26)

    run_btn = tk.Button(left, text="▶️ 백테스트 시작", bg='#0066cc', fg='#ffffff',
                        font=('Arial', 13, 'bold'), padx=20, pady=8)
    run_btn.pack(pady=12)

    # ---------- 중앙: 코인 선택 ----------
    mid = tk.Frame(root, bg=PANEL)
    mid.pack(side='left', fill='y', padx=(0, 8), pady=8)

    tk.Label(mid, text="🪙 코인 선택", bg=PANEL, fg=ACCENT,
             font=('Arial', 12, 'bold')).pack(pady=(8, 4))

    search_var = tk.StringVar()
    tk.Entry(mid, textvariable=search_var, width=18, font=('Arial', 10)).pack(padx=10)
    tk.Label(mid, text="(검색: btc, doge ...)", bg=PANEL, fg='#888888',
             font=('Arial', 8)).pack()

    lb_frame = tk.Frame(mid, bg=PANEL)
    lb_frame.pack(fill='both', expand=True, padx=10, pady=4)
    sb = tk.Scrollbar(lb_frame)
    sb.pack(side='right', fill='y')
    listbox = tk.Listbox(lb_frame, selectmode='extended', width=18, height=25,
                         bg=BG, fg=FG, font=('Consolas', 10),
                         yscrollcommand=sb.set, exportselection=False)
    listbox.pack(side='left', fill='both', expand=True)
    sb.config(command=listbox.yview)

    count_label = tk.Label(mid, text="선택: 0개", bg=PANEL, fg='#ffaa00',
                           font=('Arial', 10, 'bold'))
    count_label.pack()

    btns = tk.Frame(mid, bg=PANEL)
    btns.pack(pady=6)

    def refresh_list(*_):
        kw = search_var.get().strip().upper()
        listbox.delete(0, 'end')
        for s in all_symbols:
            if kw in s:
                listbox.insert('end', s)
    search_var.trace('w', refresh_list)

    def update_count(*_):
        count_label.config(text=f"선택: {len(listbox.curselection())}개")
    listbox.bind('<<ListboxSelect>>', update_count)

    def select_all():
        listbox.select_set(0, 'end')
        update_count()

    def select_none():
        listbox.select_clear(0, 'end')
        update_count()

    def select_default():
        select_none()
        for i in range(listbox.size()):
            if listbox.get(i) in DEFAULT_SYMBOLS:
                listbox.select_set(i)
        update_count()

    tk.Button(btns, text="전체", command=select_all, bg='#3d3d3d', fg=FG,
              font=('Arial', 9), width=5).pack(side='left', padx=2)
    tk.Button(btns, text="해제", command=select_none, bg='#3d3d3d', fg=FG,
              font=('Arial', 9), width=5).pack(side='left', padx=2)
    tk.Button(btns, text="기본3종", command=select_default, bg='#3d3d3d', fg=FG,
              font=('Arial', 9), width=7).pack(side='left', padx=2)

    # ---------- 우측: 결과 + 로그 ----------
    right = tk.Frame(root, bg=BG)
    right.pack(side='left', fill='both', expand=True, pady=8, padx=(0, 8))

    tk.Label(right, text="📊 결과 (CSV 자동 저장 — 로그에 저장 위치 표시)", bg=BG, fg=ACCENT,
             font=('Arial', 12, 'bold')).pack(anchor='w')

    cols = ['심볼', '설정', '거래수', '승률%', 'TP익절', '스위칭', '조기청산',
            '강제청산', '총수익', '수수료', '순손익', '최대낙폭', '최저ROI%']
    style = ttk.Style()
    try:
        style.theme_use('clam')
        style.configure('Treeview', background=PANEL, fieldbackground=PANEL,
                        foreground=FG, rowheight=22)
        style.configure('Treeview.Heading', background='#3d3d3d', foreground=FG)
    except Exception:
        pass

    tree_frame = tk.Frame(right, bg=BG)
    tree_frame.pack(fill='both', expand=True)
    tsb = tk.Scrollbar(tree_frame)
    tsb.pack(side='right', fill='y')
    tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=14,
                        yscrollcommand=tsb.set)
    tsb.config(command=tree.yview)
    for col in cols:
        tree.heading(col, text=col)
        tree.column(col, width=78 if col in ('심볼', '설정') else 62,
                    anchor='center')
    tree.column('설정', width=120)
    tree.pack(side='left', fill='both', expand=True)
    tree.tag_configure('win', foreground='#00ff88')
    tree.tag_configure('lose', foreground='#ff6666')

    tk.Label(right, text="📋 진행 로그", bg=BG, fg=ACCENT,
             font=('Arial', 11, 'bold')).pack(anchor='w', pady=(8, 0))
    log_frame = tk.Frame(right, bg=BG)
    log_frame.pack(fill='both', expand=True)
    lsb = tk.Scrollbar(log_frame)
    lsb.pack(side='right', fill='y')
    log_text = tk.Text(log_frame, height=12, bg='#141414', fg='#cccccc',
                       font=('Consolas', 9), yscrollcommand=lsb.set, wrap='word')
    log_text.pack(side='left', fill='both', expand=True)
    lsb.config(command=log_text.yview)

    def gui_log(msg):
        log_q.put(('log', str(msg)))

    def poll_queue():
        try:
            while True:
                kind, payload = log_q.get_nowait()
                if kind == 'log':
                    log_text.insert('end', payload + '\n')
                    log_text.see('end')
                elif kind == 'row':
                    row = payload
                    tag = 'win' if row['순손익'] >= 0 else 'lose'
                    tree.insert('', 'end', values=[row.get(col, '') for col in cols],
                                tags=(tag,))
                elif kind == 'done':
                    running[0] = False
                    run_btn.config(state='normal', text="▶️ 백테스트 시작")
        except queue.Empty:
            pass
        root.after(100, poll_queue)

    # ---------- 실행 ----------
    def read_params():
        """빈칸/공백/쉼표 등 관대하게 처리 — 이상하면 기본값으로 대체(안 죽음)"""
        p = dict(DEFAULTS)
        int_keys = {'leverage', 'adx_period', 'ut_atr', 'ema_fast', 'ema_slow',
                    'days', 'hybrid_fast', 'hybrid_slow', 'vol_ma', 'atr_tp_period'}
        fixed = []
        for key, v in vars_.items():
            raw = (v.get() or '').strip().replace(',', '')
            if raw == '':
                p[key] = DEFAULTS[key]
                v.set(str(DEFAULTS[key]))
                continue
            try:
                p[key] = int(round(float(raw))) if key in int_keys else float(raw)
            except ValueError:
                p[key] = DEFAULTS[key]
                v.set(str(DEFAULTS[key]))
                fixed.append(key)
        # 기간: 날짜(시작/끝) 우선, 없으면 최근 N일
        sd = (start_var.get() or '').strip()
        ed = (end_var.get() or '').strip()
        p['start_date'] = sd or None
        p['end_date'] = ed or None
        if not sd and not ed:
            try:
                p['days'] = max(1, int(round(float((days_var.get() or '').strip().replace(',', '')))))
            except ValueError:
                p['days'] = DEFAULTS['days']
                days_var.set(str(DEFAULTS['days']))
                fixed.append('기간(일)')
        # 상식 보정 (유연하게 — 막지 않고 조용히 정리)
        p['leverage'] = max(1, min(125, p['leverage']))
        if p['ema_fast'] >= p['ema_slow']:
            fixed.append('ema_fast/slow')
        p['interval'] = interval_var.get()
        p['adx_interval'] = adx_interval_var.get()
        p['hybrid'] = hybrid_var.get()
        p['vol_filter'] = vol_filter_var.get()
        p['vol_ma'] = max(1, int(p['vol_ma']))
        p['tp_mode'] = TP_MODE_OPTS.get(tp_mode_label_var.get(), 'adx')
        p['atr_tp_period'] = max(1, int(p['atr_tp_period']))
        if fixed:
            gui_log(f"ℹ️ 잘못된 입력 자동 보정: {', '.join(fixed)}")
        return p

    def start():
        if running[0]:
            return
        try:
            p = read_params()
        except ValueError as e:
            messagebox.showwarning("입력 오류", f"숫자를 확인해주세요!\n\n{e}")
            return
        sel = [listbox.get(i) for i in listbox.curselection()]
        if not sel:
            messagebox.showwarning("코인 선택", "코인을 1개 이상 선택해주세요!")
            return
        if len(sel) > 20:
            est = len(sel) * max(1, p['days'] // 5)
            if not messagebox.askyesno(
                    "확인", f"{len(sel)}개 코인 × {p['days']}일 = 다운로드에 "
                    f"시간이 오래 걸릴 수 있습니다 (약 {est}회 API 요청).\n계속할까요?"):
                return
        for item in tree.get_children():
            tree.delete(item)
        running[0] = True
        run_btn.config(state='disabled', text="⏳ 실행 중...")

        compare_mode = COMPARE_OPTS.get(compare_label_var.get(), 'vs')

        def work():
            try:
                run_all(sel, p, compare_mode, log=gui_log,
                        on_row=lambda row: log_q.put(('row', row)))
            except Exception as e:
                gui_log(f"❌ 오류: {e}")
            finally:
                log_q.put(('done', None))

        threading.Thread(target=work, daemon=True).start()

    run_btn.config(command=start)

    # 심볼 목록 백그라운드 로드
    def load_symbols():
        nonlocal all_symbols
        syms = fetch_symbols(log=gui_log)
        all_symbols = syms
        root.after(0, lambda: (refresh_list(), select_default()))

    refresh_list()
    select_default()
    threading.Thread(target=load_symbols, daemon=True).start()

    gui_log("✅ 준비 완료! 설정 확인 후 [▶️ 백테스트 시작]을 누르세요.")
    gui_log("   기본값 = 현재 봇 설정 (50 USDT × 3배, TP 1.2/1.0%, EMA 34/55)")
    poll_queue()
    root.mainloop()


# ==================== CLI ====================
def run_cli():
    ap = argparse.ArgumentParser(description='바이낸스 선물 백테스터 (CLI 모드)')
    ap.add_argument('--symbols', default=','.join(DEFAULT_SYMBOLS))
    ap.add_argument('--days', type=int, default=DEFAULTS['days'])
    ap.add_argument('--interval', default=DEFAULTS['interval'])
    ap.add_argument('--adx-interval', default=DEFAULTS['adx_interval'],
                    help='ADX 계산 시간봉 (기본 1h)')
    ap.add_argument('--ema', default=f"{DEFAULTS['ema_fast']},{DEFAULTS['ema_slow']}")
    ap.add_argument('--compare', default='vs',
                    choices=['off', 'vs', 'full', 'vol', 'tp'],
                    help="off=단일 / vs=설정값vs하이브리드(기본) / full=EMA4종 / "
                         "vol=볼륨OFFvsON / tp=익절방식(ADX/스위칭/ATR)")
    ap.add_argument('--start', default=None, help='시작일 YYYY-MM-DD (지정 시 --days 무시)')
    ap.add_argument('--end', default=None, help='종료일 YYYY-MM-DD (기본 오늘)')
    ap.add_argument('--vol', action='store_true', help='거래량 필터 ON')
    ap.add_argument('--vol-mult', type=float, default=DEFAULTS['vol_mult'],
                    help='거래량 배수 (기본 1.5)')
    ap.add_argument('--tp-mode', default=DEFAULTS['tp_mode'],
                    choices=['adx', 'atr', 'switch'],
                    help='익절 방식: adx(기본)/atr/switch')
    ap.add_argument('--atr-mult', type=float, default=DEFAULTS['atr_tp_mult'],
                    help='ATR TP 배수 (기본 1.5)')
    ap.add_argument('--all', action='store_true', help='바이낸스 전체 코인')
    args = ap.parse_args()

    p = dict(DEFAULTS)
    p['days'] = args.days
    p['start_date'] = args.start
    p['end_date'] = args.end
    p['interval'] = args.interval
    p['adx_interval'] = args.adx_interval
    p['vol_filter'] = args.vol
    p['vol_mult'] = args.vol_mult
    p['tp_mode'] = args.tp_mode
    p['atr_tp_mult'] = args.atr_mult
    p['ema_fast'], p['ema_slow'] = [int(x) for x in args.ema.split(',')]

    symbols = fetch_symbols() if args.all else \
        [s.strip().upper() for s in args.symbols.split(',')]
    run_all(symbols, p, args.compare)


def _pause_exit(code=0):
    """더블클릭 실행 시 창이 그냥 닫히지 않도록 대기 (원인 확인용)"""
    try:
        input("\n[엔터]를 누르면 종료합니다...")
    except Exception:
        pass
    sys.exit(code)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        # CLI 모드
        try:
            run_cli()
        except KeyboardInterrupt:
            print("\n중단됨.")
        except Exception as e:
            import traceback
            print("\n❌ 오류 발생:")
            traceback.print_exc()
            print(f"\n요약: {e}")
    else:
        # GUI 모드 — 어떤 오류에도 창이 조용히 닫히지 않게
        try:
            launch_gui()
        except ImportError:
            print("=" * 55)
            print("⚠️ 이 파이썬에 tkinter(GUI)가 없습니다.")
            print("   Windows: python.org 정식 설치본을 쓰면 기본 포함")
            print("   Linux:   sudo apt install python3-tk")
            print("=" * 55)
            print("→ 우선 CLI 모드로 실행합니다 (BTC/XRP/DOGE, 설정값 vs 하이브리드)\n")
            try:
                sys.argv = [sys.argv[0], '--compare', 'vs']
                run_cli()
            except Exception as e:
                import traceback
                traceback.print_exc()
                print(f"\n요약: {e}")
            _pause_exit(1)
        except KeyboardInterrupt:
            print("\n중단됨.")
        except Exception as e:
            import traceback
            print("\n❌ 프로그램 오류 (아래 내용을 캡처해 문의하세요):")
            traceback.print_exc()
            print(f"\n요약: {e}")
            _pause_exit(1)
