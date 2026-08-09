#!/usr/bin/env python3
"""
GUI 통합 모듈

이 파일은 CustomMessageBox와 App 클래스를 포함합니다.
"""

import tkinter as tk
from tkinter import ttk, simpledialog
import threading
import time
import os
import json
from datetime import datetime
import requests
import pandas as pd
import numpy as np
from urllib.parse import urlencode
import hashlib
import hmac
import webbrowser  # 🔥 트레이딩뷰 차트 열기용

# 엑셀 저장용 (openpyxl)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl 미설치 - pip install openpyxl 실행하세요")

# 엑셀 파일 경로
TRADE_HISTORY_FILE = 'trade_history.xlsx'

# 번호 붙은 모듈 import
import importlib

SETTINGS_FILE = 'settings.json'

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            pass
    try:
        _config = importlib.import_module('1_config')
        return {'api_key': _config.API_KEY, 'api_secret': _config.API_SECRET, 'testnet': _config.TESTNET, 'fee_rate': _config.FEE_RATE}
    except:
        return {'api_key': '', 'api_secret': '', 'testnet': True, 'fee_rate': 0.0004}

def save_settings(s):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"설정 저장 실패: {e}")

_settings = load_settings()
API_KEY = _settings.get('api_key', '')
API_SECRET = _settings.get('api_secret', '')
TESTNET = _settings.get('testnet', True)
FEE_RATE = _settings.get('fee_rate', 0.0004)

_api = importlib.import_module('2_api')
BinanceAPI = _api.BinanceAPI

# 🔥 Indicators 강제 reload (캐시 무시)
_indicators = importlib.import_module('3_indicators')
importlib.reload(_indicators)  # 강제 재로드!
Indicators = _indicators.Indicators

_bot = importlib.import_module('4_bot')
TradingBot = _bot.TradingBot

# 🔥 Volume 함수 직접 정의 (import 캐싱 문제 우회)
def get_volume_info_local(df, ma_period=60):
    """
    거래량 정보 반환 (GUI 표시용) - 로컬 버전
    """
    try:
        # Volume 컬럼 확인
        if 'volume' not in df.columns:
            return {'current': 0, 'average': 0, 'ratio': 0}
        
        volume_ma = df['volume'].rolling(ma_period).mean()
        current_volume = df['volume'].iloc[-1]
        avg_volume = volume_ma.iloc[-1]
        
        if pd.isna(current_volume) or pd.isna(avg_volume) or avg_volume == 0:
            return {'current': 0, 'average': 0, 'ratio': 0}
        
        ratio = current_volume / avg_volume
        
        return {
            'current': float(current_volume),
            'average': float(avg_volume),
            'ratio': float(ratio)
        }
    except Exception as e:
        print(f"[Volume 로컬] 오류: {e}")
        return {'current': 0, 'average': 0, 'ratio': 0}

class CustomMessageBox:
    @staticmethod
    def showinfo(title, message):
        top = tk.Toplevel()
        top.title(title)
        top.geometry("600x300")
        top.configure(bg='#2d2d2d')
        tk.Label(top, text=message, bg='#2d2d2d', fg='#ffffff', 
                font=('Arial', 12), wraplength=550, justify='left').pack(expand=True, padx=20, pady=20)
        tk.Button(top, text="확인", command=top.destroy, bg='#00aa00', fg='#ffffff',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(pady=20)
        top.transient()
        top.grab_set()
        top.wait_window()
    
    @staticmethod
    def showwarning(title, message):
        top = tk.Toplevel()
        top.title(title)
        top.geometry("600x300")
        top.configure(bg='#2d2d2d')
        tk.Label(top, text=message, bg='#2d2d2d', fg='#ffaa00',
                font=('Arial', 12), wraplength=550, justify='left').pack(expand=True, padx=20, pady=20)
        tk.Button(top, text="확인", command=top.destroy, bg='#ffaa00', fg='#000000',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(pady=20)
        top.transient()
        top.grab_set()
        top.wait_window()
    
    @staticmethod
    def askyesno(title, message):
        top = tk.Toplevel()
        top.title(title)
        top.geometry("600x300")
        top.configure(bg='#2d2d2d')
        result = [False]
        
        tk.Label(top, text=message, bg='#2d2d2d', fg='#ffffff',
                font=('Arial', 12), wraplength=550, justify='left').pack(expand=True, padx=20, pady=20)
        
        btn_frame = tk.Frame(top, bg='#2d2d2d')
        btn_frame.pack(pady=20)
        
        def yes():
            result[0] = True
            top.destroy()
        
        tk.Button(btn_frame, text="예", command=yes, bg='#00aa00', fg='#ffffff',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(side='left', padx=10)
        tk.Button(btn_frame, text="아니오", command=top.destroy, bg='#aa0000', fg='#ffffff',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(side='left', padx=10)
        
        top.transient()
        top.grab_set()
        top.wait_window()
        return result[0]

# ==================== 바이낸스 API ====================
class BinanceAPI:
    def __init__(self, api_key, api_secret, testnet=True):
        self.api_key = api_key
        self.api_secret = api_secret
        self.testnet = testnet
        
        # 주문/포지션용 URL (테스트넷/메인넷)
        self.base_url = "https://testnet.binancefuture.com" if testnet else "https://fapi.binance.com"
        
        # 차트 데이터용 URL (항상 메인넷 - 실제 시장 가격)
        self.chart_url = "https://fapi.binance.com"
        
        self.session = requests.Session()
        self.session.headers.update({'X-MBX-APIKEY': self.api_key})
        
        # 🔥 positionRisk 캐시 (LONG/SHORT 봇 공유, API 호출 감소)
        self._position_cache = None
        self._position_cache_time = 0
        self._position_cache_ttl = 3  # 3초간 캐시
        self._position_cache_lock = threading.Lock()
        
        # 🔥 klines 캐시 (같은 코인 LONG/SHORT 봇 공유)
        self._klines_cache = {}  # {symbol_interval: (data, timestamp)}
        self._klines_cache_ttl = 5  # 5초간 캐시
        self._klines_cache_lock = threading.Lock()
        
        # 시간 오프셋 초기화
        self.time_offset = 0
        self._sync_time()
    
    def _sync_time(self):
        """바이낸스 서버 시간과 동기화"""
        try:
            response = self.session.get(f"{self.base_url}/fapi/v1/time")
            server_time = response.json()['serverTime']
            local_time = int(time.time() * 1000)
            self.time_offset = server_time - local_time
            print(f"⏰ 시간 동기화 완료: 오프셋 {self.time_offset}ms")
        except Exception as e:
            print(f"⚠️ 시간 동기화 실패: {e}")
            self.time_offset = 0
    
    def _sign(self, params):
        query_string = urlencode(params)
        signature = hmac.new(self.api_secret.encode('utf-8'), query_string.encode('utf-8'), hashlib.sha256).hexdigest()
        params['signature'] = signature
        return params
    
    def _request(self, method, endpoint, params=None, signed=False):
        url = f"{self.base_url}{endpoint}"
        if params is None:
            params = {}
        if signed:
            params['timestamp'] = int(time.time() * 1000) + self.time_offset
            params = self._sign(params)
        try:
            if method == 'GET':
                response = self.session.get(url, params=params, timeout=10)
            elif method == 'POST':
                response = self.session.post(url, params=params, timeout=10)
            elif method == 'DELETE':
                response = self.session.delete(url, params=params, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            if e.response is not None:
                try:
                    error_data = e.response.json()
                    ec = error_data.get('code', 0)
                    if ec == -1021:
                        print(f"⚠️ 시간 오류! 재동기화...")
                        self._sync_time()
                    elif ec == -1003:
                        print(f"🚫 IP 밴! 30초 대기...")
                        time.sleep(30)
                        self.session = requests.Session()
                        self.session.headers.update({'X-MBX-APIKEY': self.api_key})
                        self._sync_time()
                    elif ec == -1015:
                        print(f"⚠️ 요청 제한! 3초 대기...")
                        time.sleep(3)
                    else:
                        print(f"API 오류: {error_data}")
                except:
                    print(f"API 오류: {e.response.text}")
            return None
        except requests.exceptions.ConnectionError:
            print(f"🔌 연결 끊김! 2초 후 재시도...")
            time.sleep(2)
            self.session = requests.Session()
            self.session.headers.update({'X-MBX-APIKEY': self.api_key})
            return None
        except requests.exceptions.Timeout:
            print(f"⏰ 타임아웃! 다음 루프 재시도...")
            return None
        except Exception as e:
            print(f"API 오류: {e}")
            return None
    
    def get_balance(self):
        account = self._request('GET', '/fapi/v2/account', signed=True)
        if account:
            return float(account.get('totalWalletBalance', 0))
        return 0.0
    
    def get_klines(self, symbol, interval, limit=200):
        """차트 데이터는 항상 메인넷에서 가져옴 (정확한 가격)"""
        cache_key = f"{symbol}_{interval}"
        
        # 🔥 캐시 확인 (같은 코인의 LONG/SHORT 봇이 공유)
        with self._klines_cache_lock:
            now = time.time()
            if cache_key in self._klines_cache:
                cached_data, cached_time = self._klines_cache[cache_key]
                if (now - cached_time) < self._klines_cache_ttl:
                    return cached_data.copy()
        
        params = {'symbol': symbol.replace('/', ''), 'interval': interval, 'limit': limit}
        url = f"{self.chart_url}/fapi/v1/klines"
        
        try:
            response = self.session.get(url, params=params, timeout=10)
            if response.status_code == 200:
                klines = response.json()
                df = pd.DataFrame(klines, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume',
                                                   'close_time', 'quote_volume', 'trades', 'taker_buy_base',
                                                   'taker_buy_quote', 'ignore'])
                df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
                df = df[['timestamp', 'open', 'high', 'low', 'close', 'volume']]
                for col in ['open', 'high', 'low', 'close', 'volume']:
                    df[col] = df[col].astype(float)
                df.set_index('timestamp', inplace=True)
                
                # 🔥 캐시 저장
                with self._klines_cache_lock:
                    self._klines_cache[cache_key] = (df, time.time())
                
                return df
        except Exception as e:
            print(f"차트 데이터 가져오기 오류: {e}")
        
        return None
    
    def set_leverage(self, symbol, leverage):
        params = {'symbol': symbol.replace('/', ''), 'leverage': leverage}
        return self._request('POST', '/fapi/v1/leverage', params=params, signed=True) is not None
    
    def set_margin_type(self, symbol, margin_type='ISOLATED'):
        """마진 타입 설정 (ISOLATED: 격리, CROSSED: 교차)"""
        params = {'symbol': symbol.replace('/', ''), 'marginType': margin_type}
        try:
            result = self._request('POST', '/fapi/v1/marginType', params=params, signed=True)
            return result is not None
        except Exception as e:
            # 이미 설정되어 있으면 에러 발생 (무시)
            if 'No need to change margin type' in str(e):
                return True
            return False
    
    def create_order(self, symbol, side, quantity, position_side=None):
        # 코인별 수량 소수점 자리수 설정 (precision_map)
        symbol_clean = symbol.replace('/', '')
        
        # 전체 코인 precision 맵
        precision_map = {
            'BTCUSDT': 3, 'ETHUSDT': 3, 'BNBUSDT': 2, 'SOLUSDT': 1, 'XRPUSDT': 1,
            'ADAUSDT': 1, 'DOGEUSDT': 0, 'TRXUSDT': 0, 'TONUSDT': 1, 'LINKUSDT': 2,
        }
        
        # precision 가져오기 (기본값: 2)
        precision = precision_map.get(symbol_clean, 2)
        
        # 수량 반올림
        if precision == 0:
            quantity = int(round(quantity))  # 정수로 반올림
        else:
            quantity = round(quantity, precision)
        
        # 최소 주문 수량 체크
        min_qty = 10 ** (-precision) if precision > 0 else 1
        if quantity < min_qty:
            print(f"주문 수량 부족: {quantity} < {min_qty}")
            return None
        
        params = {
            'symbol': symbol_clean,
            'side': side.upper(),
            'type': 'MARKET',
            'quantity': quantity
        }
        
        # positionSide는 Hedge Mode에서만 사용
        # One-Way Mode에서는 제거
        # 일단 positionSide 없이 시도
        
        result = self._request('POST', '/fapi/v1/order', params=params, signed=True)

        if result is None:
            print(f"주문 실패: {symbol} {side} {quantity}")
        else:
            self.invalidate_position_cache()  # 🔥 주문 성공 후 캐시 무효화

        return result

    def _load_exchange_info(self):
        """심볼별 가격 tickSize 로드 (최초 1회만 시도, 실패 시 fallback 사용)"""
        if getattr(self, '_tick_cache_loaded', False):
            return
        self._tick_cache_loaded = True
        self._tick_cache = {}
        try:
            info = self._request('GET', '/fapi/v1/exchangeInfo')
            if info and 'symbols' in info:
                for s in info['symbols']:
                    for f in s.get('filters', []):
                        if f.get('filterType') == 'PRICE_FILTER':
                            self._tick_cache[s['symbol']] = f['tickSize']
                print(f"✅ exchangeInfo 로드: {len(self._tick_cache)}개 심볼 tickSize 캐시")
        except Exception as e:
            print(f"⚠️ exchangeInfo 로드 실패 (fallback 사용): {e}")

    # tickSize fallback (exchangeInfo 조회 실패 시)
    _TICK_FALLBACK = {
        'BTCUSDT': '0.10', 'ETHUSDT': '0.01', 'BNBUSDT': '0.010', 'SOLUSDT': '0.0100',
        'XRPUSDT': '0.0001', 'ADAUSDT': '0.00010', 'DOGEUSDT': '0.000010',
        'TRXUSDT': '0.00001', 'TONUSDT': '0.0001', 'LINKUSDT': '0.001',
    }

    def round_price(self, symbol, price):
        """tickSize 배수로 가격 반올림 → 문자열 반환 (주문 파라미터용)"""
        from decimal import Decimal, ROUND_HALF_UP
        self._load_exchange_info()
        symbol_clean = symbol.replace('/', '')
        tick_str = self._tick_cache.get(symbol_clean) or self._TICK_FALLBACK.get(symbol_clean, '0.0001')
        tick = Decimal(tick_str)
        steps = (Decimal(str(price)) / tick).quantize(Decimal('1'), rounding=ROUND_HALF_UP)
        result = steps * tick
        decimals = -tick.normalize().as_tuple().exponent
        if decimals < 0:
            decimals = 0
        return f"{result:.{decimals}f}"

    def create_tp_order(self, symbol, position_side, stop_price):
        """🔥 거래소 측 TP 주문 (TAKE_PROFIT_MARKET + closePosition)

        거래소가 직접 트리거하므로 프로그램 지연 없이 설정가 근처에서 청산됨.
        position_side: 'long'이면 SELL TP, 'short'이면 BUY TP
        """
        params = {
            'symbol': symbol.replace('/', ''),
            'side': 'SELL' if position_side == 'long' else 'BUY',
            'type': 'TAKE_PROFIT_MARKET',
            'stopPrice': self.round_price(symbol, stop_price),
            'closePosition': 'true',
            'workingType': 'CONTRACT_PRICE',  # 차트(최종가) 기준 트리거
        }
        return self._request('POST', '/fapi/v1/order', params=params, signed=True)

    def cancel_all_orders(self, symbol):
        """심볼의 모든 대기 주문 취소 (잔여 TP 주문 정리)"""
        params = {'symbol': symbol.replace('/', '')}
        return self._request('DELETE', '/fapi/v1/allOpenOrders', params=params, signed=True)

    def get_position(self, symbol):
        # 🔥 캐시된 positionRisk 사용 (3초 TTL)
        with self._position_cache_lock:
            now = time.time()
            if self._position_cache is not None and (now - self._position_cache_time) < self._position_cache_ttl:
                positions = self._position_cache
            else:
                positions = self._request('GET', '/fapi/v2/positionRisk', signed=True)
                if positions is not None:
                    self._position_cache = positions
                    self._position_cache_time = now
                else:
                    # API 실패 — 캐시가 있으면 캐시 사용 (10초 이내)
                    if self._position_cache is not None and (now - self._position_cache_time) < 10:
                        positions = self._position_cache
                    else:
                        return None  # API 실패 + 캐시도 없음
        
        if positions is None:
            return None  # API 실패
        if positions:
            for pos in positions:
                if pos['symbol'] == symbol.replace('/', '') and float(pos['positionAmt']) != 0:
                    position_amt = float(pos['positionAmt'])
                    entry_price = float(pos['entryPrice'])
                    mark_price = float(pos['markPrice'])
                    leverage = int(pos['leverage'])
                    
                    if position_amt > 0:
                        roi_pct = ((mark_price - entry_price) / entry_price) * 100 * leverage
                    else:
                        roi_pct = ((entry_price - mark_price) / entry_price) * 100 * leverage
                    
                    return {
                        'side': 'long' if position_amt > 0 else 'short',
                        'amount': abs(position_amt),
                        'entry_price': entry_price,
                        'mark_price': mark_price,
                        'leverage': leverage,
                        'pnl': float(pos['unRealizedProfit']),
                        'roi_pct': roi_pct
                    }
        return False  # API 성공 + 포지션 없음
    
    def invalidate_position_cache(self):
        """포지션 캐시 무효화 — 진입/청산 직후 호출"""
        with self._position_cache_lock:
            self._position_cache = None
            self._position_cache_time = 0
    
    def close_position(self, symbol):
        # 🔥 대기 중인 TP 주문 먼저 취소 (고아 주문 방지)
        try:
            self.cancel_all_orders(symbol)
        except Exception:
            pass
        position = self.get_position(symbol)
        if position:
            side = 'SELL' if position['side'] == 'long' else 'BUY'
            result = self.create_order(symbol, side, position['amount'])
            self.invalidate_position_cache()  # 🔥 청산 후 캐시 무효화
            return result
        return True
    
    def get_order(self, symbol, order_id):
        params = {'symbol': symbol.replace('/', ''), 'orderId': order_id}
        return self._request('GET', '/fapi/v1/order', params=params, signed=True)
    
    def get_income_history(self, symbol, income_type=None, limit=10):
        params = {'symbol': symbol.replace('/', ''), 'limit': limit}
        if income_type:
            params['incomeType'] = income_type
        return self._request('GET', '/fapi/v1/income', params=params, signed=True)
    
    def get_last_trade_info(self, symbol):
        try:
            income = self.get_income_history(symbol, 'REALIZED_PNL', limit=1)
            realized_pnl = float(income[0]['income']) if income and len(income) > 0 else 0
            commission = self.get_income_history(symbol, 'COMMISSION', limit=5)
            total_commission = 0
            if commission:
                for c in commission[:2]:
                    total_commission += abs(float(c['income']))
            return {'realized_pnl': realized_pnl, 'commission': total_commission, 'net_pnl': realized_pnl - total_commission}
        except Exception as e:
            print(f"[⚠️ 거래 정보 조회 실패] {symbol}: {e}")
            return None

# ==================== 지표 계산 ====================
# 🔥 Indicators는 3_indicators.py에서 import됨 (중복 제거!)
# 위의 import 섹션에서 이미 로드됨: from 3_indicators import Indicators

# 🔥 스레드 Lock (로그 중복 방지)
import threading
LOG_LOCK = threading.Lock()
ENTRY_LOCK = threading.Lock()
CLOSE_LOCK = threading.Lock()  # 🔥 거래소 TP 체결 감지 중복 방지

# ==================== 트레이딩 봇 ====================
class TradingBot:
    def __init__(self, api, coin_config, log_callback, stats_callback, excel_callback=None, bot_type='long', app=None, is_reconnect=False):
        self.api = api
        self.config = coin_config
        self.log_callback = log_callback
        self.stats_callback = stats_callback
        self.excel_callback = excel_callback
        self.bot_type = bot_type
        self.app = app
        self.is_reconnect = is_reconnect
        self.running = False
        self.position = None
        self.entry_price = None
        # last_close_time, is_closing, is_entering은 coin에 저장 (LONG/SHORT 공유)
        
        # 🔥 로그 ID (중복 방지)
        self._last_log_id = None
        
        # 통계 추적
        self.stats = {
            'total_pnl': 0,
            'long_count': 0,
            'short_count': 0,
            'long_win': 0,
            'short_win': 0,
            'long_loss': 0,
            'short_loss': 0,
            'long_profit': 0,
            'short_profit': 0,
            'long_fee': 0,  # 🔥 LONG 수수료
            'short_fee': 0,  # 🔥 SHORT 수수료
            'total_fee': 0,  # 🔥 총 수수료
        }
    
    def format_price(self, price):
        """코인별 소수점 자릿수에 맞게 가격 포맷"""
        symbol = self.config['symbol'].replace('/', '')
        # 가격대별 소수점 결정
        if price >= 10000:  # BTC
            return f"{price:.2f}"
        elif price >= 1000:  # ETH
            return f"{price:.2f}"
        elif price >= 100:  # BNB, SOL
            return f"{price:.2f}"
        elif price >= 10:
            return f"{price:.3f}"
        elif price >= 1:
            return f"{price:.4f}"
        elif price >= 0.1:
            return f"{price:.5f}"
        elif price >= 0.01:
            return f"{price:.6f}"
        else:  # TRX, DOGE 등
            return f"{price:.8f}"
    
    def log(self, message, position_type):
        """포지션 타입에 따라 올바른 로그에 기록 (중복 방지)"""
        # 🔥 중복 로그 방지 (같은 메시지 연속 출력 차단)
        log_id = f"{position_type}:{message[:50]}"  # 메시지 앞 50자로 ID 생성
        with LOG_LOCK:
            if self._last_log_id == log_id:
                return  # 중복이면 무시
            self._last_log_id = log_id
        self.log_callback(message, position_type)
    
    def save_trade_excel(self, pos_type, roi_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, trade_type):
        """엑셀에 거래 기록 저장 - 바이낸스 실제 데이터 우선"""
        if not self.excel_callback:
            return
        a_pnl, a_fee, a_net, a_roi = pnl_usd, total_fee, net_profit, roi_pct
        try:
            time.sleep(1)
            bd = self.api.get_last_trade_info(self.config['symbol'])
            if bd and bd.get('realized_pnl', 0) != 0:
                a_pnl = bd['realized_pnl']
                a_fee = bd['commission']
                a_net = bd['net_pnl']
                if self.config['amount'] > 0:
                    a_roi = (a_pnl / self.config['amount']) * 100
                print(f"[📊 바이낸스] {self.config['symbol']} {pos_type} 실현:{a_pnl:.4f} 수수료:{a_fee:.4f} 순:{a_net:.4f}")
        except Exception as e:
            print(f"[⚠️ 바이낸스 조회 실패] {e}")
        if True:
            self.excel_callback(
                coin=self.config,
                position_type=pos_type,
                entry_amount=self.config['amount'],
                leverage=self.config['leverage'],
                tp_pct=self.config.get('entry_tp_long', 0) if pos_type == 'LONG' else self.config.get('entry_tp_short', 0),
                sl_pct=0,
                roi_pct=a_roi,
                profit_usdt=a_pnl if a_pnl > 0 else 0,
                loss_usdt=abs(a_pnl) if a_pnl < 0 else 0,
                entry_fee=a_fee/2 if a_fee != total_fee else entry_fee,
                close_fee=a_fee/2 if a_fee != total_fee else close_fee,
                total_fee=a_fee,
                net_profit=a_net,
                trade_type=trade_type
            )
    
    def place_exchange_tp(self, pos_type, entry_price, tp_pct):
        """🔥 진입 직후 거래소에 TP 트리거 주문 등록

        설정한 TP 가격에 도달하면 거래소가 즉시 청산 → 폴링 지연/슬리피지 최소화.
        등록 실패 시 False 반환 (기존 봇 폴링 익절이 백업으로 동작).
        """
        try:
            if pos_type == 'LONG':
                tp_price = entry_price * (1 + tp_pct / 100)
            else:
                tp_price = entry_price * (1 - tp_pct / 100)
            result = self.api.create_tp_order(self.config['symbol'], pos_type.lower(), tp_price)
            if result:
                tp_price_str = self.api.round_price(self.config['symbol'], tp_price)
                self.log(f"   📌 거래소 TP 주문 등록: ${tp_price_str} (도달 시 즉시 청산)", pos_type)
                print(f"[📌 TP 주문] {self.config['symbol']} {pos_type} → ${tp_price_str}")
                return True
            self.log(f"   ⚠️ 거래소 TP 주문 등록 실패 → 봇 폴링으로 익절 처리", pos_type)
        except Exception as e:
            print(f"[{self.config['symbol']}] TP 주문 등록 오류: {e}")
        return False

    def _handle_external_close(self):
        """🔥 포지션이 봇 청산 없이 사라짐 = 거래소 TP 주문 체결 (또는 수동 청산)

        바이낸스 실현 손익을 조회해서 통계/로그/엑셀에 기록.
        """
        with CLOSE_LOCK:
            prev = self.config.get('_cached_position')
            if not prev or not isinstance(prev, dict) or self.config.get('is_closing'):
                return
            side = prev.get('side')
            if side not in ('long', 'short') or not self.config.get(f'entry_tp_{side}'):
                return
            self.config['_cached_position'] = None  # 소비 (중복 기록 방지)

        pos_type = side.upper()
        symbol = self.config['symbol']
        print(f"[🎯 TP 체결 감지] {symbol} {pos_type} 포지션이 거래소에서 청산됨")

        # 잔여 주문 정리
        try:
            self.api.cancel_all_orders(symbol)
        except Exception:
            pass

        # 바이낸스 실현 손익 조회
        pnl_usd = prev.get('pnl', 0)
        position_size = self.config['amount'] * self.config['leverage']
        entry_fee = self.config.get(f'entry_fee_{side}', position_size * FEE_RATE)
        close_fee = position_size * FEE_RATE
        total_fee = entry_fee + close_fee
        try:
            time.sleep(1)
            info = self.api.get_last_trade_info(symbol)
            if info and info.get('realized_pnl', 0) != 0:
                pnl_usd = info['realized_pnl']
                if info.get('commission', 0) > 0:
                    total_fee = info['commission']
        except Exception as e:
            print(f"[{symbol}] 실현 손익 조회 실패: {e}")

        net_profit = pnl_usd - total_fee
        roi_pct = (pnl_usd / self.config['amount']) * 100 if self.config['amount'] > 0 else 0

        # 통계 (기존 익절 경로와 동일한 방식)
        self.config['stats'][f'{side}_fee'] = self.config['stats'].get(f'{side}_fee', 0) + total_fee
        self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + total_fee
        self.config['stats'][f'{side}_count'] += 1
        if net_profit > 0:
            self.config['stats'][f'{side}_win'] += 1
        else:
            self.config['stats'][f'{side}_loss'] += 1
        self.config['stats'][f'{side}_profit'] += net_profit
        self.config['stats']['total_pnl'] += net_profit

        count = self.config['stats'][f'{side}_count']
        profit_sign = '+' if net_profit >= 0 else ''

        # 엑셀 저장
        self.save_trade_excel(pos_type, roi_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, '익절(TP주문)')

        # 로그
        self.log(f"✅ {pos_type} 익절! 거래소 TP 주문 체결 (#{count})", pos_type)
        self.log(f"   💰 실현 수익: ${pnl_usd:.2f}", pos_type)
        self.log(f"   💸 수수료: ${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}", pos_type)
        print(f"[✅ 익절] {symbol} {pos_type} 거래소 TP 체결 | 순수익 {profit_sign}${abs(net_profit):.2f}")

        self.stats_callback()
        self.config['last_close_time'] = time.time()

        # 신호/ROI 초기화 (재진입 대비)
        self.config['prev_signals'] = {
            'ut_long': False, 'ut_short': False,
            'ema_long': False, 'ema_short': False,
            'long_signal': False, 'short_signal': False
        }
        self.config[f'entry_tp_{side}'] = None
        self.config[f'chart_entry_{side}'] = None
        self.config[f'entry_fee_{side}'] = 0
        self.config['roi'][f'{side}_entry'] = None
        self.config['roi'][f'{side}_current'] = 0
        self.config['roi'][f'{side}_max'] = 0
        self.config['roi'][f'{side}_min'] = 0

    def get_htf_adx(self, df_fallback):
        """🔥 ADX를 1시간봉(완성봉)으로 계산 — TP 결정용

        - 60초 캐시 (LONG/SHORT 봇 공유, API 부하 최소화)
        - 1시간봉 조회 실패 시 매매봉 데이터로 폴백
        """
        adx_period = self.config.get('adx_period', 10)
        adx_interval = self.config.get('adx_interval', '1h')

        cache = self.config.get('_adx_htf_cache')
        now = time.time()
        if cache and (now - cache[1]) < 60:
            return cache[0]

        df_adx = df_fallback
        if adx_interval and adx_interval != self.config.get('timeframe'):
            try:
                df_h = self.api.get_klines(self.config['symbol'], adx_interval)
                if df_h is not None and len(df_h) > adx_period + 5:
                    df_adx = df_h[:-1]  # 완성된 봉만
            except Exception as e:
                print(f"[{self.config['symbol']}] {adx_interval} ADX 조회 실패(매매봉 사용): {e}")

        adx_value = Indicators.calculate_adx(df_adx, period=adx_period)
        self.config['_adx_htf_cache'] = (adx_value, now)
        return adx_value

    def get_dynamic_tp(self, df):
        """
        ADX 기반 동적 TP 계산 (🔥 ADX는 1시간봉 기준!)

        Args:
            df: OHLCV 데이터프레임 (매매봉 - 1시간봉 조회 실패 시 폴백용)

        Returns:
            tuple: (tp, adx_value, market_type)
                - tp: 적용할 TP 값
                - adx_value: 현재 ADX 값 (1시간봉)
                - market_type: '추세장' 또는 '횡보장'
        """
        try:
            # 🔥 1시간봉 ADX (60초 캐시)
            adx_value = self.get_htf_adx(df)

            # ADX 기준으로 TP 결정
            if adx_value >= 21:
                # 추세장
                tp = self.config['tp_trend']
                market_type = '추세장'
            else:
                # 횡보장 (ADX < 21)
                tp = self.config['tp_sideways']
                market_type = '횡보장'
            
            return tp, adx_value, market_type
            
        except Exception as e:
            print(f"[{self.config['symbol']}] 동적 TP 계산 실패: {e}")
            # 오류 시 기본 TP 사용
            return self.config['tp'], 20.0, '알 수 없음'
        
    def start(self):
        self.running = True
        threading.Thread(target=self._run, daemon=True).start()
        
    def stop(self):
        self.running = False
        
    def _run(self):
        if self.is_reconnect:
            print(f"[🔄 재연결] {self.config['symbol']} {self.bot_type.upper()} - 포지션/ROI 유지!")
            try:
                ep = self.api.get_position(self.config['symbol'])
                if ep:
                    self.config['has_position'] = True
                    print(f"[🔄 재연결] {self.config['symbol']} 포지션 이어받음: {ep['side'].upper()}")
                else:
                    print(f"[🔄 재연결] {self.config['symbol']} 포지션 없음")
            except Exception as e:
                print(f"[🔄 재연결] {self.config['symbol']} 확인 실패: {e}")
        else:
            print(f"[🔍 시작 체크] {self.config['symbol']} - 기존 포지션 확인 중...")
            try:
                existing_position = self.api.get_position(self.config['symbol'])
                if existing_position:
                    print(f"[⚠️ 기존 포지션 발견!] {self.config['symbol']} {existing_position['side'].upper()}")
                    print(f"   진입가: ${existing_position['entry_price']:.2f}")
                    print(f"   수량: {existing_position['amount']}")
                    print(f"   PNL: ${existing_position['pnl']:.2f}")
                    print(f"")
                    print(f"[🔥 청산 시작] 깨끗한 시작을 위해 기존 포지션 청산 중...")
                    for attempt in range(3):
                        try:
                            result = self.api.close_position(self.config['symbol'])
                            if result:
                                print(f"[✅ 청산 완료] {self.config['symbol']} 기존 포지션 청산 성공!")
                                print(f"")
                                time.sleep(2)
                                break
                        except Exception as e:
                            print(f"[재시도 {attempt+1}/3] 청산 실패: {e}")
                            time.sleep(1)
                    else:
                        print(f"[❌ 청산 실패!] {self.config['symbol']} 수동으로 청산해주세요!")
                        print(f"")
                else:
                    print(f"[✅ 포지션 없음] {self.config['symbol']} 깨끗한 상태로 시작!")
                    print(f"")
                    # 🔥 이전 세션의 잔여 TP 주문 정리
                    try:
                        self.api.cancel_all_orders(self.config['symbol'])
                    except Exception:
                        pass
            except Exception as e:
                print(f"[⚠️ 체크 실패] {self.config['symbol']} 포지션 확인 실패: {e}")
                print(f"")
            
            print(f"[🔄 초기화] {self.config['symbol']} ROI 및 진입가 초기화 중...")
            self.config['roi'] = {
                'long_entry': None, 'long_current': 0, 'long_max': 0, 'long_min': 0,
                'short_entry': None, 'short_current': 0, 'short_max': 0, 'short_min': 0,
            }
            self.config['chart_entry_long'] = None
            self.config['chart_entry_short'] = None
            self.config['restart_entry_long'] = None
            self.config['restart_entry_short'] = None
            self.config['entry_fee_long'] = 0
            self.config['entry_fee_short'] = 0
            self.config['entry_tp_long'] = None
            self.config['entry_tp_short'] = None
            print(f"[✅ 초기화 완료] {self.config['symbol']} 모든 데이터 리셋!")
            print(f"")
        
        # 시작하자마자 즉시 체크 (0초 대기)
        first_check = True
        last_heartbeat = time.time()  # 🔥 하트비트
        loop_count = 0
        consecutive_errors = 0  # 🔥 연속 오류 카운터
        last_api_resync = time.time()  # 🔥 API 재동기화
        
        print(f"[🚀 봇 시작] {self.config['symbol']} - 스레드 시작!")
        print(f"")
        
        while self.running:
            try:
                loop_count += 1
                
                # 🔥 하트비트 (1시간마다 = 3600초)
                if time.time() - last_heartbeat > 3600:
                    print(f"[💓] {self.config['symbol']} 동작 중 (루프 #{loop_count}, {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})")
                    print(f"   플래그: entering_long={self.config.get('is_entering_long')}, entering_short={self.config.get('is_entering_short')}, closing={self.config.get('is_closing')}")
                    print(f"   연속오류: {consecutive_errors}회")
                    last_heartbeat = time.time()
                
                # 🔥 API 시간 재동기화 (6시간마다)
                if time.time() - last_api_resync > 21600:
                    try:
                        self.api._sync_time()
                        print(f"[🔄 시간 재동기화] {self.config['symbol']}")
                        last_api_resync = time.time()
                    except:
                        pass
                
                # 🔥 플래그 타임아웃 (stuck 방지) - 봇 타입별!
                entering_key = f'is_entering_{self.bot_type}'
                if self.config.get(entering_key):
                    start_time = self.config.get(f'{entering_key}_start', 0)
                    if start_time > 0 and time.time() - start_time > 60:
                        print(f"[⚠️ 리셋] {self.config['symbol']}: {entering_key} 60초 stuck")
                        self.config[entering_key] = False
                
                if self.config.get('is_closing'):
                    start_time = self.config.get('is_closing_start', 0)
                    if start_time > 0 and time.time() - start_time > 30:
                        print(f"[⚠️ 리셋] {self.config['symbol']}: is_closing 30초 stuck")
                        self.config['is_closing'] = False
                
                # 🔥🔥🔥 루프 주기: 포지션 있으면 3초, 없으면 5초
                if not first_check:
                    loop_interval = 3 if self.config.get('has_position') else 5
                    time.sleep(loop_interval)
                first_check = False
                
                # 🔥🔥🔥 캔들 조회 주기
                # 포지션 있으면 4초마다 (2초 루프 × 2회)
                # 포지션 없으면 15초마다 (5초 루프 × 3회)
                if not hasattr(self, '_kline_check_count'):
                    self._kline_check_count = 0
                self._kline_check_count += 1
                
                kline_interval = 2 if self.config.get('has_position') else 3
                
                if self._kline_check_count >= kline_interval:
                    self._kline_check_count = 0
                    df = self.api.get_klines(self.config['symbol'], self.config['timeframe'])
                    if df is None or len(df) < 60:
                        consecutive_errors += 1
                        if consecutive_errors >= 10:
                            print(f"[⚠️ 네트워크] {self.config['symbol']}: 연속 {consecutive_errors}회 실패, 30초 대기")
                            time.sleep(30)
                        continue
                    
                    # 성공 시 오류 카운터 리셋
                    consecutive_errors = 0
                    
                    # 🔥🔥🔥 완성된 봉만 사용 (트레이딩뷰와 동일하게!)
                    df_closed = df[:-1]
                    
                    # 캐시에 저장
                    self._cached_df = df_closed
                    self._cached_signals = Indicators.get_signals(df_closed, self.config['ut_sens'], self.config['ut_atr'],
                                                    self.config['ema_fast'], self.config['ema_slow'])
                    self._cached_dynamic_tp = self.get_dynamic_tp(df_closed)
                    

                else:
                    # 캐시된 데이터 사용
                    if not hasattr(self, '_cached_df') or self._cached_df is None:
                        continue  # 아직 캐시 없으면 스킵
                    df_closed = self._cached_df
                
                # 캐시된 신호 사용
                if not hasattr(self, '_cached_signals') or self._cached_signals is None:
                    continue
                signals = self._cached_signals
                dynamic_tp, adx_value, market_type = self._cached_dynamic_tp if hasattr(self, '_cached_dynamic_tp') else (self.config.get('tp', 1.5), 25, '추세장')
                
                # 🔥🔥🔥 포지션 확인
                # 포지션 있으면 매 루프마다 (2초)
                # 포지션 없으면 60초마다 (5초 루프 × 12회)
                if self.config.get('is_entering_long') or self.config.get('is_entering_short') or self.config.get('is_closing'):
                    current_position = self.api.get_position(self.config['symbol'])
                elif self.config.get('has_position'):  # 포지션 있으면 매번 체크 (3초마다)
                    current_position = self.api.get_position(self.config['symbol'])
                    if current_position is None:
                        # 🔥 API 실패(None)면 has_position 유지 + 이전 캐시 사용!
                        # 이래야 TP/SL 체크가 계속 됨 (익절 가능!)
                        current_position = self.config.get('_cached_position')
                    elif current_position is False or not current_position:
                        self.config['has_position'] = False  # 실제로 포지션 없음
                        # 🔥 봇이 청산한 게 아니면 = 거래소 TP 주문 체결 → 기록!
                        self._handle_external_close()
                else:
                    # 포지션 없으면 60초마다 (5초 × 12회)
                    if not hasattr(self, '_position_check_count'):
                        self._position_check_count = 0
                    self._position_check_count += 1
                    
                    if self._position_check_count >= 12:  # 60초마다
                        current_position = self.api.get_position(self.config['symbol'])
                        self._position_check_count = 0
                        if current_position:
                            self.config['has_position'] = True
                    else:
                        current_position = None
                
                # 🔥 봇 루프에서 가져온 포지션을 coin에 캐시 (ROI 표시용)
                if current_position and isinstance(current_position, dict):
                    self.config['_cached_position'] = current_position
                elif current_position is False:
                    self.config['_cached_position'] = None
                
                # 🔥 신호는 항상 업데이트 (포지션 유무 무관!)
                # 📊 현재 신호 - ut_position 사용 (현재 UT Bot 상태)
                current_long_signal = signals.get('ut_position_long', False) and signals['ema_long']
                current_short_signal = signals.get('ut_position_short', False) and signals['ema_short']
                
                # 📊 이전 신호 (없으면 None)
                prev_signals = self.config.get('prev_signals', {})
                prev_long_signal = prev_signals.get('long_signal')
                prev_short_signal = prev_signals.get('short_signal')
                
                # 진입 로직
                if not current_position and not self.config.get('has_position'):
                    # 포지션이 없을 때만 진입 (has_position 플래그도 체크!)
                    side = None
                    pos_type = None
                    
                    # 🚨 최근 청산 확인 (5초 강제 대기) - coin에서 공유
                    if self.config.get('last_close_time'):
                        time_since_close = time.time() - self.config['last_close_time']
                        if time_since_close < 5:
                            # 🔥 대기 중에는 신호 업데이트 안 함! (False 유지)
                            # 대기 완료 후 False→True 감지
                            continue

                    # 🔍 디버깅 - 신호 상태 (5초마다)
                    if not hasattr(self, '_last_signal_debug'):
                        self._last_signal_debug = 0
                    if time.time() - self._last_signal_debug > 5:
                        self._last_signal_debug = time.time()
                        if current_long_signal or current_short_signal:
                            print(f"[신호 체크] {self.config['symbol']}")
                            print(f"  Current: LONG={current_long_signal} SHORT={current_short_signal}")
                            print(f"  Prev: LONG={prev_long_signal} SHORT={prev_short_signal}")
                            print(f"  Active: LONG={self.config['long_active']} SHORT={self.config['short_active']}")
                            print(f"  Position: {current_position is not None}")
                    
                    # 🔔 신호 변경 감지
                    # 1. 처음 시작 (prev=None) + 현재 신호 True → 진입 OK!
                    # 2. False→True 변경 → 진입 OK!
                    # 3. True→True 유지 → 진입 안 함
                    
                    long_signal_new = False
                    short_signal_new = False
                    
                    # LONG 신호 체크
                    if current_long_signal:
                        if prev_long_signal is None:
                            # 처음 시작 + 신호 있음 → 진입!
                            long_signal_new = True
                            print(f"[진입 신호] {self.config['symbol']} LONG: None→True")
                        elif prev_long_signal is False:
                            # False→True 변경 → 진입!
                            long_signal_new = True
                            print(f"[진입 신호] {self.config['symbol']} LONG: False→True")
                        elif prev_long_signal is True:
                            # 🔥 포지션 없으면 재진입 허용!
                            if current_position:  # 포지션 있으면 차단
                                print(f"[진입 차단] {self.config['symbol']} LONG: 포지션 있음")
                            else:
                                long_signal_new = True
                                print(f"[진입 신호] {self.config['symbol']} LONG: True→True (신호 유지)")
                    
                    # SHORT 신호 체크
                    if current_short_signal:
                        if prev_short_signal is None:
                            # 처음 시작 + 신호 있음 → 진입!
                            short_signal_new = True
                            print(f"[진입 신호] {self.config['symbol']} SHORT: None→True")
                        elif prev_short_signal is False:
                            # False→True 변경 → 진입!
                            short_signal_new = True
                            print(f"[진입 신호] {self.config['symbol']} SHORT: False→True")
                        elif prev_short_signal is True:
                            # 🔥 포지션 없으면 재진입 허용!
                            if current_position:  # 포지션 있으면 차단
                                print(f"[진입 차단] {self.config['symbol']} SHORT: 포지션 있음")
                            else:
                                short_signal_new = True
                                print(f"[진입 신호] {self.config['symbol']} SHORT: True→True (신호 유지)")
                    
                    # 🚫 이미 진입 중이면 신호 감지 차단 - 봇 타입별!
                    entering_key = f'is_entering_{self.bot_type}'
                    if self.config.get(entering_key):
                        print(f"[진입 차단] {self.config['symbol']}: {entering_key}=True")
                        continue
                    
                    # 🔥 각 봇은 자기 타입만 진입!
                    # LONG 봇은 LONG만, SHORT 봇은 SHORT만
                    # 🔥 손절 쿨다운 체크: 손절 후에는 신호 2개 (UT + EMA) 모두 True일 때만 진입!
                    if self.bot_type == 'long' and self.config['long_active'] and long_signal_new:
                        # 🔥🔥🔥 4H UT 필터 체크!
                        if self.app:
                            filter_4h = self.app.get_4h_filter(self.config['symbol'])
                            if not filter_4h.get('ut_long', True):
                                print(f"[4H 필터] {self.config['symbol']} LONG 차단: 4H UT가 SHORT")
                                continue
                        

                        # 손절 쿨다운 중이면 신호 2개 모두 확인
                        if self.config.get('sl_cooldown_long'):
                            if current_long_signal:  # UT + EMA 둘 다 True
                                print(f"[🔓 쿨다운 해제] {self.config['symbol']} LONG: 신호 2개 충족!")
                                self.config['sl_cooldown_long'] = False
                            else:
                                print(f"[⏸️ 쿨다운 중] {self.config['symbol']} LONG: 신호 2개 충족 대기")
                                continue  # 진입 안 함
                        
                        side = 'BUY'
                        pos_type = 'LONG'
                        print(f"[진입 시도] {self.config['symbol']} LONG")
                        print(f"  UT Bot 상태: {signals.get('ut_position_long', False)} | EMA LONG: {signals['ema_long']}")
                        print(f"  → AND 조건: {current_long_signal}")
                        print(f"  🔍 long_active={self.config['long_active']}")
                        print(f"  🔍 long_signal_new={long_signal_new}")
                        print(f"  🔍 current_position={current_position}")
                        print(f"  🔍 is_entering_long={self.config.get('is_entering_long')}")
                    # 숏이 활성화되어 있고 NEW 숏 신호 발생
                    elif self.bot_type == 'short' and self.config['short_active'] and short_signal_new:
                        # 🔥🔥🔥 4H UT 필터 체크!
                        if self.app:
                            filter_4h = self.app.get_4h_filter(self.config['symbol'])
                            if not filter_4h.get('ut_short', True):
                                print(f"[4H 필터] {self.config['symbol']} SHORT 차단: 4H UT가 LONG")
                                continue
                        

                        # 손절 쿨다운 중이면 신호 2개 모두 확인
                        if self.config.get('sl_cooldown_short'):
                            if current_short_signal:  # UT + EMA 둘 다 True
                                print(f"[🔓 쿨다운 해제] {self.config['symbol']} SHORT: 신호 2개 충족!")
                                self.config['sl_cooldown_short'] = False
                            else:
                                print(f"[⏸️ 쿨다운 중] {self.config['symbol']} SHORT: 신호 2개 충족 대기")
                                continue  # 진입 안 함
                        
                        side = 'SELL'
                        pos_type = 'SHORT'
                        print(f"[진입 시도] {self.config['symbol']} SHORT")
                        print(f"  UT Bot 상태: {signals.get('ut_position_short', False)} | EMA SHORT: {signals['ema_short']}")
                        print(f"  → AND 조건: {current_short_signal}")
                        print(f"  🔍 short_active={self.config['short_active']}")
                        print(f"  🔍 short_signal_new={short_signal_new}")
                        print(f"  🔍 current_position={current_position}")
                        print(f"  🔍 is_entering_short={self.config.get('is_entering_short')}")
                    
                    # 🔥 진입 안 하는 경우 상세 로그
                    if not side:
                        if long_signal_new and not self.config['long_active']:
                            print(f"[❌ 진입 불가] {self.config['symbol']} LONG: 봇 비활성화")
                        if short_signal_new and not self.config['short_active']:
                            print(f"[❌ 진입 불가] {self.config['symbol']} SHORT: 봇 비활성화")
                        
                        # 신호는 있지만 진입 안된 경우
                        if current_long_signal and not long_signal_new:
                            print(f"[⚠️ 진입 안됨] {self.config['symbol']} LONG: prev={prev_long_signal}, current={current_long_signal}")
                        if current_short_signal and not short_signal_new:
                            print(f"[⚠️ 진입 안됨] {self.config['symbol']} SHORT: prev={prev_short_signal}, current={current_short_signal}")
                        
                        # 🔥 진입 안 하는 경우 신호 업데이트 안 함!
                        # prev 상태 유지 → 다음에 버튼 켜면 진입 가능
                    
                    if side:
                        # 🔒 진입 전 Lock으로 원자적 체크!
                        entering_key = f'is_entering_{self.bot_type}'
                        
                        with ENTRY_LOCK:
                            # 🔥🔥🔥 LONG이든 SHORT이든 어떤 봇이 진입 중이면 차단!
                            if self.config.get('is_entering_long') or self.config.get('is_entering_short'):
                                continue
                            # 🔒 즉시 플래그 설정 (Lock 안에서!)
                            self.config[entering_key] = True
                            self.config[f'{entering_key}_start'] = time.time()
                        
                        # 🔒 포지션 재확인 (다른 봇이 진입했을 수 있음)
                        double_check_position = self.api.get_position(self.config['symbol'])
                        if double_check_position:
                            print(f"[진입 차단] {self.config['symbol']}: 포지션 이미 존재!")
                            self.config[entering_key] = False
                            continue
                        # 🔥 API 실패(None)면 안전하게 진입 차단!
                        if double_check_position is None and self.config.get('has_position'):
                            print(f"[진입 차단] {self.config['symbol']}: API 실패 + has_position=True")
                            self.config[entering_key] = False
                            continue
                        
                        try:
                            # 🔥 격리 모드 설정 (고정)
                            self.api.set_margin_type(self.config['symbol'], 'ISOLATED')
                            
                            # 진입
                            self.api.set_leverage(self.config['symbol'], self.config['leverage'])
                            
                            # 수량 계산 (레버리지 고려)
                            qty = (self.config['amount'] * self.config['leverage']) / signals['price']
                            
                            # 🔥 코인별 정밀도 (바이낸스 선물 실제 기준!)
                            symbol_clean = self.config['symbol'].replace('/', '')
                            
                            precision_map = {
                                'BTCUSDT': 3, 'ETHUSDT': 3, 'BNBUSDT': 2, 'LINKUSDT': 2,
                                'SOLUSDT': 1, 'XRPUSDT': 1, 'ADAUSDT': 1, 'TONUSDT': 1,
                                'DOGEUSDT': 0, 'TRXUSDT': 0,
                            }
                            precision = precision_map.get(symbol_clean, 2)
                            
                            if precision == 0:
                                qty = int(round(qty))  # 반올림 후 정수
                                min_qty = 1
                            else:
                                qty = round(qty, precision)
                                min_qty = 10 ** (-precision)
                            
                            # 최소 주문량 체크
                            if qty < min_qty:
                                self.log(f"⚠️ 주문 금액이 너무 작습니다! 현재:{qty} 최소:{min_qty}", pos_type)
                                self.log(f"   💡 진입금을 늘리거나 레버리지를 높이세요", pos_type)
                                self.config[entering_key] = False
                                continue
                            
                            # One-Way Mode: positionSide 없이 주문
                            order = self.api.create_order(self.config['symbol'], side, qty)
                            if order:
                                # 🔥 진입 시점의 TP 저장 (LONG/SHORT 분리!)
                                if pos_type == 'LONG':
                                    self.config['entry_tp_long'] = dynamic_tp
                                    self.config['tp_reached_long'] = False
                                else:  # SHORT
                                    self.config['entry_tp_short'] = dynamic_tp
                                    self.config['tp_reached_short'] = False
                                
                                # 진입 시간 기록
                                self.config['entry_time'] = time.time()
                                
                                # 목표 ROI 계산 (동적 TP 사용)
                                target_tp_roi = dynamic_tp * self.config['leverage']
                                
                                # 🔥 목표 USDT 계산 (LONG/SHORT 분리!)
                                target_usdt = self.config['amount'] * (target_tp_roi / 100)
                                if pos_type == 'LONG':
                                    self.config['target_usdt_long'] = target_usdt
                                else:  # SHORT
                                    self.config['target_usdt_short'] = target_usdt
                                
                                # 🔥 수수료 계산 및 즉시 통계 반영!
                                leverage = self.config['leverage']
                                position_size = self.config['amount'] * leverage
                                entry_fee = position_size * FEE_RATE  # 0.04%
                                
                                # 🔥 진입 수수료 저장
                                if pos_type == 'LONG':
                                    self.config['entry_fee_long'] = entry_fee
                                    # 🔥 즉시 통계에 반영!
                                    self.config['stats']['long_fee'] = self.config['stats'].get('long_fee', 0) + entry_fee
                                else:
                                    self.config['entry_fee_short'] = entry_fee
                                    # 🔥 즉시 통계에 반영!
                                    self.config['stats']['short_fee'] = self.config['stats'].get('short_fee', 0) + entry_fee
                                
                                # 🔥 전체 수수료 및 PNL에 즉시 반영!
                                self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + entry_fee
                                self.config['stats']['total_pnl'] -= entry_fee  # 진입 시점에 차감!
                                
                                # 🔥 통계 UI 즉시 업데이트!
                                self.stats_callback()
                                
                                # 🔥🔥🔥 메인넷 차트 진입가 저장! (ROI 0% 시작용)
                                # 테스트넷 체결가가 아닌 메인넷 차트 가격 사용
                                chart_entry_price = signals['price']  # 메인넷 현재가
                                if pos_type == 'LONG':
                                    self.config['chart_entry_long'] = chart_entry_price
                                else:
                                    self.config['chart_entry_short'] = chart_entry_price
                                
                                # 로그 출력 (동적 TP 값으로 표시)
                                entry_price_str = self.format_price(signals['price'])
                                self.log(f"✅ {pos_type} 진입! ${entry_price_str} | {qty}개 | {self.config['timeframe']} | {self.config['leverage']}x", pos_type)
                                self.log(f"   📊 ADX: {adx_value:.1f} ({market_type}) → TP {dynamic_tp}%", pos_type)
                                self.log(f"   🎯 TP: 가격+{dynamic_tp:.2f}% → ROI +{target_tp_roi:.2f}% → 💰 ${target_usdt:.2f}", pos_type)
                                self.log(f"   💸 예상 수수료: ${entry_fee:.2f} + ${entry_fee:.2f} = ${entry_fee*2:.2f} (0.08%)", pos_type)
                                self.log(f"   🛡️ SL: AUTO (스위칭)", pos_type)
                                
                                # 🔥 포지션 있음 플래그 설정!
                                self.config['has_position'] = True
                                

                                
                                # 콘솔 출력
                                print(f"[✅ 진입] {self.config['symbol']} {pos_type} ${entry_price_str} | ADX={adx_value:.1f} ({market_type}) | TP={dynamic_tp}%")
                                print(f"   💰 목표 USDT: ${target_usdt:.2f} | 💸 예상 수수료: ${entry_fee*2:.2f} (0.08%)")
                                
                                # 🔥 order response에서 평균 체결가 가져오기!
                                filled_price = None
                                
                                # 1차 시도: order response에서 직접 가져오기
                                if order and 'avgPrice' in order and order['avgPrice']:
                                    filled_price = float(order['avgPrice'])
                                    print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f}")
                                
                                # 2차 시도: order 정보 재조회
                                if not filled_price and order and 'orderId' in order:
                                    try:
                                        time.sleep(0.5)  # 0.5초 대기
                                        order_info = self.api.get_order(self.config['symbol'], order['orderId'])
                                        if order_info and 'avgPrice' in order_info and order_info['avgPrice']:
                                            filled_price = float(order_info['avgPrice'])
                                            print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (재조회)")
                                    except Exception as e:
                                        print(f"[{self.config['symbol']}] order 재조회 실패: {e}")
                                
                                # 3차 시도: 체결 내역(fills)에서 계산
                                if not filled_price and order and 'fills' in order and order['fills']:
                                    try:
                                        total_qty = sum(float(fill['qty']) for fill in order['fills'])
                                        total_cost = sum(float(fill['price']) * float(fill['qty']) for fill in order['fills'])
                                        if total_qty > 0:
                                            filled_price = total_cost / total_qty
                                            print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (fills 계산)")
                                    except Exception as e:
                                        print(f"[{self.config['symbol']}] fills 계산 실패: {e}")
                                
                                # 출력 (체결가를 못 가져왔을 경우)
                                if not filled_price:
                                    print(f"[{self.config['symbol']}] 체결 가격 조회 실패 - position에서 확인 예정")
                                
                                # 🔥 ROI 완전 초기화 (진입가 같아도 초기화!)
                                if pos_type == 'LONG':
                                    self.config['roi']['long_entry'] = None
                                    self.config['roi']['long_current'] = 0
                                    self.config['roi']['long_max'] = 0
                                    self.config['roi']['long_min'] = 0
                                    # 🔥 order 체결가를 restart_entry로 설정! (가장 정확!)
                                    if filled_price:
                                        self.config['restart_entry_long'] = filled_price
                                        print(f"[{self.config['symbol']}] LONG restart_entry 설정 (order): ${filled_price:.2f}")
                                    else:
                                        # fallback: 포지션 재조회
                                        time.sleep(1)  # 1초 대기 (기존 0.5초보다 길게)
                                        position_check = self.api.get_position(self.config['symbol'])
                                        if position_check:
                                            self.config['restart_entry_long'] = position_check['entry_price']
                                            print(f"[{self.config['symbol']}] LONG restart_entry 설정 (position): ${position_check['entry_price']:.2f}")
                                        else:
                                            self.config['restart_entry_long'] = signals['price']
                                            print(f"[{self.config['symbol']}] LONG restart_entry 설정 (신호): ${signals['price']:.2f}")
                                else:
                                    self.config['roi']['short_entry'] = None
                                    self.config['roi']['short_current'] = 0
                                    self.config['roi']['short_max'] = 0
                                    self.config['roi']['short_min'] = 0
                                    # 🔥 order 체결가를 restart_entry로 설정! (가장 정확!)
                                    if filled_price:
                                        self.config['restart_entry_short'] = filled_price
                                        print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (order): ${filled_price:.2f}")
                                    else:
                                        # fallback: 포지션 재조회
                                        time.sleep(1)  # 1초 대기
                                        position_check = self.api.get_position(self.config['symbol'])
                                        if position_check:
                                            self.config['restart_entry_short'] = position_check['entry_price']
                                            print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (position): ${position_check['entry_price']:.2f}")
                                        else:
                                            self.config['restart_entry_short'] = signals['price']
                                            print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (신호): ${signals['price']:.2f}")
                                
                                # 🔥🔥🔥 거래소 TP 주문 등록 (설정가 도달 시 거래소가 즉시 청산!)
                                tp_base = filled_price or self.config.get(f'restart_entry_{self.bot_type}') or signals['price']
                                self.place_exchange_tp(pos_type, tp_base, dynamic_tp)

                                # 🔥 is_closing 해제 (새 진입이므로)
                                self.config['is_closing'] = False
                                
                                # 🔥 진입 성공 후 신호 False 초기화 (익절 후 재진입 대비!)
                                self.config['prev_signals'] = {
                                    'ut_long': False,
                                    'ut_short': False,
                                    'ema_long': False,
                                    'ema_short': False,
                                    'long_signal': False,
                                    'short_signal': False
                                }
                                
                                # 🔥 진입 성공 시 10초 대기 (중복 완전 차단!)
                                time.sleep(10)  # 10초 대기
                        
                        except Exception as e:
                            self.log(f"❌ 진입 오류: {str(e)}", pos_type)
                            print(f"[❌ 진입 오류] {self.config['symbol']} {pos_type}: {str(e)}")
                        
                        finally:
                            # 진입 완료 - 플래그 해제 (봇 타입별!)
                            self.config[entering_key] = False
                    
                    # 🔥 진입 실패 시 신호 업데이트 안 함!
                    # prev 상태 유지하여 다음에 재시도 가능
                
                # 청산 로직
                elif current_position:
                    # 🔥🔥🔥 핵심: 자기 타입의 포지션만 처리!
                    # LONG 봇은 LONG 포지션만, SHORT 봇은 SHORT 포지션만
                    if current_position['side'] != self.bot_type:
                        # 다른 타입의 포지션이면 무시 (반대 봇이 처리함)
                        continue
                    
                    entry = current_position['entry_price']
                    current = current_position.get('mark_price', signals['price'])
                    leverage = current_position.get('leverage', self.config['leverage'])
                    
                    # 🔥 ROI 업데이트 (익절 판단용)
                    roi_pct = current_position.get('roi_pct', 0)
                    if current_position['side'] == 'long':
                        self.config['roi']['long_current'] = roi_pct
                    else:
                        self.config['roi']['short_current'] = roi_pct
                    
                    # 🔥 바이낸스 ROI 직접 사용
                    pnl_pct = roi_pct
                    pnl_usd = current_position['pnl']
                    
                    if current_position['side'] == 'long':
                        # 🔥 진입 시 저장한 동적 TP 사용! (ADX 기반)
                        user_tp = self.config.get('entry_tp_long') or self.config.get('tp', 0.3)
                        target_roi = user_tp * leverage
                        
                        # ROI >= 목표면 즉시 익절
                        if pnl_pct >= target_roi and not self.config.get("is_closing"):
                            print(f"[🎯 익절] {self.config['symbol']} LONG: ROI {pnl_pct:.2f}% >= 목표 {target_roi:.2f}% (TP {user_tp}%)")
                            
                            self.config["is_closing"] = True
                            self.config["is_closing_start"] = time.time()
                            
                            close_success = False
                            for attempt in range(3):
                                try:
                                    result = self.api.close_position(self.config['symbol'])
                                    if result:
                                        close_success = True
                                        break
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] 청산 재시도 {attempt+1}/3")
                                time.sleep(1)
                            
                            if not close_success:
                                self.log(f"❌ 청산 실패! 수동으로 청산하세요!", 'LONG')
                                print(f"[❌ 청산 실패] {self.config['symbol']} LONG - 수동 청산 필요!")
                                self.config["is_closing"] = False
                                continue
                            
                            # 🔥 수수료 계산
                            position_size = self.config['amount'] * self.config['leverage']
                            entry_fee = self.config.get('entry_fee_long', position_size * FEE_RATE)
                            close_fee = position_size * FEE_RATE
                            total_fee = entry_fee + close_fee
                            net_profit = pnl_usd - total_fee
                            
                            # 통계
                            self.config['stats']['long_fee'] = self.config['stats'].get('long_fee', 0) + total_fee
                            self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + total_fee
                            
                            close_price = current
                            
                            self.config['stats']['long_count'] += 1
                            self.config['stats']['long_win'] += 1
                            self.config['stats']['long_profit'] += net_profit
                            self.config['stats']['total_pnl'] += net_profit
                            
                            profit_sign = '+' if net_profit >= 0 else ''
                            count = self.config['stats']['long_count']
                            
                            # 엑셀 저장
                            self.save_trade_excel('LONG', pnl_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, '익절')
                            
                            # 로그
                            self.log(f"✅ LONG 익절! ROI +{pnl_pct:.2f}% (목표 {target_roi:.2f}%) (#{count})", 'LONG')
                            self.log(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f}", 'LONG')
                            self.log(f"   💸 수수료: ${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}", 'LONG')
                            
                            print(f"[✅ 익절] {self.config['symbol']} LONG ROI +{pnl_pct:.2f}% (목표 {target_roi:.2f}%)")
                            
                            self.stats_callback()
                            self.config['last_close_time'] = time.time()
                            
                            # 신호 초기화
                            self.config['prev_signals'] = {
                                'ut_long': False, 'ut_short': False,
                                'ema_long': False, 'ema_short': False,
                                'long_signal': False, 'short_signal': False
                            }
                            
                            # ROI 초기화
                            self.config['entry_tp_long'] = None
                            self.config['chart_entry_long'] = None
                            self.config['entry_fee_long'] = 0
                            self.config['roi']['long_entry'] = None
                            self.config['roi']['long_current'] = 0
                            self.config['roi']['long_max'] = 0
                            self.config['roi']['long_min'] = 0
                            
                            print(f"[{self.config['symbol']}] LONG 익절 완료 - 재진입 대기")
                            self.config["is_closing"] = False
                            continue
                        
                        # 🔥 손절: AUTO(스위칭) - UT Bot 현재 상태 + EMA 크로스
                        auto_sl = signals.get('ut_position_short', False) and signals['ema_short']
                        

                        # 🔍 디버깅: 신호 상태 출력 (신호 있을 때만)
                        if signals.get('ut_position_short', False) or signals['ema_short']:
                            if not self.config.get("is_closing"):
                                print(f"[DEBUG] {self.config['symbol']} LONG 포지션 - 스위칭 신호 체크:")
                                print(f"   UT Bot 상태: position_short={signals.get('ut_position_short', False)}")
                                print(f"   EMA SHORT: {signals['ema_short']}")
                                print(f"   → auto_sl (AND): {auto_sl}")
                        
                        if auto_sl and not self.config.get("is_closing"):
                            print(f"")
                            print(f"[🔄 스위칭 실행!] {self.config['symbol']} LONG→SHORT")
                            print(f"   ✅ UT Bot 상태: SHORT={signals.get('ut_position_short', False)}")
                            print(f"   ✅ EMA SHORT: {signals['ema_short']}")
                            print(f"   ✅ AND 조건: True")
                            print(f"")
                            
                            self.config["is_closing"] = True
                            self.config["is_closing_start"] = time.time()
                            
                            close_success = False
                            for attempt in range(3):
                                try:
                                    result = self.api.close_position(self.config['symbol'])
                                    if result:
                                        close_success = True
                                        break
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] 스위칭 청산 시도 {attempt+1}/3 실패: {e}")
                                time.sleep(1)
                            
                            if not close_success:
                                self.log(f"❌ 스위칭 청산 실패! 수동으로 청산하세요!", 'LONG')
                                print(f"[❌ 스위칭 실패] {self.config['symbol']} LONG→SHORT - 수동 청산 필요!")
                                self.config["is_closing"] = False
                                continue
                            
                            # 🔥 가격 변동률 계산
                            price_change_pct = ((current - entry) / entry) * 100
                            
                            # 🔥 수수료 계산
                            position_size = self.config['amount'] * self.config['leverage']
                            entry_fee = self.config.get('entry_fee_long', position_size * FEE_RATE)
                            close_fee = position_size * FEE_RATE
                            total_fee = entry_fee + close_fee
                            net_profit = pnl_usd - total_fee
                            
                            self.config['stats']['long_fee'] = self.config['stats'].get('long_fee', 0) + total_fee
                            self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + total_fee
                            
                            close_price = current
                            
                            self.config['stats']['long_count'] += 1
                            if net_profit > 0:
                                self.config['stats']['long_win'] += 1
                            else:
                                self.config['stats']['long_loss'] += 1
                            self.config['stats']['long_profit'] += net_profit
                            self.config['stats']['total_pnl'] += net_profit
                            
                            profit_sign = '+' if net_profit >= 0 else ''
                            
                            # 🔥 엑셀 저장!
                            self.save_trade_excel('LONG', pnl_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, '스위칭')
                            
                            # 🔥 로그에 청산 가격, 수수료, 순수익 표시
                            self.log(f"🔄 LONG→SHORT 스위칭! 가격{'+' if price_change_pct >= 0 else ''}{price_change_pct:.2f}% | ROI {'+' if pnl_pct >= 0 else ''}{pnl_pct:.2f}%", 'LONG')
                            self.log(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f}", 'LONG')
                            self.log(f"   💸 총 수수료: -${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}", 'LONG')
                            
                            # 콘솔 출력
                            print(f"[🔄 스위칭] {self.config['symbol']} LONG→SHORT {'+' if price_change_pct >= 0 else ''}{price_change_pct:.2f}%")
                            print(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f} | 총 수수료: -${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}")
                            
                            # 통계 업데이트
                            self.stats_callback()
                            
                            # 🔥 즉시 SHORT 진입 (스위칭)
                            self.log(f"🔄 즉시 SHORT 진입 준비...", 'SHORT')
                            print(f"[{self.config['symbol']}] LONG→SHORT 스위칭: 즉시 SHORT 진입!")
                            
                            # 🔥 LONG ROI 초기화 (청산됨)
                            self.config['roi']['long_entry'] = None
                            self.config['roi']['long_current'] = 0
                            self.config['roi']['long_max'] = 0
                            self.config['roi']['long_min'] = 0
                            self.config['restart_entry_long'] = None  # 재시작 price도 초기화
                            self.config['target_usdt_long'] = 0  # 목표 USDT 초기화!
                            self.config['entry_tp_long'] = None
                            self.config['tp_reached_long'] = False
                            self.config['entry_fee_long'] = 0  # 진입 수수료 초기화
                            
                            # 5초 대기
                            time.sleep(5)
                            
                            # 청산 완료 - 플래그 해제
                            self.config["warned_position_exists"] = False
                            self.config["is_closing"] = False
                            
                            # 🔥 스위칭은 역신호만 있어도 진입! (2개 조건 불필요)
                            # SHORT 봇이 활성화되어 있으면 즉시 진입
                            if self.config['short_active']:
                                try:
                                    # 🔒 진입 플래그 설정 - Lock으로 보호!
                                    with ENTRY_LOCK:
                                        if self.config.get('is_entering_long') or self.config.get('is_entering_short'):
                                            print(f"[{self.config['symbol']}] 스위칭 진입 차단: 다른 진입 중")
                                            continue
                                        self.config["is_entering_short"] = True
                                        self.config["is_entering_short_start"] = time.time()
                                    
                                    # 포지션 재확인
                                    double_check = self.api.get_position(self.config['symbol'])
                                    if double_check:
                                        print(f"[{self.config['symbol']}] 스위칭 실패: 이미 포지션 있음")
                                        self.config["is_entering_short"] = False
                                        continue
                                    
                                    # 소수점 처리 (바이낸스 실제 기준)
                                    symbol_clean = self.config['symbol'].replace('/', '')
                                    p_map = {'BTCUSDT': 3, 'ETHUSDT': 3, 'BNBUSDT': 2, 'LINKUSDT': 2,
                                             'SOLUSDT': 1, 'XRPUSDT': 1, 'ADAUSDT': 1, 'TONUSDT': 1,
                                             'DOGEUSDT': 0, 'TRXUSDT': 0}
                                    p = p_map.get(symbol_clean, 2)
                                    raw_qty = self.config['amount'] * self.config['leverage'] / signals['price']
                                    qty = int(round(raw_qty)) if p == 0 else round(raw_qty, p)
                                    
                                    # SHORT 주문
                                    order = self.api.create_order(self.config['symbol'], 'SELL', qty)
                                    
                                    if order:
                                        # 진입 시간 기록
                                        self.config['entry_time'] = time.time()
                                        
                                        # 목표 ROI (🔥 ADX 동적 TP 사용!)
                                        target_tp_roi = dynamic_tp * self.config['leverage']

                                        # 🔥 목표 USDT 계산 (SHORT 전용!)
                                        target_usdt = self.config['amount'] * (target_tp_roi / 100)
                                        self.config['target_usdt_short'] = target_usdt  # 저장!
                                        self.config['entry_tp_short'] = dynamic_tp
                                        self.config['tp_reached_short'] = False
                                        
                                        # 🔥🔥🔥 메인넷 차트 진입가 저장!
                                        self.config['chart_entry_short'] = signals['price']
                                        
                                        # 🔥 수수료 계산 및 즉시 통계 반영!
                                        leverage = self.config['leverage']
                                        position_size = self.config['amount'] * leverage
                                        entry_fee = position_size * FEE_RATE  # 0.04%
                                        
                                        # 🔥 진입 수수료 저장 및 즉시 통계 반영!
                                        self.config['entry_fee_short'] = entry_fee
                                        self.config['stats']['short_fee'] = self.config['stats'].get('short_fee', 0) + entry_fee
                                        self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + entry_fee
                                        self.config['stats']['total_pnl'] -= entry_fee  # 진입 시점에 차감!
                                        
                                        # 🔥 통계 UI 즉시 업데이트!
                                        self.stats_callback()
                                        
                                        # 🔥 포지션 있음 플래그 설정!
                                        self.config['has_position'] = True
                                        

                                        self.log(f"✅ SHORT 진입! ${signals['price']:.2f} | {qty}개 | {self.config['timeframe']} | {self.config['leverage']}x", 'SHORT')
                                        self.log(f"   🎯 TP: 가격+{dynamic_tp:.2f}% → ROI +{target_tp_roi:.2f}% → 💰 ${target_usdt:.2f}", 'SHORT')
                                        self.log(f"   💸 진입 수수료: ${entry_fee:.2f} ({FEE_RATE*100:.2f}%) (청산 시 합산)", 'SHORT')
                                        self.log(f"   🛡️ SL: AUTO (스위칭)", 'SHORT')
                                        
                                        print(f"[✅ 진입] {self.config['symbol']} SHORT ${signals['price']:.2f} | {qty}개 | {self.config['leverage']}x")
                                        print(f"   💰 목표 USDT: ${target_usdt:.2f} | 💸 진입 수수료: ${entry_fee:.2f}")
                                        
                                        # 🔥 order response에서 평균 체결가 가져오기!
                                        filled_price = None
                                        
                                        # 1차 시도: order response에서 직접 가져오기
                                        if order and 'avgPrice' in order and order['avgPrice']:
                                            filled_price = float(order['avgPrice'])
                                            print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f}")
                                        
                                        # 2차 시도: order 정보 재조회
                                        if not filled_price and order and 'orderId' in order:
                                            try:
                                                time.sleep(0.5)
                                                order_info = self.api.get_order(self.config['symbol'], order['orderId'])
                                                if order_info and 'avgPrice' in order_info and order_info['avgPrice']:
                                                    filled_price = float(order_info['avgPrice'])
                                                    print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (재조회)")
                                            except Exception as e:
                                                print(f"[{self.config['symbol']}] order 재조회 실패: {e}")
                                        
                                        # 3차 시도: 체결 내역(fills)에서 계산
                                        if not filled_price and order and 'fills' in order and order['fills']:
                                            try:
                                                total_qty = sum(float(fill['qty']) for fill in order['fills'])
                                                total_cost = sum(float(fill['price']) * float(fill['qty']) for fill in order['fills'])
                                                if total_qty > 0:
                                                    filled_price = total_cost / total_qty
                                                    print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (fills 계산)")
                                            except Exception as e:
                                                print(f"[{self.config['symbol']}] fills 계산 실패: {e}")
                                        
                                        if not filled_price:
                                            print(f"[{self.config['symbol']}] 체결가 조회 실패 - position에서 확인 예정")
                                        
                                        # 🔥 SHORT ROI 완전 초기화 (새 진입)
                                        self.config['roi']['short_entry'] = None
                                        self.config['roi']['short_current'] = 0
                                        self.config['roi']['short_max'] = 0
                                        self.config['roi']['short_min'] = 0
                                        # 🔥 order 체결가를 restart_entry로 설정!
                                        if filled_price:
                                            self.config['restart_entry_short'] = filled_price
                                            print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (order): ${filled_price:.2f}")
                                        else:
                                            # fallback
                                            time.sleep(1)
                                            position_check = self.api.get_position(self.config['symbol'])
                                            if position_check:
                                                self.config['restart_entry_short'] = position_check['entry_price']
                                                print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (position): ${position_check['entry_price']:.2f}")
                                            else:
                                                self.config['restart_entry_short'] = signals['price']
                                                print(f"[{self.config['symbol']}] SHORT restart_entry 설정 (신호): ${signals['price']:.2f}")

                                        # 🔥🔥🔥 거래소 TP 주문 등록 (설정가 도달 시 거래소가 즉시 청산!)
                                        tp_base = filled_price or self.config.get('restart_entry_short') or signals['price']
                                        self.place_exchange_tp('SHORT', tp_base, dynamic_tp)

                                        # 🔥 청산 시간 기록 (진입 성공 후)
                                        self.config['last_close_time'] = time.time()

                                        # 10초 대기
                                        time.sleep(10)
                                    
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] 스위칭 후 SHORT 진입 실패: {e}")
                                finally:
                                    self.config["is_entering_short"] = False
                            else:
                                print(f"[{self.config['symbol']}] SHORT 봇 비활성화 - 스위칭 진입 안 함")
                            
                            # 🔥 신호 초기화 (False로 - 재진입 허용!)
                            self.config["prev_signals"] = {
                                "ut_long": False,
                                "ut_short": False,  # False로!
                                "ema_long": False,
                                "ema_short": False,
                                "long_signal": False,
                                "short_signal": False  # False로!
                            }
                            
                            continue
                    
                    else:  # short
                        # 🔥 진입 시 저장한 동적 TP 사용! (ADX 기반)
                        user_tp = self.config.get('entry_tp_short') or self.config.get('tp', 0.3)
                        target_roi = user_tp * leverage
                        
                        # ROI >= 목표면 즉시 익절
                        if pnl_pct >= target_roi and not self.config.get("is_closing"):
                            print(f"[🎯 익절] {self.config['symbol']} SHORT: ROI {pnl_pct:.2f}% >= 목표 {target_roi:.2f}% (TP {user_tp}%)")
                            
                            self.config["is_closing"] = True
                            self.config["is_closing_start"] = time.time()

                            close_success = False
                            for attempt in range(3):
                                try:
                                    result = self.api.close_position(self.config['symbol'])
                                    if result:
                                        close_success = True
                                        print(f"[✅ 청산 성공] {self.config['symbol']} SHORT")
                                        break
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] SHORT 청산 시도 {attempt+1}/3 실패: {e}")
                                time.sleep(1)
                            
                            if not close_success:
                                self.log(f"❌ 청산 실패! 수동으로 청산하세요!", 'SHORT')
                                print(f"[❌ 청산 실패] {self.config['symbol']} SHORT - 수동 청산 필요!")
                                self.config["is_closing"] = False
                                continue
                            
                            # 🔥 수수료 계산
                            position_size = self.config['amount'] * self.config['leverage']
                            entry_fee = self.config.get('entry_fee_short', position_size * FEE_RATE)
                            close_fee = position_size * FEE_RATE
                            total_fee = entry_fee + close_fee
                            net_profit = pnl_usd - total_fee
                            
                            # 통계
                            self.config['stats']['short_fee'] = self.config['stats'].get('short_fee', 0) + total_fee
                            self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + total_fee
                            
                            close_price = current
                            
                            self.config['stats']['short_count'] += 1
                            self.config['stats']['short_win'] += 1
                            self.config['stats']['short_profit'] += net_profit
                            self.config['stats']['total_pnl'] += net_profit
                            
                            profit_sign = '+' if net_profit >= 0 else ''
                            count = self.config['stats']['short_count']
                            
                            # 엑셀 저장
                            self.save_trade_excel('SHORT', pnl_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, '익절')
                            
                            # 로그
                            self.log(f"✅ SHORT 익절! ROI +{pnl_pct:.2f}% (목표 {target_roi:.2f}%) (#{count})", 'SHORT')
                            self.log(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f}", 'SHORT')
                            self.log(f"   💸 수수료: ${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}", 'SHORT')
                            
                            print(f"[✅ 익절] {self.config['symbol']} SHORT ROI +{pnl_pct:.2f}% (목표 {target_roi:.2f}%)")
                            
                            self.stats_callback()
                            self.config['last_close_time'] = time.time()
                            
                            # 신호 초기화
                            self.config['prev_signals'] = {
                                'ut_long': False, 'ut_short': False,
                                'ema_long': False, 'ema_short': False,
                                'long_signal': False, 'short_signal': False
                            }
                            
                            # ROI 초기화
                            self.config['entry_tp_short'] = None
                            self.config['chart_entry_short'] = None
                            self.config['entry_fee_short'] = 0
                            self.config['roi']['short_entry'] = None
                            self.config['roi']['short_current'] = 0
                            self.config['roi']['short_max'] = 0
                            self.config['roi']['short_min'] = 0
                            
                            print(f"[{self.config['symbol']}] SHORT 익절 완료 - 재진입 대기")
                            self.config["is_closing"] = False
                            continue
                        
                        # 🔥 손절: AUTO(스위칭) - UT Bot 현재 상태 + EMA 크로스
                        auto_sl = signals.get('ut_position_long', False) and signals['ema_long']
                        

                        # 🔍 디버깅: 신호 상태 출력 (신호 있을 때만)
                        if signals.get('ut_position_long', False) or signals['ema_long']:
                            if not self.config.get("is_closing"):
                                print(f"[DEBUG] {self.config['symbol']} SHORT 포지션 - 스위칭 신호 체크:")
                                print(f"   UT Bot 상태: position_long={signals.get('ut_position_long', False)}")
                                print(f"   EMA LONG: {signals['ema_long']}")
                                print(f"   → auto_sl (AND): {auto_sl}")
                        
                        if auto_sl and not self.config.get("is_closing"):
                            print(f"")
                            print(f"[🔄 스위칭 실행!] {self.config['symbol']} SHORT→LONG")
                            print(f"   ✅ UT Bot 상태: LONG={signals.get('ut_position_long', False)}")
                            print(f"   ✅ EMA LONG: {signals['ema_long']}")
                            print(f"   ✅ AND 조건: True")
                            print(f"")
                            
                            # 청산 시작 플래그
                            self.config["is_closing"] = True
                            self.config["is_closing_start"] = time.time()

                            # 🔥 청산 재시도 (최대 3번)
                            close_success = False
                            for attempt in range(3):
                                try:
                                    result = self.api.close_position(self.config['symbol'])
                                    if result:
                                        close_success = True
                                        break
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] SHORT 스위칭 청산 시도 {attempt+1}/3 실패: {e}")
                                time.sleep(1)
                            
                            if not close_success:
                                self.log(f"❌ 스위칭 청산 실패! 수동으로 청산하세요!", 'SHORT')
                                print(f"[❌ 스위칭 실패] {self.config['symbol']} SHORT→LONG - 수동 청산 필요!")
                                self.config["is_closing"] = False
                                continue
                            
                            # 🔥 가격 변동률 계산 (SHORT: 진입가 - 현재가)
                            price_change_pct = ((entry - current) / entry) * 100
                            
                            # 🔥 수수료 계산
                            position_size = self.config['amount'] * self.config['leverage']
                            entry_fee = self.config.get('entry_fee_short', position_size * FEE_RATE)
                            close_fee = position_size * FEE_RATE
                            total_fee = entry_fee + close_fee
                            net_profit = pnl_usd - total_fee
                            
                            self.config['stats']['short_fee'] = self.config['stats'].get('short_fee', 0) + total_fee
                            self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + total_fee
                            
                            close_price = current
                            
                            self.config['stats']['short_count'] += 1
                            if net_profit > 0:
                                self.config['stats']['short_win'] += 1
                            else:
                                self.config['stats']['short_loss'] += 1
                            self.config['stats']['short_profit'] += net_profit
                            self.config['stats']['total_pnl'] += net_profit
                            
                            profit_sign = '+' if net_profit >= 0 else ''
                            
                            # 🔥 엑셀 저장!
                            self.save_trade_excel('SHORT', pnl_pct, pnl_usd, entry_fee, close_fee, total_fee, net_profit, '스위칭')
                            
                            # 🔥 로그에 청산 가격, 수수료, 순수익 표시
                            self.log(f"🔄 SHORT→LONG 스위칭! 가격{'+' if price_change_pct >= 0 else ''}{price_change_pct:.2f}% | ROI {'+' if pnl_pct >= 0 else ''}{pnl_pct:.2f}%", 'SHORT')
                            self.log(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f}", 'SHORT')
                            self.log(f"   💸 총 수수료: -${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}", 'SHORT')
                            
                            # 콘솔 출력
                            print(f"[🔄 스위칭] {self.config['symbol']} SHORT→LONG {'+' if price_change_pct >= 0 else ''}{price_change_pct:.2f}%")
                            print(f"   💰 청산가: ${close_price:.2f} | 수익: ${pnl_usd:.2f} | 총 수수료: -${total_fee:.2f} | 순수익: {profit_sign}${abs(net_profit):.2f}")
                            
                            # 통계 업데이트
                            self.stats_callback()
                            
                            # 🔥 즉시 LONG 진입 (스위칭)
                            self.log(f"🔄 즉시 LONG 진입 준비...", 'LONG')
                            print(f"[{self.config['symbol']}] SHORT→LONG 스위칭: 즉시 LONG 진입!")
                            
                            # 🔥 SHORT ROI 초기화 (청산됨)
                            self.config['roi']['short_entry'] = None
                            self.config['roi']['short_current'] = 0
                            self.config['roi']['short_max'] = 0
                            self.config['roi']['short_min'] = 0
                            self.config['restart_entry_short'] = None  # 재시작 price도 초기화
                            self.config['target_usdt_short'] = 0  # 목표 USDT 초기화!
                            self.config['entry_tp_short'] = None
                            self.config['tp_reached_short'] = False
                            self.config['entry_fee_short'] = 0  # 진입 수수료 초기화
                            
                            # 5초 대기
                            time.sleep(5)
                            
                            # 청산 완료 - 플래그 해제
                            self.config["warned_position_exists"] = False
                            self.config["is_closing"] = False
                            
                            # 🔥 스위칭은 역신호만 있어도 진입! (2개 조건 불필요)
                            # LONG 봇이 활성화되어 있으면 즉시 진입
                            if self.config['long_active']:
                                try:
                                    # 🔒 진입 플래그 설정 - Lock으로 보호!
                                    with ENTRY_LOCK:
                                        if self.config.get('is_entering_long') or self.config.get('is_entering_short'):
                                            print(f"[{self.config['symbol']}] 스위칭 진입 차단: 다른 진입 중")
                                            continue
                                        self.config["is_entering_long"] = True
                                        self.config["is_entering_long_start"] = time.time()
                                    
                                    # 포지션 재확인
                                    double_check = self.api.get_position(self.config['symbol'])
                                    if double_check:
                                        print(f"[{self.config['symbol']}] 스위칭 실패: 이미 포지션 있음")
                                        self.config["is_entering_long"] = False
                                        continue
                                    
                                    # 소수점 처리 (바이낸스 실제 기준)
                                    symbol_clean = self.config['symbol'].replace('/', '')
                                    p_map = {'BTCUSDT': 3, 'ETHUSDT': 3, 'BNBUSDT': 2, 'LINKUSDT': 2,
                                             'SOLUSDT': 1, 'XRPUSDT': 1, 'ADAUSDT': 1, 'TONUSDT': 1,
                                             'DOGEUSDT': 0, 'TRXUSDT': 0}
                                    p = p_map.get(symbol_clean, 2)
                                    raw_qty = self.config['amount'] * self.config['leverage'] / signals['price']
                                    qty = int(round(raw_qty)) if p == 0 else round(raw_qty, p)
                                    
                                    # LONG 주문
                                    order = self.api.create_order(self.config['symbol'], 'BUY', qty)
                                    
                                    if order:
                                        # 진입 시간 기록
                                        self.config['entry_time'] = time.time()
                                        
                                        # 목표 ROI (🔥 ADX 동적 TP 사용!)
                                        target_tp_roi = dynamic_tp * self.config['leverage']

                                        # 🔥 목표 USDT 계산 (LONG 전용!)
                                        target_usdt = self.config['amount'] * (target_tp_roi / 100)
                                        self.config['target_usdt_long'] = target_usdt  # 저장!
                                        self.config['entry_tp_long'] = dynamic_tp
                                        self.config['tp_reached_long'] = False
                                        
                                        # 🔥🔥🔥 메인넷 차트 진입가 저장!
                                        self.config['chart_entry_long'] = signals['price']
                                        
                                        # 🔥 수수료 계산 및 즉시 통계 반영!
                                        leverage = self.config['leverage']
                                        position_size = self.config['amount'] * leverage
                                        entry_fee = position_size * FEE_RATE  # 0.04%
                                        
                                        # 🔥 진입 수수료 저장 및 즉시 통계 반영!
                                        self.config['entry_fee_long'] = entry_fee
                                        self.config['stats']['long_fee'] = self.config['stats'].get('long_fee', 0) + entry_fee
                                        self.config['stats']['total_fee'] = self.config['stats'].get('total_fee', 0) + entry_fee
                                        self.config['stats']['total_pnl'] -= entry_fee  # 진입 시점에 차감!
                                        
                                        # 🔥 통계 UI 즉시 업데이트!
                                        self.stats_callback()
                                        
                                        # 🔥 포지션 있음 플래그 설정!
                                        self.config['has_position'] = True
                                        

                                        self.log(f"✅ LONG 진입! ${signals['price']:.2f} | {qty}개 | {self.config['timeframe']} | {self.config['leverage']}x", 'LONG')
                                        self.log(f"   🎯 TP: 가격+{dynamic_tp:.2f}% → ROI +{target_tp_roi:.2f}% → 💰 ${target_usdt:.2f}", 'LONG')
                                        self.log(f"   💸 진입 수수료: ${entry_fee:.2f} ({FEE_RATE*100:.2f}%) (청산 시 합산)", 'LONG')
                                        self.log(f"   🛡️ SL: AUTO (스위칭)", 'LONG')
                                        
                                        print(f"[✅ 진입] {self.config['symbol']} LONG ${signals['price']:.2f} | {qty}개 | {self.config['leverage']}x")
                                        print(f"   💰 목표 USDT: ${target_usdt:.2f} | 💸 진입 수수료: ${entry_fee:.2f}")
                                        
                                        # 🔥 order response에서 평균 체결가 가져오기!
                                        filled_price = None
                                        
                                        # 1차 시도: order response에서 직접 가져오기
                                        if order and 'avgPrice' in order and order['avgPrice']:
                                            filled_price = float(order['avgPrice'])
                                            print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f}")
                                        
                                        # 2차 시도: order 정보 재조회
                                        if not filled_price and order and 'orderId' in order:
                                            try:
                                                time.sleep(0.5)
                                                order_info = self.api.get_order(self.config['symbol'], order['orderId'])
                                                if order_info and 'avgPrice' in order_info and order_info['avgPrice']:
                                                    filled_price = float(order_info['avgPrice'])
                                                    print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (재조회)")
                                            except Exception as e:
                                                print(f"[{self.config['symbol']}] order 재조회 실패: {e}")
                                        
                                        # 3차 시도: 체결 내역(fills)에서 계산
                                        if not filled_price and order and 'fills' in order and order['fills']:
                                            try:
                                                total_qty = sum(float(fill['qty']) for fill in order['fills'])
                                                total_cost = sum(float(fill['price']) * float(fill['qty']) for fill in order['fills'])
                                                if total_qty > 0:
                                                    filled_price = total_cost / total_qty
                                                    print(f"[{self.config['symbol']}] 체결 가격: ${filled_price:.2f} (fills 계산)")
                                            except Exception as e:
                                                print(f"[{self.config['symbol']}] fills 계산 실패: {e}")
                                        
                                        if not filled_price:
                                            print(f"[{self.config['symbol']}] 체결가 조회 실패 - position에서 확인 예정")
                                        
                                        # 🔥 LONG ROI 완전 초기화 (새 진입)
                                        self.config['roi']['long_entry'] = None
                                        self.config['roi']['long_current'] = 0
                                        self.config['roi']['long_max'] = 0
                                        self.config['roi']['long_min'] = 0
                                        # 🔥 order 체결가를 restart_entry로 설정!
                                        if filled_price:
                                            self.config['restart_entry_long'] = filled_price
                                            print(f"[{self.config['symbol']}] LONG restart_entry 설정 (order): ${filled_price:.2f}")
                                        else:
                                            # fallback
                                            time.sleep(1)
                                            position_check = self.api.get_position(self.config['symbol'])
                                            if position_check:
                                                self.config['restart_entry_long'] = position_check['entry_price']
                                                print(f"[{self.config['symbol']}] LONG restart_entry 설정 (position): ${position_check['entry_price']:.2f}")
                                            else:
                                                self.config['restart_entry_long'] = signals['price']
                                                print(f"[{self.config['symbol']}] LONG restart_entry 설정 (신호): ${signals['price']:.2f}")

                                        # 🔥🔥🔥 거래소 TP 주문 등록 (설정가 도달 시 거래소가 즉시 청산!)
                                        tp_base = filled_price or self.config.get('restart_entry_long') or signals['price']
                                        self.place_exchange_tp('LONG', tp_base, dynamic_tp)

                                        # 🔥 청산 시간 기록 (진입 성공 후)
                                        self.config['last_close_time'] = time.time()

                                        # 10초 대기
                                        time.sleep(10)
                                    
                                except Exception as e:
                                    print(f"[{self.config['symbol']}] 스위칭 후 LONG 진입 실패: {e}")
                                finally:
                                    self.config["is_entering_long"] = False
                            else:
                                print(f"[{self.config['symbol']}] LONG 봇 비활성화 - 스위칭 진입 안 함")
                            
                            # 🔥 신호 초기화 (False로 - 재진입 허용!)
                            self.config["prev_signals"] = {
                                "ut_long": False,  # False로!
                                "ut_short": False,
                                "ema_long": False,
                                "ema_short": False,
                                "long_signal": False,  # False로!
                                "short_signal": False
                            }
                            
                            continue
                
                # 🔥 포지션 있을 때는 익절/손절만 처리 (신호 업데이트 안 함!)
                # 익절/손절 후 prev_signals가 False로 초기화되므로 재진입 가능
            
            except Exception as e:
                # 에러는 양쪽 로그에 모두 표시
                error_msg = f"❌ 오류: {str(e)}"
                try:
                    if self.config.get('long_active'):
                        self.log(error_msg, 'LONG')
                    if self.config.get('short_active'):
                        self.log(error_msg, 'SHORT')
                except:
                    print(error_msg)

# ==================== GUI ====================
STATS_FILE = "bot_stats.json"  # 🔥 통계 저장 파일

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("🤖 자동매매 봇")
        self.root.geometry("1600x900")
        self.root.configure(bg='#1e1e1e')
        
        self.api = BinanceAPI(API_KEY, API_SECRET, TESTNET)
        self.coins = []
        self.bots = {}
        self.tab_buttons = {}
        self.program_number = None
        self._is_starting = False  # 🔥 시작/재연결 중복 방지
        
        # 🔥🔥🔥 4H UT Bot 필터 (전역 설정)
        self.filter_4h_enabled = False  # OFF 고정
        self.filter_4h_cache = {}  # 코인별 4H 신호 캐시
        self.filter_4h_last_update = 0  # 마지막 업데이트 시간
        
        # 전체 통계 (파일에서 불러오기 시도)
        self.global_stats = {
            'total_pnl': 0,
            'total_trades': 0,
            'total_wins': 0,
            'total_fee': 0,
        }
        
        print("=" * 60)
        print("📊 통계: 파일에서 불러오기 시도...")
        print("📝 거래 기록: trade_history.xlsx에 저장됨")
        print("=" * 60)
        
        self.create_ui()
        self.add_default_coins()  # 기본 10개 코인 추가!
        
        # 🔥 저장된 통계 불러오기!
        self.load_stats()
        
        # 🔥 엑셀 파일 초기화
        self.init_trade_history_excel()
        
        self.update_balance()
        
        # 🔥 4H UT 필터 비활성화
        # self.update_4h_filter()
        
        # 🔥 X 버튼 핸들러 등록
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 🔥 자동 저장 타이머 (5분마다)
        self.auto_save_stats()
        
        # 🔥 자동 재연결: bot_running.json 있으면 자동으로 이어받기
        BOT_RUNNING_FILE = 'bot_running.json'
        if os.path.exists(BOT_RUNNING_FILE):
            print("\n" + "=" * 60)
            print("🔄 이전 세션 발견! 자동 재연결 중...")
            print("=" * 60)
            # 연결 테스트
            balance = self.api.get_balance()
            if balance > 0:
                mode = "테스트넷" if self.api.testnet else "메인넷"
                print(f"💰 잔고: ${balance:,.2f} | 모드: {mode}")
                self.root.after(500, lambda: self._auto_reconnect())
            else:
                print("⚠️ API 연결 실패 — 수동 시작 필요")
                os.remove(BOT_RUNNING_FILE)
                self._show_connection_info()
        else:
            self._show_connection_info()
    
    def _show_connection_info(self):
        """연결 테스트 + 안내 팝업"""
        balance = self.api.get_balance()
        mode = "테스트넷" if self.api.testnet else "메인넷"
        chart_source = "메인넷 (실시간)"
        
        CustomMessageBox.showinfo("API 연결", 
            f"✅ 연결 성공!\n\n"
            f"💰 잔고: ${balance:,.2f} USDT\n"
            f"🔧 주문 모드: {mode}\n"
            f"📊 차트 데이터: {chart_source}\n\n"
            f"🔗 선물 차트 확인:\n"
            f"https://www.binance.com/en/futures/ADAUSDT\n\n"
            f"🪙 기본 10개 코인 추가됨!"
        )
    
    def _auto_reconnect(self):
        """자동 재연결 — 이전 세션 이어받기"""
        print("🔄 자동 재연결 시작...")
        threading.Thread(target=self._reconnect_all_thread, daemon=True).start()
    
    def _save_running_state(self):
        """봇 실행 상태 저장 — 프로그램 재시작 시 자동 재연결용"""
        try:
            state = {'running': True, 'timestamp': time.time()}
            with open('bot_running.json', 'w') as f:
                json.dump(state, f)
        except:
            pass
    
    def _clear_running_state(self):
        """봇 종료 상태 — 다음 시작 시 새로 시작"""
        try:
            if os.path.exists('bot_running.json'):
                os.remove('bot_running.json')
        except:
            pass
    
    def select_program_number(self):
        """프로그램 #1 고정"""
        prog_num = 1
        self.program_number = prog_num
        self.root.title(f"🤖 자동매매 봇 - 프로그램 #{prog_num}")
        return prog_num
    
    def add_default_coins(self):
        """프로그램 번호에 따라 10개 코인 자동 추가"""
        
        # 🪙 바이낸스 선물 거래 가능한 100개 코인 (2025년 1월 기준)
        all_coins = [
            # 프로그램 #1 (1-10) - 메이저 코인
            {'symbol': 'BTC/USDT', 'name': 'Bitcoin'},
            {'symbol': 'ETH/USDT', 'name': 'Ethereum'},
            {'symbol': 'BNB/USDT', 'name': 'BNB'},
            {'symbol': 'SOL/USDT', 'name': 'Solana'},
            {'symbol': 'XRP/USDT', 'name': 'Ripple'},
            {'symbol': 'ADA/USDT', 'name': 'Cardano'},
            {'symbol': 'DOGE/USDT', 'name': 'Dogecoin'},
            {'symbol': 'TRX/USDT', 'name': 'TRON'},
            {'symbol': 'TON/USDT', 'name': 'Toncoin'},
            {'symbol': 'LINK/USDT', 'name': 'Chainlink'},
        ]
        
        # 프로그램 번호 선택
        prog_num = self.select_program_number()
        
        # 해당 프로그램의 10개 코인 선택 (이미 10개만 있음)
        selected_coins = all_coins  # 전체 사용
        
        print("=" * 60)
        print(f"🚀 프로그램 #{prog_num} 시작!")
        print(f"📌 담당 코인: 1번 ~ 10번")
        print("=" * 60)
        
        for i, coin_info in enumerate(selected_coins, 1):
            coin = {
                'symbol': coin_info['symbol'],
                'timeframe': '15m',  # 15분봉
                'amount': 50,  # 진입금 50 USDT
                'leverage': 3,  # 레버리지 3배 (안전)
                'tp': 0.3,  # 기본 TP (동적 TP 비활성화 시)
                'sl': 0,  # AUTO 고정
                'ut_sens': 10,  # 🔥 UT Bot Key Value
                'ut_atr': 2,   # 🔥 UT Bot ATR Period
                'ema_fast': 34,
                'ema_slow': 55,
                'long_active': False,
                'short_active': False,
                # 🔥 동적 TP 설정
                'tp_trend': 1.2,  # ✅ 추세장 TP 1.2% (ADX >= 21)
                'tp_sideways': 1.0,  # ✅ 횡보장 TP 1.0% (ADX < 21)
                'adx_period': 10,  # ADX 기간
                'adx_interval': '1h',  # 🔥 ADX 계산 시간봉 (TP 결정용)
                # 🔥 거래량 필터 비활성화
                'volume_filter_enabled': False,  # 거래량 필터 비활성화
                'volume_multiplier': 1.0,  # 사용 안 함
            }
            
            self.add_coin(coin)
            num = (prog_num - 1) * 10 + i
            print(f"  #{num:3d}. {coin_info['symbol']:15s} ({coin_info['name']})")
        
        print("=" * 60)
        print(f"🪙 프로그램 #{prog_num}: 10개 코인 로드 완료!")
        print("=" * 60)
        print()
    
    def create_ui(self):
        # 상단 헤더
        header = tk.Frame(self.root, bg='#2d2d2d', height=80)
        header.pack(fill='x')
        
        # 프로그램 번호 표시
        prog_text = f"🤖 자동매매 봇 - 프로그램 #{self.program_number}" if self.program_number else "🤖 자동매매 봇"
        tk.Label(header, text=prog_text, font=('Arial', 16, 'bold'),
                bg='#2d2d2d', fg='#ffffff').pack(side='left', padx=20, pady=10)
        
        # 전체 통계
        stats_frame = tk.Frame(header, bg='#2d2d2d')
        stats_frame.pack(side='right', padx=20, pady=10)
        
        self.global_pnl_label = tk.Label(stats_frame, text="📊 총 수익: $0.00", 
                                         font=('Arial', 12, 'bold'), bg='#2d2d2d', fg='#00ff00')
        self.global_pnl_label.pack(side='top')
        
        # 🔥 수수료 라벨 (빨간색)
        self.global_fee_label = tk.Label(stats_frame, text="💸 수수료: -$0.00",
                                         font=('Arial', 10), bg='#2d2d2d', fg='#ff6666')
        self.global_fee_label.pack(side='top')
        
        self.global_trades_label = tk.Label(stats_frame, text="거래: 0회 | 승률: 0%",
                                            font=('Arial', 10), bg='#2d2d2d', fg='#aaaaaa')
        self.global_trades_label.pack(side='top')
        
        # 잔고 표시
        self.balance_label = tk.Label(header, text="잔고: $0.00", font=('Arial', 14, 'bold'),
                                     bg='#2d2d2d', fg='#00ff00')
        self.balance_label.pack(side='right', padx=20, pady=10)
        
        # 탭 프레임
        self.tab_frame = tk.Frame(self.root, bg='#2d2d2d')
        self.tab_frame.pack(fill='x', padx=10, pady=5)
        
        # 전체 제어 버튼들 (크기 통일)
        tk.Button(self.tab_frame, text="▶️ 전체 롱숏 시작", command=self.start_all,
                 bg='#0066cc', fg='#ffffff', font=('Arial', 11, 'bold'),
                 width=14, pady=5).pack(side='left', padx=5)
        
        tk.Button(self.tab_frame, text="🔄 재연결", command=self.reconnect_all,
                 bg='#00aa00', fg='#ffffff', font=('Arial', 11, 'bold'),
                 width=14, pady=5).pack(side='left', padx=5)
        
        tk.Button(self.tab_frame, text="⏹️ 전체 롱숏 종료", command=self.stop_all,
                 bg='#cc0000', fg='#ffffff', font=('Arial', 11, 'bold'),
                 width=14, pady=5).pack(side='left', padx=5)
        
        # 메인 컨텐츠
        self.main_frame = tk.Frame(self.root, bg='#1e1e1e')
        self.main_frame.pack(fill='both', expand=True)
    
    def toggle_4h_filter(self):
        """🔥 4H UT 필터 ON/OFF 토글"""
        self.filter_4h_enabled = not self.filter_4h_enabled

        if not hasattr(self, 'filter_4h_btn'):
            print(f"[4H UT 필터] {'ON' if self.filter_4h_enabled else 'OFF'}")
            return

        if self.filter_4h_enabled:
            self.filter_4h_btn.config(text="📊 4H필터: ON", bg='#9900cc')
            print("[4H UT 필터] ON - 4시간봉 방향과 같은 방향만 진입!")
        else:
            self.filter_4h_btn.config(text="📊 4H필터: OFF", bg='#666666')
            print("[4H UT 필터] OFF - 모든 방향 진입 허용!")
    
    def start_all(self):
        """전체 롱숏 시작 - 포지션 무시하고 새로 시작"""
        
        # 🔥🔥🔥 디버그: 이 로그가 안 나오면 다른 파일이 실행 중!
        print("\n" + "=" * 80)
        print("🔥🔥🔥 [NEW VERSION 2025-01-11 23:50] start_all() 실행! 🔥🔥🔥")
        print("🔥🔥🔥 이 로그가 보이면 최신 파일입니다! 🔥🔥🔥")
        print("🔥🔥🔥 안 보이면 캐시 삭제 후 다시 시작하세요! 🔥🔥🔥")
        print("=" * 80 + "\n")
        
        prog_info = f"프로그램 #{self.program_number}: " if self.program_number else ""
        if CustomMessageBox.askyesno("전체 시작", 
            f"{prog_info}모든 코인의 롱과 숏을 시작하시겠습니까?\n\n"
            f"대상: {len(self.coins)}개 코인\n"
            f"총 {len(self.coins) * 2}개 봇 시작\n\n"
            f"⚠️ 기존 포지션은 무시하고 새로 시작합니다.\n"
            f"⚠️ 포지션 재연결은 '재연결' 버튼을 사용하세요.\n\n"
            f"💡 순차적으로 시작되므로 약 {len(self.coins)}초 소요됩니다."):
            
            # 🔥 스레드에서 실행 (GUI 멈춤 방지!)
            threading.Thread(target=self._start_all_thread, daemon=True).start()
    
    def _start_all_thread(self):
        if self._is_starting:
            print("[⚠️] 이미 시작/재연결 중!")
            return
        self._is_starting = True
        try:
            count = 0
            for i, coin in enumerate(self.coins, 1):
                print(f"\n[전체 시작] {i}/{len(self.coins)}: {coin['symbol']}")
                print(f"  🔍 [DEBUG] 코인 처리 시작: {coin['symbol']}")
            
                # 🔥🔥🔥 LONG 완전 초기화! (active 상태 무관)
                print(f"  🔍 [DEBUG] LONG ROI 초기화 시작...")
            
                # 🔥 ROI 관련 모든 값 초기화
                coin['roi']['long_entry'] = None
                coin['roi']['long_current'] = 0
                coin['roi']['long_max'] = 0
                coin['roi']['long_min'] = 0
            
                # 🔥 진입가 관련 모든 값 초기화 (이전 값 절대 사용 안함!)
                coin['chart_entry_long'] = None
                coin['restart_entry_long'] = None
                coin['entry_fee_long'] = 0
            
                # 🔥 TP 관련 초기화
                coin['target_usdt_long'] = 0
                coin['entry_tp_long'] = None
                coin['tp_reached_long'] = False
            
                print(f"  ✅ [DEBUG] LONG 완전 초기화 완료!")
            
                # 🔥 UI 즉시 업데이트 (화면에 - 표시!)
                print(f"  🔍 [DEBUG] UI 업데이트 시작...")
                try:
                    if 'labels' in coin and 'long_current_roi' in coin['labels']:
                        try:
                            self.root.after(0, lambda c=coin: (
                                c['labels']['long_current_roi'].config(text="현재: -", fg='#ffffff'),
                                c['labels']['long_max_roi'].config(text="최고: -"),
                                c['labels']['long_min_roi'].config(text="최저: -")
                            ))
                        except: pass
                        print(f"  ✅ {coin['symbol']} LONG ROI → - (UI 업데이트)")
                    else:
                        print(f"  ⚠️ [DEBUG] labels 없음 또는 long_current_roi 없음")
                except Exception as e:
                    print(f"  ⚠️ {coin['symbol']} LONG UI 업데이트 실패: {e}")
            
                # LONG 시작 (무조건 재시작!)
                print(f"  🔍 [DEBUG] LONG 봇 재시작 시작...")
                # 기존 봇 정지
                if id(coin) in self.bots and 'long' in self.bots[id(coin)]:
                    print(f"  🔍 [DEBUG] 기존 LONG 봇 발견! 정지 시작...")
                    self.bots[id(coin)]['long'].stop()
                    time.sleep(1.5)  # 스레드 완전 종료 대기
                    del self.bots[id(coin)]['long']
                    print(f"  🛑 {coin['symbol']} 기존 LONG 봇 정지")
                else:
                    print(f"  🔍 [DEBUG] 기존 LONG 봇 없음 (새로 시작)")
            
                coin['long_active'] = True
                coin['last_close_time'] = None
                coin['is_entering_long'] = False; coin['is_entering_short'] = False
                coin['is_closing'] = False
                coin['warned_position_exists'] = False
            
                # 🔥 신호를 None으로 초기화
                coin['prev_signals'] = {
                    'ut_long': None,
                    'ut_short': None,
                    'ema_long': None,
                    'ema_short': None,
                    'long_signal': None,
                    'short_signal': None
                }
            
                if id(coin) not in self.bots:
                    self.bots[id(coin)] = {}
            
                # 🔥 완전히 새 봇 생성!
                self.bots[id(coin)]['long'] = TradingBot(
                    self.api, coin,
                    lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                    lambda c=coin: self.update_stats(c),
                    self.save_trade_to_excel,
                    bot_type='long',
                    app=self  # 🔥 App 참조 전달
                )
                self.bots[id(coin)]['long'].start()
                count += 1
            
                # 🔥 시작 로그
                self.add_log(coin, 'LONG', "=" * 40)
                self.add_log(coin, 'LONG', "✅ LONG 봇 시작!")
                self.add_log(coin, 'LONG', f"💰 잔고: ${self.api.get_balance():,.2f}")
                self.add_log(coin, 'LONG', "📭 새로 시작 (포지션 무시)")
            
                # 통계 정보
                s = coin['stats']
                self.add_log(coin, 'LONG', f"📊 LONG 통계:")
                self.add_log(coin, 'LONG', f"   거래: {s['long_count']}회 | 익절: {s['long_win']}회 | 손절: {s['long_loss']}회")
                self.add_log(coin, 'LONG', f"   수익: ${s['long_profit']:+,.2f}")
                self.add_log(coin, 'LONG', f"📊 총 통계: ${s['total_pnl']:+,.2f}")
                self.add_log(coin, 'LONG', "=" * 40)
            
                # 🔥🔥🔥 SHORT 완전 초기화! (active 상태 무관)
                print(f"  🔍 [DEBUG] SHORT ROI 초기화 시작...")
            
                # 🔥 ROI 관련 모든 값 초기화
                coin['roi']['short_entry'] = None
                coin['roi']['short_current'] = 0
                coin['roi']['short_max'] = 0
                coin['roi']['short_min'] = 0
            
                # 🔥 진입가 관련 모든 값 초기화 (이전 값 절대 사용 안함!)
                coin['chart_entry_short'] = None
                coin['restart_entry_short'] = None
                coin['entry_fee_short'] = 0
            
                # 🔥 TP 관련 초기화
                coin['target_usdt_short'] = 0
                coin['entry_tp_short'] = None
                coin['tp_reached_short'] = False
            
                print(f"  ✅ [DEBUG] SHORT 완전 초기화 완료!")
            
                # 🔥 UI 즉시 업데이트 (화면에 - 표시!)
                print(f"  🔍 [DEBUG] SHORT UI 업데이트 시작...")
                try:
                    if 'labels' in coin and 'short_current_roi' in coin['labels']:
                        try:
                            self.root.after(0, lambda c=coin: (
                                c['labels']['short_current_roi'].config(text="현재: -", fg='#ffffff'),
                                c['labels']['short_max_roi'].config(text="최고: -"),
                                c['labels']['short_min_roi'].config(text="최저: -")
                            ))
                        except: pass
                        print(f"  ✅ {coin['symbol']} SHORT ROI → - (UI 업데이트)")
                    else:
                        print(f"  ⚠️ [DEBUG] labels 없음 또는 short_current_roi 없음")
                except Exception as e:
                    print(f"  ⚠️ {coin['symbol']} SHORT UI 업데이트 실패: {e}")
            
                # SHORT 시작 (무조건 재시작!)
                print(f"  🔍 [DEBUG] SHORT 봇 재시작 시작...")
                # 기존 봇 정지
                if id(coin) in self.bots and 'short' in self.bots[id(coin)]:
                    print(f"  🔍 [DEBUG] 기존 SHORT 봇 발견! 정지 시작...")
                    self.bots[id(coin)]['short'].stop()
                    time.sleep(1.5)  # 스레드 완전 종료 대기
                    del self.bots[id(coin)]['short']
                    print(f"  🛑 {coin['symbol']} 기존 SHORT 봇 정지")
                else:
                    print(f"  🔍 [DEBUG] 기존 SHORT 봇 없음 (새로 시작)")
            
                coin['short_active'] = True
            
                # 🔥 완전히 새 봇 생성!
                self.bots[id(coin)]['short'] = TradingBot(
                    self.api, coin,
                    lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                    lambda c=coin: self.update_stats(c),
                    self.save_trade_to_excel,
                    bot_type='short',
                    app=self  # 🔥 App 참조 전달
                )
                self.bots[id(coin)]['short'].start()
                count += 1
            
                # 🔥 시작 로그
                self.add_log(coin, 'SHORT', "=" * 40)
                self.add_log(coin, 'SHORT', "✅ SHORT 봇 시작!")
                self.add_log(coin, 'SHORT', f"💰 잔고: ${self.api.get_balance():,.2f}")
                self.add_log(coin, 'SHORT', "📭 새로 시작 (포지션 무시)")
            
                # 통계 정보
                s = coin['stats']
                self.add_log(coin, 'SHORT', f"📊 SHORT 통계:")
                self.add_log(coin, 'SHORT', f"   거래: {s['short_count']}회 | 익절: {s['short_win']}회 | 손절: {s['short_loss']}회")
                self.add_log(coin, 'SHORT', f"   수익: ${s['short_profit']:+,.2f}")
                self.add_log(coin, 'SHORT', f"📊 총 통계: ${s['total_pnl']:+,.2f}")
                self.add_log(coin, 'SHORT', "=" * 40)
            
                # ⚠️ LONG ROI는 이미 위에서 초기화됨 (재초기화하지 않음!)
                # 포지션이 있으면 이미 restart_entry_long과 roi['long_entry']가 설정됨
            
                # 🔥 각 코인마다 0.5초 대기 (API 부하 분산)
                if i < len(self.coins):
                    time.sleep(0.5)
        
            # 🔥🔥🔥 봇들이 첫 신호 체크하기 전에 대기! (중요!)
            print("\n" + "=" * 60)
            print("⏰ 5초 대기 중... (봇 초기화 완료 대기)")
            print("=" * 60)
            time.sleep(5)  # 봇들이 첫 신호 체크하기 전에 대기
        
            # 🔥🔥🔥 모든 봇 시작 후 UI 강제 업데이트!
            print("\n" + "=" * 60)
            print("🔄 UI 강제 업데이트 시작...")
            print("=" * 60)
            for coin in self.coins:
                try:
                    # 🔥 포지션 확인
                    position = self.api.get_position(coin['symbol'])
                
                    if position:
                        # 포지션 있으면 실시간 ROI 표시
                        print(f"  🔍 {coin['symbol']} 포지션: {position['side']} entry=${position['entry_price']:.4f} pnl=${position.get('pnl', 0):.4f}")
                        coin['_cached_position'] = position  # 캐시에 저장
                        self._update_roi_display(coin)
                    else:
                        # 🔥 포지션 없으면 모든 ROI 데이터 초기화 + "-" 표시!
                        coin['roi']['long_entry'] = None
                        coin['roi']['short_entry'] = None
                        coin['roi']['long_current'] = 0
                        coin['roi']['short_current'] = 0
                        coin['roi']['long_max'] = 0
                        coin['roi']['short_max'] = 0
                        coin['roi']['long_min'] = 0
                        coin['roi']['short_min'] = 0
                    
                        if 'labels' in coin and 'long_current_roi' in coin['labels']:
                            try:
                                self.root.after(0, lambda c=coin: (
                                    c['labels']['long_current_roi'].config(text="현재: -", fg='#ffffff'),
                                    c['labels']['long_max_roi'].config(text="최고: -"),
                                    c['labels']['long_min_roi'].config(text="최저: -")
                                ))
                            except: pass
                    
                        if 'labels' in coin and 'short_current_roi' in coin['labels']:
                            try:
                                self.root.after(0, lambda c=coin: (
                                    c['labels']['short_current_roi'].config(text="현재: -", fg='#ffffff'),
                                    c['labels']['short_max_roi'].config(text="최고: -"),
                                    c['labels']['short_min_roi'].config(text="최저: -")
                                ))
                            except: pass
                        print(f"  ✅ {coin['symbol']} 포지션 없음 - ROI 초기화 완료")
                except Exception as e:
                    print(f"  ⚠️ {coin['symbol']} UI 강제 업데이트 실패: {e}")
        
            # 화면 업데이트 (메인 스레드에서)
            print("=" * 60)
            print("✅ 봇 시작 완료!")
            print("=" * 60 + "\n")
        
            # 🔥 실행 상태 저장 (자동 재연결용)
            self._save_running_state()
        
            self.root.after(0, lambda: CustomMessageBox.showinfo("전체 시작", 
                f"✅ {count}개 봇을 시작했습니다!\n\n"
                f"📊 기존 포지션이 있으면 바이낸스 실시간 ROI를 표시합니다.\n"
                f"🎯 신호 발생 시 자동으로 진입합니다."))
        finally:
            self._is_starting = False
    
    def reconnect_all(self):
        """전체 재연결 - 기존 포지션 + 통계 + ROI 이어받기"""
        prog_info = f"프로그램 #{self.program_number}: " if self.program_number else ""
        if CustomMessageBox.askyesno("🔄 전체 재연결", 
            f"{prog_info}모든 코인의 포지션을 재연결하시겠습니까?\n\n"
            f"대상: {len(self.coins)}개 코인\n"
            f"총 {len(self.coins) * 2}개 봇 시작\n\n"
            f"✅ 기존 포지션 자동 이어받기\n"
            f"✅ 실시간 ROI 계산 (바이낸스 기준)\n"
            f"✅ 통계 유지 (총 수익, 수수료 등)\n\n"
            f"💡 순차적으로 시작되므로 약 {len(self.coins)}초 소요됩니다."):
            
            # 🔥 스레드에서 실행 (GUI 멈춤 방지!)
            threading.Thread(target=self._reconnect_all_thread, daemon=True).start()
    
    def _reconnect_all_thread(self):
        if self._is_starting:
            print("[⚠️] 이미 시작/재연결 중!")
            return
        self._is_starting = True
        try:
            count = 0
            reconnected_positions = []
        
            for i, coin in enumerate(self.coins, 1):
                print(f"[전체 재연결] {i}/{len(self.coins)}: {coin['symbol']}")
            
                # 🔥 바이낸스 포지션 확인
                position = None
                for retry in range(3):
                    position = self.api.get_position(coin['symbol'])
                    if position is not None:
                        break
                    if retry < 2:
                        time.sleep(1)
            
                # 🔥🔥🔥 포지션이 있으면 실시간 ROI 계산해서 이어받기
                if position:
                    entry_price = position['entry_price']
                    current_price = None
                
                    # 현재가 조회
                    df = self.api.get_klines(coin['symbol'], coin['timeframe'], limit=1)
                    if df is not None and len(df) > 0:
                        current_price = float(df['close'].iloc[-1])
                
                    if current_price:
                        # 실시간 ROI 계산
                        leverage = coin['leverage']
                        if position['side'] == 'long':
                            roi_pct = ((current_price - entry_price) / entry_price) * 100 * leverage
                        else:  # short
                            roi_pct = ((entry_price - current_price) / entry_price) * 100 * leverage
                    
                        reconnected_positions.append({
                            'symbol': coin['symbol'],
                            'side': position['side'],
                            'entry': entry_price,
                            'current': current_price,
                            'roi': roi_pct
                        })
                    
                        # 🔥 ROI 이어받기 (최고/최저는 현재 ROI 기준으로 초기화)
                        if position['side'] == 'long':
                            coin['roi']['long_entry'] = entry_price
                            coin['roi']['long_current'] = roi_pct
                            coin['roi']['long_max'] = max(0, roi_pct)  # 0 이상만
                            coin['roi']['long_min'] = min(0, roi_pct)  # 0 이하만
                        else:
                            coin['roi']['short_entry'] = entry_price
                            coin['roi']['short_current'] = roi_pct
                            coin['roi']['short_max'] = max(0, roi_pct)
                            coin['roi']['short_min'] = min(0, roi_pct)
                    
                        print(f"  🔄 포지션 이어받기: {position['side'].upper()} | 진입가: ${entry_price:.2f} | ROI: {roi_pct:+.2f}%")
                else:
                    # 포지션 없으면 ROI 초기화
                    coin['roi']['long_entry'] = None
                    coin['roi']['long_current'] = 0
                    coin['roi']['long_max'] = 0
                    coin['roi']['long_min'] = 0
                    coin['roi']['short_entry'] = None
                    coin['roi']['short_current'] = 0
                    coin['roi']['short_max'] = 0
                    coin['roi']['short_min'] = 0
            
                # 🔥 플래그 초기화 (공통)
                coin['last_close_time'] = None
                coin['is_entering_long'] = False; coin['is_entering_short'] = False
                coin['is_closing'] = False
                coin['is_closing_start'] = 0
                coin['warned_position_exists'] = False
                coin['prev_short_signal_initialized'] = False
                coin['prev_long_signal_initialized'] = False
            
                # 🔥 LONG 봇 시작
                if not coin['long_active']:
                    coin['long_active'] = True
                
                    # 신호 초기화
                    if position and position['side'] == 'long':
                        coin['prev_signals'] = {
                            'ut_long': True, 'ut_short': False,
                            'ema_long': True, 'ema_short': False,
                            'long_signal': True, 'short_signal': False
                        }
                    else:
                        coin['prev_signals'] = {
                            'ut_long': None, 'ut_short': None,
                            'ema_long': None, 'ema_short': None,
                            'long_signal': None, 'short_signal': None
                        }
                
                    if id(coin) not in self.bots:
                        self.bots[id(coin)] = {}
                
                    if 'long' in self.bots[id(coin)]:
                        self.bots[id(coin)]['long'].stop()
                        time.sleep(0.3)
                        del self.bots[id(coin)]['long']
                
                    self.bots[id(coin)]['long'] = TradingBot(
                        self.api, coin,
                        lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                        lambda c=coin: self.update_stats(c),
                        self.save_trade_to_excel,
                        bot_type='long',
                        app=self,
                        is_reconnect=True
                    )
                    self.bots[id(coin)]['long'].start()
                    count += 1
                
                    # 🔥 재연결 로그
                    self.add_log(coin, 'LONG', "=" * 40)
                    self.add_log(coin, 'LONG', "🔄 LONG 봇 재연결!")
                
                    if position and position['side'] == 'long':
                        roi = coin['roi']['long_current']
                        self.add_log(coin, 'LONG', f"📍 포지션 이어받음!")
                        self.add_log(coin, 'LONG', f"   진입가: ${position['entry_price']:.2f}")
                        self.add_log(coin, 'LONG', f"   현재 ROI: {roi:+.2f}%")
                    else:
                        self.add_log(coin, 'LONG', "📭 포지션 없음 - 신호 대기")
                
                    # 🔥 코인별 통계 표시 (이어받기)
                    s = coin['stats']
                    self.add_log(coin, 'LONG', f"📊 [통계 이어받음]")
                    self.add_log(coin, 'LONG', f"   LONG: {s['long_count']}회 (익절 {s['long_win']} / 손절 {s['long_loss']})")
                    self.add_log(coin, 'LONG', f"   LONG 수익: ${s['long_profit']:+,.2f}")
                    self.add_log(coin, 'LONG', f"   SHORT: {s['short_count']}회 (익절 {s['short_win']} / 손절 {s['short_loss']})")
                    self.add_log(coin, 'LONG', f"   SHORT 수익: ${s['short_profit']:+,.2f}")
                    self.add_log(coin, 'LONG', f"📊 코인 총계: ${s['total_pnl']:+,.2f} | 수수료: ${s['total_fee']:.2f}")
                    self.add_log(coin, 'LONG', "=" * 40)
            
                # 🔥 SHORT 봇 시작
                if not coin['short_active']:
                    coin['short_active'] = True
                
                    # 신호 초기화
                    if position and position['side'] == 'short':
                        coin['prev_signals'] = {
                            'ut_long': False, 'ut_short': True,
                            'ema_long': False, 'ema_short': True,
                            'long_signal': False, 'short_signal': True
                        }
                
                    if id(coin) not in self.bots:
                        self.bots[id(coin)] = {}
                
                    if 'short' in self.bots[id(coin)]:
                        self.bots[id(coin)]['short'].stop()
                        time.sleep(0.3)
                        del self.bots[id(coin)]['short']
                
                    self.bots[id(coin)]['short'] = TradingBot(
                        self.api, coin,
                        lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                        lambda c=coin: self.update_stats(c),
                        self.save_trade_to_excel,
                        bot_type='short',
                        app=self,
                        is_reconnect=True
                    )
                    self.bots[id(coin)]['short'].start()
                    count += 1
                
                    # 🔥 재연결 로그
                    self.add_log(coin, 'SHORT', "=" * 40)
                    self.add_log(coin, 'SHORT', "🔄 SHORT 봇 재연결!")
                
                    if position and position['side'] == 'short':
                        roi = coin['roi']['short_current']
                        self.add_log(coin, 'SHORT', f"📍 포지션 이어받음!")
                        self.add_log(coin, 'SHORT', f"   진입가: ${position['entry_price']:.2f}")
                        self.add_log(coin, 'SHORT', f"   현재 ROI: {roi:+.2f}%")
                    else:
                        self.add_log(coin, 'SHORT', "📭 포지션 없음 - 신호 대기")
                
                    # 🔥 코인별 통계 표시 (이어받기)
                    s = coin['stats']
                    self.add_log(coin, 'SHORT', f"📊 [통계 이어받음]")
                    self.add_log(coin, 'SHORT', f"   LONG: {s['long_count']}회 (익절 {s['long_win']} / 손절 {s['long_loss']})")
                    self.add_log(coin, 'SHORT', f"   LONG 수익: ${s['long_profit']:+,.2f}")
                    self.add_log(coin, 'SHORT', f"   SHORT: {s['short_count']}회 (익절 {s['short_win']} / 손절 {s['short_loss']})")
                    self.add_log(coin, 'SHORT', f"   SHORT 수익: ${s['short_profit']:+,.2f}")
                    self.add_log(coin, 'SHORT', f"📊 코인 총계: ${s['total_pnl']:+,.2f} | 수수료: ${s['total_fee']:.2f}")
                    self.add_log(coin, 'SHORT', "=" * 40)
            
                # API 부하 분산
                if i < len(self.coins):
                    time.sleep(0.3)
        
            # 결과 메시지
            msg = f"✅ {count}개 봇을 재연결했습니다!\n\n"
            if reconnected_positions:
                msg += f"📍 이어받은 포지션: {len(reconnected_positions)}개\n"
                for p in reconnected_positions[:5]:  # 최대 5개만 표시
                    msg += f"  • {p['symbol']} {p['side'].upper()}: {p['roi']:+.2f}%\n"
                if len(reconnected_positions) > 5:
                    msg += f"  ... 외 {len(reconnected_positions)-5}개\n"
            else:
                msg += "📭 이어받은 포지션 없음\n"
        
            msg += "\n💡 통계와 ROI가 유지됩니다!"
            # 🔥 실행 상태 저장 (자동 재연결용)
            self._save_running_state()
        
            self.root.after(0, lambda: CustomMessageBox.showinfo("전체 재연결", msg))
        finally:
            self._is_starting = False
    
    def stop_all(self):
        """전체 롱숏 종료 (정지 + 청산)"""
        prog_info = f"프로그램 #{self.program_number}: " if self.program_number else ""
        if CustomMessageBox.askyesno("⚠️ 전체 종료", 
            f"{prog_info}모든 코인의 롱과 숏을 종료하시겠습니까?\n\n"
            f"대상: GUI {len(self.coins)}개 코인 + 바이낸스 모든 포지션\n"
            f"동작:\n"
            f"  1. 모든 봇 정지 (진입 차단) 🛑\n"
            f"  2. 바이낸스 모든 포지션 강제 청산 🔥\n\n"
            f"※ 이 작업은 되돌릴 수 없습니다!"):
            
            # 🔥 스레드로 실행 (UI 프리즈 방지!)
            def _stop_all_thread():
                closed_count = 0
                stopped_count = 0
                
                # 🔥 1단계: 먼저 모든 봇 정지! (진입 차단)
                print("=" * 60)
                print("🛑 모든 봇 정지 중...")
                print("=" * 60)
                
                for coin in self.coins:
                    if coin['long_active']:
                        coin['long_active'] = False
                        stopped_count += 1
                    if coin['short_active']:
                        coin['short_active'] = False
                        stopped_count += 1
                    
                    # 봇 중지
                    if id(coin) in self.bots:
                        if 'long' in self.bots[id(coin)]:
                            self.bots[id(coin)]['long'].stop()
                        if 'short' in self.bots[id(coin)]:
                            self.bots[id(coin)]['short'].stop()
                    
                    # 플래그 초기화
                    coin['last_close_time'] = None
                    coin['is_entering_long'] = False; coin['is_entering_short'] = False
                    coin['is_closing'] = True  # 청산 중!
                    coin['warned_position_exists'] = False
                    coin['prev_signals'] = {
                        'ut_long': None, 'ut_short': None,
                        'ema_long': None, 'ema_short': None,
                        'long_signal': None, 'short_signal': None
                    }
                    
                    # 🔥 ROI 완전 초기화
                    coin['roi']['long_entry'] = None
                    coin['roi']['short_entry'] = None
                    coin['roi']['long_current'] = 0
                    coin['roi']['short_current'] = 0
                    coin['roi']['long_max'] = 0
                    coin['roi']['short_max'] = 0
                    coin['roi']['long_min'] = 0
                    coin['roi']['short_min'] = 0
                    coin['restart_entry_long'] = None
                    coin['restart_entry_short'] = None
                    coin['entry_tp'] = None
                    coin['tp_reached'] = False
                    coin['target_usdt'] = 0  # 목표 USDT 초기화!                    
                    # ROI UI 초기화 (메인 스레드에서!)
                    if 'labels' in coin:
                        for side in ['long', 'short']:
                            if f'{side}_current_roi' in coin['labels']:
                                try:
                                    s = side
                                    self.root.after(0, lambda c=coin, s=s: (
                                        c['labels'][f'{s}_current_roi'].config(text="현재: -", fg='#ffffff'),
                                        c['labels'][f'{s}_max_roi'].config(text="최고: -"),
                                        c['labels'][f'{s}_min_roi'].config(text="최저: -")
                                    ))
                                except:
                                    pass
                
                print(f"✅ {stopped_count}개 봇 정지 완료!")
                print("=" * 60)
                
                # 🔥 1초 대기 (봇 완전 정지)
                time.sleep(1)
                
                # 🔥 2단계: 바이낸스 모든 포지션 청산
                print("=" * 60)
                print("🔥 바이낸스 전체 포지션 청산 시작...")
                print("=" * 60)
                
                try:
                    # 바이낸스에서 모든 포지션 조회
                    all_positions = self.api._request('GET', '/fapi/v2/positionRisk', signed=True)
                    
                    if all_positions:
                        for pos in all_positions:
                            pos_amt = float(pos['positionAmt'])
                            
                            # 포지션이 있으면 청산
                            if pos_amt != 0:
                                symbol = pos['symbol']  # BTCUSDT
                                symbol_display = symbol[:-4] + '/' + symbol[-4:]  # BTC/USDT
                                side = 'LONG' if pos_amt > 0 else 'SHORT'
                                pnl = float(pos['unRealizedProfit'])
                                
                                print(f"📍 발견: {symbol_display} {side} (수익: ${pnl:.2f})")

                                # 🔥 잔여 TP 주문 먼저 취소
                                try:
                                    self.api.cancel_all_orders(symbol_display)
                                except Exception:
                                    pass

                                # 🔥 청산 재시도 (최대 3번)
                                close_success = False
                                for attempt in range(3):
                                    try:
                                        # 청산 (symbol은 BTCUSDT 형식 그대로)
                                        close_side = 'SELL' if pos_amt > 0 else 'BUY'
                                        self.api.create_order(symbol_display, close_side, abs(pos_amt))
                                        
                                        close_success = True
                                        closed_count += 1
                                        print(f"✅ {symbol_display} {side} 청산 완료! (${pnl:+.2f})")
                                        break
                                        
                                    except Exception as e:
                                        print(f"❌ {symbol_display} {side} 청산 시도 {attempt+1}/3 실패: {e}")
                                        if attempt < 2:
                                            time.sleep(1)
                                
                                if close_success:
                                    
                                    # GUI에 있는 코인이면 로그 추가
                                    for coin in self.coins:
                                        if coin['symbol'] == symbol_display:
                                            profit_sign = '+' if pnl >= 0 else ''
                                            self.add_log(coin, side, f"⏹️ 전체 종료 청산!")
                                            self.add_log(coin, side, f"   수익금: {profit_sign}${pnl:.2f}")
                                            
                                            # 통계 업데이트
                                            if side == 'LONG':
                                                coin['stats']['long_count'] += 1
                                                if pnl > 0:
                                                    coin['stats']['long_win'] += 1
                                                else:
                                                    coin['stats']['long_loss'] += 1
                                                coin['stats']['long_profit'] += pnl
                                                coin['stats']['total_pnl'] += pnl
                                            else:
                                                coin['stats']['short_count'] += 1
                                                if pnl > 0:
                                                    coin['stats']['short_win'] += 1
                                                else:
                                                    coin['stats']['short_loss'] += 1
                                                coin['stats']['short_profit'] += pnl
                                                coin['stats']['total_pnl'] += pnl
                                            
                                            self.update_stats(coin)
                                            break
                                else:
                                    print(f"❌ {symbol_display} {side} 청산 실패! (3회 시도)")
                                    
                                # 청산 간 짧은 대기
                                time.sleep(0.5)
                        
                        if closed_count == 0:
                            print("📭 청산할 포지션이 없습니다.")
                        else:
                            print(f"✅ 총 {closed_count}개 포지션 청산 완료!")
                    else:
                        print("⚠️ 포지션 조회 실패")
                        
                except Exception as e:
                    print(f"❌ 전체 포지션 조회 오류: {e}")
                
                print("=" * 60)
                
                # 🔥 3단계: 재확인 (혹시 남은 포지션 체크)
                print("=" * 60)
                print("🔍 남은 포지션 재확인 중...")
                print("=" * 60)
                
                try:
                    time.sleep(2)  # 2초 대기 후 재확인
                    remaining_positions = self.api._request('GET', '/fapi/v2/positionRisk', signed=True)
                    
                    if remaining_positions:
                        remaining_count = 0
                        for pos in remaining_positions:
                            if float(pos['positionAmt']) != 0:
                                remaining_count += 1
                                symbol = pos['symbol']
                                print(f"⚠️ 남은 포지션: {symbol} (수량: {pos['positionAmt']})")
                        
                        if remaining_count > 0:
                            print(f"⚠️ {remaining_count}개 포지션이 남아있습니다!")
                            print("웹에서 수동으로 청산해주세요.")
                        else:
                            print("✅ 모든 포지션 청산 확인!")
                    else:
                        print("✅ 모든 포지션 청산 확인!")
                        
                except Exception as e:
                    print(f"❌ 재확인 오류: {e}")
                
                print("=" * 60)
                
                # 🔥 실행 상태 삭제 (다음 시작 시 새로 시작)
                self._clear_running_state()
                
                # 🔥 완료 메시지 (UI 스레드에서 실행)
                self.root.after(0, lambda: CustomMessageBox.showinfo("전체 종료", 
                    f"✅ 전체 종료 완료!\n\n"
                    f"청산: {closed_count}개\n"
                    f"정지: {stopped_count}개 봇\n\n"
                    f"📊 통계: 메모리 초기화됨\n"
                    f"📝 거래 기록: trade_history.xlsx"))
            
            # 🔥 스레드 시작 (daemon=True로 백그라운드 실행)
            threading.Thread(target=_stop_all_thread, daemon=True).start()
    
    def init_trade_history_excel(self):
        """엑셀 거래 기록 파일 초기화 - 코인별 시트 10개"""
        if not OPENPYXL_AVAILABLE:
            print("⚠️ openpyxl 미설치 - 엑셀 저장 불가")
            return
        
        try:
            # 헤더 설정
            headers = [
                '년도', '월', '일', '시간', '분', '코인', '포지션',
                '진입금액', '레버리지', 'TP%', 'SL%', 'ROI%',
                '수익(USDT)', '손실(USDT)', '진입수수료', '청산수수료',
                '총수수료', '순수익', '거래유형'
            ]
            
            # 코인 시트 목록 (10개)
            coin_sheets = ['BTC', 'ETH', 'SOL', 'XRP', 'DOGE', 'ADA', 'BNB', 'AVAX', 'DOT', 'ETC']
            
            if os.path.exists(TRADE_HISTORY_FILE):
                # 기존 파일이 있으면 시트만 추가
                try:
                    wb = load_workbook(TRADE_HISTORY_FILE)
                    sheets_added = False
                    for sheet_name in coin_sheets:
                        if sheet_name not in wb.sheetnames:
                            ws = wb.create_sheet(title=sheet_name)
                            # 헤더 추가
                            header_font = Font(bold=True, color='FFFFFF')
                            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                            for col, header in enumerate(headers, 1):
                                cell = ws.cell(row=1, column=col, value=header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal='center')
                            sheets_added = True
                    
                    if sheets_added:
                        wb.save(TRADE_HISTORY_FILE)
                        print(f"📝 기존 파일에 시트 추가됨: {TRADE_HISTORY_FILE}")
                    else:
                        print(f"📝 기존 거래 기록 파일 사용: {TRADE_HISTORY_FILE}")
                    wb.close()
                except Exception as e:
                    print(f"⚠️ 기존 파일 열기 실패, 새로 생성: {e}")
                    # 새로 생성
                    pass
                else:
                    return
            
            # 새 워크북 생성
            wb = Workbook()
            
            # 기본 시트 제거
            default_sheet = wb.active
            
            # 코인별 시트 생성
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            for i, sheet_name in enumerate(coin_sheets):
                if i == 0:
                    ws = default_sheet
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                # 헤더 추가
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # 열 너비 조정
                ws.column_dimensions['F'].width = 12  # 코인
                ws.column_dimensions['G'].width = 10  # 포지션
                ws.column_dimensions['S'].width = 10  # 거래유형
            
            wb.save(TRADE_HISTORY_FILE)
            wb.close()
            print(f"✅ 거래 기록 파일 생성 (코인별 시트 10개): {TRADE_HISTORY_FILE}")
            
        except Exception as e:
            print(f"❌ 엑셀 파일 초기화 실패: {e}")
    
    def save_trade_to_excel(self, coin, position_type, entry_amount, leverage,
                           tp_pct, sl_pct, roi_pct, profit_usdt, loss_usdt,
                           entry_fee, close_fee, total_fee, net_profit, trade_type):
        """거래 기록을 엑셀에 저장 - 코인별 시트 + 파일 잠금 처리"""
        if not OPENPYXL_AVAILABLE:
            return
        
        # 🔥 코인 심볼에서 시트명 추출 (BTC/USDT → BTC)
        symbol = coin['symbol'].replace('/USDT', '').replace('USDT', '')
        
        # 현재 시간
        now = datetime.now()
        
        # 새 행 데이터
        row_data = [
            now.year,
            now.month,
            now.day,
            now.hour,
            now.minute,
            coin['symbol'],
            position_type,
            entry_amount,
            leverage,
            tp_pct if tp_pct else 0,
            sl_pct if sl_pct else 0,
            round(roi_pct, 2),
            round(profit_usdt, 2) if profit_usdt > 0 else 0,
            round(loss_usdt, 2) if loss_usdt > 0 else 0,
            round(entry_fee, 4),
            round(close_fee, 4),
            round(total_fee, 4),
            round(net_profit, 2),
            trade_type
        ]
        
        # 🔥 파일 잠금 대비 - 최대 3회 재시도
        for attempt in range(3):
            try:
                wb = load_workbook(TRADE_HISTORY_FILE)
                
                # 🔥 코인별 시트 찾기 또는 생성
                if symbol in wb.sheetnames:
                    ws = wb[symbol]
                else:
                    # 시트가 없으면 생성
                    ws = wb.create_sheet(title=symbol)
                    # 헤더 추가
                    headers = [
                        '년도', '월', '일', '시간', '분', '코인', '포지션',
                        '진입금액', '레버리지', 'TP%', 'SL%', 'ROI%',
                        '수익(USDT)', '손실(USDT)', '진입수수료', '청산수수료',
                        '총수수료', '순수익', '거래유형'
                    ]
                    header_font = Font(bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                
                ws.append(row_data)
                wb.save(TRADE_HISTORY_FILE)
                wb.close()
                
                print(f"📝 거래 기록 저장 [{symbol}]: {position_type} {trade_type} - 순수익 ${net_profit:.2f}")
                return  # 성공 시 종료
                
            except PermissionError:
                # 🔥 파일이 열려있는 경우
                print(f"⚠️ 엑셀 파일 잠김 (시도 {attempt+1}/3) - 2초 후 재시도...")
                time.sleep(2)
            except Exception as e:
                print(f"❌ 엑셀 저장 실패: {e}")
                return
        
        # 3회 실패 시 CSV로 백업 저장
        try:
            backup_file = f"trade_backup_{now.strftime('%Y%m%d')}.csv"
            import csv
            with open(backup_file, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(row_data)
            print(f"📝 CSV 백업 저장: {backup_file}")
        except Exception as e:
            print(f"❌ CSV 백업도 실패: {e}")
    
    def on_closing(self):
        """X 버튼 클릭 시 종료 처리"""
        print("=" * 60)
        print("🔚 프로그램 종료 중...")
        print("=" * 60)
        print("📊 통계 저장 중...")
        print("📝 거래 기록: trade_history.xlsx에 저장됨")
        print("=" * 60)
        
        # 🔥 통계 저장!
        self.save_stats()
        
        # 모든 봇 정지
        for coin in self.coins:
            if id(coin) in self.bots:
                if 'long' in self.bots[id(coin)]:
                    self.bots[id(coin)]['long'].stop()
                if 'short' in self.bots[id(coin)]:
                    self.bots[id(coin)]['short'].stop()
        
        # 프로그램 종료
        self.root.destroy()
    
    def save_stats(self):
        """🔥 통계 저장 - JSON 파일"""
        try:
            import json
            
            data = {
                'global_stats': self.global_stats,
                'coins': {}
            }
            
            # 각 코인별 통계 저장
            for coin in self.coins:
                symbol = coin['symbol']
                data['coins'][symbol] = {
                    'stats': coin.get('stats', {}),
                    'logs': {
                        'long': coin.get('long_logs', [])[-100:],  # 최근 100개만
                        'short': coin.get('short_logs', [])[-100:]
                    }
                }
            
            with open(STATS_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"✅ 통계 저장 완료: {STATS_FILE}")
        except Exception as e:
            print(f"❌ 통계 저장 실패: {e}")
    
    def load_stats(self):
        """🔥 통계 불러오기 - JSON 파일"""
        try:
            import json
            
            if not os.path.exists(STATS_FILE):
                print(f"📭 저장된 통계 없음 (새로 시작)")
                return
            
            with open(STATS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 전체 통계 불러오기
            if 'global_stats' in data:
                self.global_stats = data['global_stats']
                print(f"✅ 전체 통계 불러옴: 총 수익 ${self.global_stats.get('total_pnl', 0):+,.2f}")
            
            # 각 코인별 통계 불러오기
            if 'coins' in data:
                for coin in self.coins:
                    symbol = coin['symbol']
                    if symbol in data['coins']:
                        coin_data = data['coins'][symbol]
                        
                        # 통계 복원
                        if 'stats' in coin_data:
                            coin['stats'] = coin_data['stats']
                            print(f"  ✅ {symbol} 통계 복원: ${coin['stats'].get('total_pnl', 0):+,.2f}")
                        
                        # 로그 복원
                        if 'logs' in coin_data:
                            coin['long_logs'] = coin_data['logs'].get('long', [])
                            coin['short_logs'] = coin_data['logs'].get('short', [])
                            
                            # GUI 로그 텍스트 위젯에도 추가
                            if 'labels' in coin and 'long_log' in coin['labels'] and coin['long_logs']:
                                try:
                                    log_widget = coin['labels']['long_log']
                                    for log in coin['long_logs']:
                                        log_widget.insert('end', log + '\n')
                                    log_widget.see('end')
                                except:
                                    pass
                            
                            if 'labels' in coin and 'short_log' in coin['labels'] and coin['short_logs']:
                                try:
                                    log_widget = coin['labels']['short_log']
                                    for log in coin['short_logs']:
                                        log_widget.insert('end', log + '\n')
                                    log_widget.see('end')
                                except:
                                    pass
            
            print(f"✅ 통계 불러오기 완료!")
            
            # 🔥 UI에 통계 반영
            self.root.after(500, self._refresh_stats_ui)
            
        except Exception as e:
            print(f"❌ 통계 불러오기 실패: {e}")
    
    def _refresh_stats_ui(self):
        """🔥 저장된 통계를 UI에 반영"""
        try:
            # 전체 통계 UI 업데이트
            total_pnl = self.global_stats.get('total_pnl', 0)
            total_fee = self.global_stats.get('total_fee', 0)
            total_trades = self.global_stats.get('total_trades', 0)
            total_wins = self.global_stats.get('total_wins', 0)
            
            pnl_color = '#00ff00' if total_pnl >= 0 else '#ff0000'
            pnl_sign = '+' if total_pnl >= 0 else ''
            win_rate = (total_wins / total_trades * 100) if total_trades > 0 else 0
            
            self.global_pnl_label.config(text=f"📊 순수익: {pnl_sign}${total_pnl:.2f}", fg=pnl_color)
            self.global_fee_label.config(text=f"💸 수수료: -${total_fee:.2f}", fg='#ff6666')
            self.global_trades_label.config(text=f"거래: {total_trades}회 | 승률: {win_rate:.1f}%")
            
            # 각 코인별 통계 UI 업데이트
            for coin in self.coins:
                if 'stats_label' in coin:
                    try:
                        self.update_stats(coin)
                    except:
                        pass
            
            print("✅ UI 통계 반영 완료!")
        except Exception as e:
            print(f"❌ UI 통계 반영 실패: {e}")
    
    def auto_save_stats(self):
        """🔥 자동 저장 (5분마다)"""
        try:
            # 봇이 실행 중일 때만 저장
            any_active = any(coin.get('long_active') or coin.get('short_active') for coin in self.coins)
            if any_active:
                self.save_stats()
        except:
            pass
        
        # 5분 후 다시 실행
        self.root.after(300000, self.auto_save_stats)  # 300000ms = 5분
    
    def update_balance(self):
        try:
            balance = self.api.get_balance()
            self.balance_label.config(text=f"잔고: ${balance:,.2f} USDT")
        except:
            pass
        self.root.after(5000, self.update_balance)
    
    def update_4h_filter(self):
        """🔥 4H UT Bot 필터 업데이트 (60초마다)"""
        if not self.filter_4h_enabled:
            self.root.after(1800000, self.update_4h_filter)  # 30분마다
            return
        
        try:
            for coin in self.coins:
                symbol = coin['symbol']
                
                # 4H 캔들 조회
                df_4h = self.api.get_klines(symbol, '4h', limit=100)
                if df_4h is None or len(df_4h) < 60:
                    continue
                
                # 완성된 봉만 사용
                df_4h_closed = df_4h[:-1]
                
                # UT Bot 신호 계산 (코인 설정값 사용)
                ut_sens = coin.get('ut_sens', 10)
                ut_atr = coin.get('ut_atr', 2)
                
                signals_4h = Indicators.get_signals(
                    df_4h_closed, ut_sens, ut_atr,
                    coin.get('ema_fast', 34), coin.get('ema_slow', 55)
                )
                
                # 4H UT 포지션 상태 저장
                self.filter_4h_cache[symbol] = {
                    'ut_long': signals_4h.get('ut_position_long', False),
                    'ut_short': signals_4h.get('ut_position_short', False),
                    'updated': time.time()
                }
                
                # API 부하 분산
                time.sleep(0.1)
            
            self.filter_4h_last_update = time.time()
            
            # 콘솔에 상태 출력 (5분마다)
            if int(time.time()) % 300 < 60:
                print(f"\n[4H UT 필터] 업데이트 완료!")
                for symbol, data in self.filter_4h_cache.items():
                    status = "🟢 LONG" if data['ut_long'] else ("🔴 SHORT" if data['ut_short'] else "⚪ NEUTRAL")
                    print(f"  {symbol}: {status}")
                print()
        
        except Exception as e:
            print(f"[4H UT 필터 오류] {e}")
        
        # 60초 후 다시 실행
        self.root.after(1800000, self.update_4h_filter)  # 30분마다
    
    def get_4h_filter(self, symbol):
        """🔥 4H UT 필터 상태 조회"""
        if not self.filter_4h_enabled:
            return {'ut_long': True, 'ut_short': True}  # 필터 OFF면 모두 허용
        
        return self.filter_4h_cache.get(symbol, {'ut_long': True, 'ut_short': True})
    
    def add_coin_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("코인 추가")
        dialog.geometry("400x700")
        dialog.configure(bg='#2d2d2d')
        
        tk.Label(dialog, text="코인:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        coin_var = tk.StringVar(value='BTC/USDT')
        ttk.Combobox(dialog, textvariable=coin_var, values=['BTC/USDT', 'ETH/USDT', 'SOL/USDT'],
                    state='readonly', font=('Arial', 11)).pack(pady=5)
        
        tk.Label(dialog, text="시간봉:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        tf_var = tk.StringVar(value='15m')  # 기본 15분봉
        ttk.Combobox(dialog, textvariable=tf_var, values=['1m', '5m', '15m', '30m', '1h', '4h', '1d'],
                    state='readonly', font=('Arial', 11)).pack(pady=5)
        
        tk.Label(dialog, text="주문 금액 (USDT):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        amount_var = tk.StringVar(value='50')  # 기본 50
        amount_entry = tk.Entry(dialog, textvariable=amount_var, font=('Arial', 11))
        amount_entry.pack(pady=5)
        
        tk.Label(dialog, text="레버리지:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        lev_var = tk.IntVar(value=3)  # 🔥 기본 3x (안전!)
        lev_spin = tk.Spinbox(dialog, from_=1, to=20, textvariable=lev_var, font=('Arial', 11))
        lev_spin.pack(pady=5)
        
        # 🔥 수수료 미리 계산 표시
        fee_label = tk.Label(dialog, text="", bg='#2d2d2d', fg='#ffaa00', font=('Arial', 10))
        fee_label.pack(pady=5)
        
        def update_fee_estimate(*args):
            try:
                amount = float(amount_var.get())
                leverage = int(lev_var.get())
                position_size = amount * leverage
                entry_fee = position_size * FEE_RATE
                close_fee = position_size * FEE_RATE
                total_fee = entry_fee + close_fee
                fee_percent = (total_fee / amount) * 100
                
                fee_label.config(text=f"💸 거래당 수수료: ${total_fee:.2f} ({fee_percent:.2f}%)\n"
                                     f"   진입: ${entry_fee:.2f} + 청산: ${close_fee:.2f}")
            except:
                fee_label.config(text="")
        
        # 값 변경 시 자동 업데이트
        amount_var.trace('w', update_fee_estimate)
        lev_var.trace('w', update_fee_estimate)
        update_fee_estimate()  # 초기 표시
        
        tk.Label(dialog, text="익절 (%) - 기본:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        tp_var = tk.DoubleVar(value=0.5)  # 🔥 0.5% (손익비 개선!)
        tk.Entry(dialog, textvariable=tp_var, font=('Arial', 11)).pack(pady=5)
        
        # 🔥 동적 TP 설정
        tk.Label(dialog, text="━━━━━━━━━━━━━━━━━━━━━", bg='#2d2d2d', fg='#666666', font=('Arial', 11)).pack(pady=5)
        tk.Label(dialog, text="📊 동적 TP (ADX 기반)", bg='#2d2d2d', fg='#00ffff', font=('Arial', 11, 'bold')).pack(pady=5)
        
        tk.Label(dialog, text="추세장 TP (%) - ADX >= 21:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        tp_trend_var = tk.DoubleVar(value=1.2)  # ✅ 1.2% (추세장)
        tk.Entry(dialog, textvariable=tp_trend_var, font=('Arial', 11)).pack(pady=5)
        
        tk.Label(dialog, text="횡보장 TP (%) - ADX < 21:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        tp_sideways_var = tk.DoubleVar(value=1.0)  # ✅ 1.0% (횡보장)
        tk.Entry(dialog, textvariable=tp_sideways_var, font=('Arial', 11)).pack(pady=5)
        
        tk.Label(dialog, text="ADX 기간 (코인 최적: 10):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        adx_period_var = tk.IntVar(value=10)
        tk.Spinbox(dialog, from_=5, to=20, textvariable=adx_period_var, font=('Arial', 11)).pack(pady=5)
        
        tk.Label(dialog, text="손절: AUTO (스위칭) - 고정", bg='#2d2d2d', fg='#00ffff', font=('Arial', 11, 'bold')).pack(pady=10)
        sl_var = tk.DoubleVar(value=0)  # AUTO 고정
        
        def add():
            symbol = coin_var.get()
            timeframe = tf_var.get()
            
            # 1️⃣ 같은 코인+시간봉 중복 체크
            for existing in self.coins:
                if existing['symbol'] == symbol and existing['timeframe'] == timeframe:
                    CustomMessageBox.showwarning("중복 탭", f"⚠️ {symbol} {timeframe} 탭이 이미 존재합니다!")
                    return
            
            # 2️⃣ 같은 코인의 다른 시간봉 차단
            for existing in self.coins:
                if existing['symbol'] == symbol:
                    CustomMessageBox.showwarning(
                        "코인 중복",
                        f"⚠️ {symbol} 탭이 이미 존재합니다!\n\n"
                        f"기존 시간봉: {existing['timeframe']}\n"
                        f"추가 시도: {timeframe}\n\n"
                        f"같은 코인은 한 번에 1개 시간봉만 거래 가능합니다.\n"
                        f"기존 탭을 삭제한 후 추가해주세요."
                    )
                    return
            
            # 3️⃣ 같은 코인에 포지션이 있는지 체크
            symbol_clean = symbol.replace('/', '')
            try:
                all_positions = self.api._request('GET', '/fapi/v2/positionRisk', signed=True)
                if all_positions:
                    for pos in all_positions:
                        pos_amt = float(pos['positionAmt'])
                        if pos['symbol'] == symbol_clean and pos_amt != 0:
                            side_name = 'LONG' if pos_amt > 0 else 'SHORT'
                            CustomMessageBox.showwarning(
                                "포지션 존재",
                                f"⚠️ {symbol}에 이미 포지션이 있습니다!\n\n"
                                f"포지션: {side_name}\n\n"
                                f"같은 코인은 하나의 시간봉만 거래 가능합니다.\n"
                                f"기존 포지션을 청산한 후 추가해주세요."
                            )
                            return
            except Exception as e:
                print(f"포지션 체크 오류: {e}")
            
            coin = {'symbol': symbol, 'timeframe': timeframe, 'amount': float(amount_var.get()),
                   'leverage': lev_var.get(), 'tp': tp_var.get(), 'sl': sl_var.get(),
                   'ut_sens': 10, 'ut_atr': 2, 'ema_fast': 34, 'ema_slow': 55,
                   'long_active': False, 'short_active': False,
                   'tp_trend': tp_trend_var.get(), 'tp_sideways': tp_sideways_var.get(),
                   'adx_period': adx_period_var.get()}
            self.add_coin(coin)
            dialog.destroy()
        
        tk.Button(dialog, text="추가", command=add, bg='#00aa00', fg='#ffffff',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(pady=20)
    
    def add_coin(self, coin):
        self.coins.append(coin)
        
        # 로그 저장용 리스트 추가
        coin['long_logs'] = []
        coin['short_logs'] = []
        
        # 통계 초기화
        coin['stats'] = {
            'total_pnl': 0,
            'long_count': 0,
            'short_count': 0,
            'long_win': 0,
            'short_win': 0,
            'long_loss': 0,
            'short_loss': 0,
            'long_profit': 0,
            'short_profit': 0,
            # 🔥 수수료 추가!
            'long_fee': 0,  # 롱 수수료 합계
            'short_fee': 0,  # 숏 수수료 합계
            'total_fee': 0,  # 전체 수수료
        }
        
        # ROI 추적
        coin['roi'] = {
            'long_entry': None,
            'long_current': 0,
            'long_max': 0,
            'long_min': 0,
            'short_entry': None,
            'short_current': 0,
            'short_max': 0,
            'short_min': 0,
        }
        
        # 청산 시간 추적 (LONG/SHORT 공유)
        coin['last_close_time'] = None
        
        # 🔥 LONG/SHORT 각각 분리!
        coin['target_usdt_long'] = 0
        coin['target_usdt_short'] = 0
        coin['entry_tp_long'] = None
        coin['entry_tp_short'] = None
        coin['tp_reached_long'] = False
        coin['tp_reached_short'] = False
        
        # 진행 중 플래그 (LONG/SHORT 공유)
        coin['is_entering_long'] = False; coin['is_entering_short'] = False
        coin['is_closing'] = False
        
        # 경고 표시 플래그
        coin['warned_position_exists'] = False
        
        # 🔥 신호 추적 (None으로 초기화 - 첫 신호 감지용)
        coin['prev_signals'] = {
            'ut_long': None,
            'ut_short': None,
            'ema_long': None,
            'ema_short': None,
            'long_signal': None,  # ut_long AND ema_long
            'short_signal': None  # ut_short AND ema_short
        }
        
        # 탭 버튼
        tab_text = f"{coin['symbol'].split('/')[0]} {coin['timeframe']}"
        btn_frame = tk.Frame(self.tab_frame, bg='#2d2d2d')
        btn_frame.pack(side='left', padx=5)
        
        btn = tk.Button(btn_frame, text=tab_text, command=lambda: self.show_coin(coin),
                       bg='#3d3d3d', fg='#ffffff', font=('Arial', 11), padx=15, pady=5)
        btn.pack(side='left')
        
        close_btn = tk.Button(btn_frame, text="✖", command=lambda: self.remove_coin(coin, btn_frame),
                             bg='#aa0000', fg='#ffffff', font=('Arial', 10, 'bold'), padx=5, pady=5)
        close_btn.pack(side='left')
        
        self.tab_buttons[id(coin)] = btn_frame
        
        if len(self.coins) == 1:
            self.show_coin(coin)
    
    def remove_coin(self, coin, btn_frame):
        # 포지션 확인
        position = self.api.get_position(coin['symbol'])
        if position:
            CustomMessageBox.showwarning("삭제 불가", "⚠️ 포지션이 열려있습니다!\n\n먼저 롱과 숏을 모두 강제정지한 후\n삭제해주세요.")
            return
        
        if not coin['long_active'] and not coin['short_active']:
            if CustomMessageBox.askyesno("탭 삭제", f"{coin['symbol']} {coin['timeframe']} 탭을 삭제하시겠습니까?"):
                self.coins.remove(coin)
                btn_frame.destroy()
                if id(coin) in self.bots:
                    self.bots[id(coin)]['long'].stop()
                    self.bots[id(coin)]['short'].stop()
                    del self.bots[id(coin)]
                # 메인 프레임 초기화
                for w in self.main_frame.winfo_children():
                    w.destroy()
        else:
            CustomMessageBox.showwarning("삭제 불가", "⚠️ 봇이 실행 중입니다!\n\n먼저 롱과 숏을 모두 강제정지한 후\n삭제해주세요.")
    
    def show_coin(self, coin):
        for w in self.main_frame.winfo_children():
            w.destroy()
        
        # 잔고 + 신호 상태 (상단) - 컴팩트
        status_frame = tk.LabelFrame(self.main_frame, text=f"📊 {coin['symbol']} - {coin['timeframe']}",
                                    bg='#2d2d2d', fg='#ffffff', font=('Arial', 11, 'bold'))
        status_frame.pack(fill='x', padx=10, pady=5)
        
        # 한 줄에 표시
        status_line = tk.Frame(status_frame, bg='#2d2d2d')
        status_line.pack(pady=5)
        
        # 🔥 트레이딩뷰 차트 버튼
        symbol_clean = coin['symbol'].replace('/', '')  # BTC/USDT → BTCUSDT
        tv_url = f"https://www.tradingview.com/chart/?symbol=BINANCE:{symbol_clean}.P"
        
        tv_btn = tk.Button(status_line, text="📈 차트", 
                          command=lambda url=tv_url: webbrowser.open(url),
                          bg='#1e88e5', fg='#ffffff', font=('Arial', 9, 'bold'),
                          padx=8, pady=2)
        tv_btn.pack(side='left', padx=5)
        
        # 사용 가능한 USDT (짧게)
        coin['usdt_label'] = tk.Label(status_line, text="💰 $0.00", bg='#2d2d2d', 
                                      fg='#00ff00', font=('Arial', 10, 'bold'))
        coin['usdt_label'].pack(side='left', padx=5)
        
        # 현재가 (짧게)
        coin['price_label'] = tk.Label(status_line, text="💵 $0.00", bg='#2d2d2d',
                                       fg='#ffffff', font=('Arial', 10))
        coin['price_label'].pack(side='left', padx=5)
        
        # UT Bot 상태 (짧게)
        coin['ut_status_label'] = tk.Label(status_line, text="🤖 -", bg='#2d2d2d',
                                           fg='#aaaaaa', font=('Arial', 10))
        coin['ut_status_label'].pack(side='left', padx=5)
        
        # EMA 상태 (짧게)
        coin['ema_status_label'] = tk.Label(status_line, text="📈 -", bg='#2d2d2d', 
                                            fg='#aaaaaa', font=('Arial', 10))
        coin['ema_status_label'].pack(side='left', padx=5)
        

        # 디버깅은 제거 (너무 길어짐)
        coin['debug_label'] = None
        
        # 🔥 ADX & Volume 프레임 (반으로 나눔)
        adx_vol_frame = tk.LabelFrame(self.main_frame, text="📊 ADX(1시간봉) - 동적 TP | 📊 Volume", bg='#2d2d2d', fg='#00ffff',
                                 font=('Arial', 10, 'bold'))
        adx_vol_frame.pack(fill='x', padx=10, pady=5)
        
        # 전체 프레임을 좌우로 나눔
        main_content = tk.Frame(adx_vol_frame, bg='#2d2d2d')
        main_content.pack(fill='x', padx=10, pady=5)
        
        # === 왼쪽: ADX ===
        left_frame = tk.Frame(main_content, bg='#2d2d2d')
        left_frame.pack(side='left', fill='x', expand=True)
        
        # ADX 라벨
        coin['adx_label'] = tk.Label(left_frame, text="ADX:", bg='#2d2d2d', fg='#ffffff',
                                     font=('Arial', 9))
        coin['adx_label'].pack(side='left', padx=5)
        
        # 막대바 캔버스 (길이 증가)
        coin['adx_canvas'] = tk.Canvas(left_frame, width=200, height=20, bg='#1e1e1e', 
                                       highlightthickness=0)
        coin['adx_canvas'].pack(side='left', padx=5)
        
        # ADX 값 라벨
        coin['adx_value_label'] = tk.Label(left_frame, text="--", bg='#2d2d2d', fg='#ffffff',
                                           font=('Arial', 10, 'bold'))
        coin['adx_value_label'].pack(side='left', padx=5)
        
        # 시장 타입 라벨
        coin['market_type_label'] = tk.Label(left_frame, text="(알 수 없음)", bg='#2d2d2d', fg='#aaaaaa',
                                            font=('Arial', 9))
        coin['market_type_label'].pack(side='left', padx=5)
        
        # === 오른쪽: Volume ===
        right_frame = tk.Frame(main_content, bg='#2d2d2d')
        right_frame.pack(side='left', fill='x', expand=True, padx=(20, 0))
        
        # Volume 라벨
        coin['volume_label'] = tk.Label(right_frame, text="Vol:", bg='#2d2d2d', fg='#ffffff',
                                        font=('Arial', 9))
        coin['volume_label'].pack(side='left', padx=5)
        
        # Volume 막대바 (길이 증가)
        coin['volume_canvas'] = tk.Canvas(right_frame, width=200, height=20, bg='#1e1e1e',
                                          highlightthickness=0)
        coin['volume_canvas'].pack(side='left', padx=5)
        
        # Volume 배수 라벨
        coin['volume_ratio_label'] = tk.Label(right_frame, text="--", bg='#2d2d2d', fg='#ffffff',
                                              font=('Arial', 10, 'bold'))
        coin['volume_ratio_label'].pack(side='left', padx=5)
        
        # Volume 상태 라벨
        coin['volume_status_label'] = tk.Label(right_frame, text="(알 수 없음)", bg='#2d2d2d', fg='#aaaaaa',
                                               font=('Arial', 9))
        coin['volume_status_label'].pack(side='left', padx=5)
        
        # TP 라벨 (ADX/Volume 프레임 하단에 추가)
        tp_label_frame = tk.Frame(adx_vol_frame, bg='#2d2d2d')
        tp_label_frame.pack(fill='x', padx=10, pady=(0, 5))
        
        coin['dynamic_tp_label'] = tk.Label(tp_label_frame, text="TP: --", bg='#2d2d2d', fg='#00ff00',
                                           font=('Arial', 9, 'bold'))
        coin['dynamic_tp_label'].pack(side='left', padx=5)
        
        # 설정 - 컴팩트
        settings_frame = tk.LabelFrame(self.main_frame, text="⚙️ 설정", bg='#2d2d2d', fg='#ffffff',
                                      font=('Arial', 10, 'bold'))
        settings_frame.pack(fill='x', padx=10, pady=5)
        
        st = tk.Frame(settings_frame, bg='#2d2d2d')
        st.pack(padx=10, pady=5)
        
        # 한 줄에 모두 표시
        settings_line = tk.Frame(st, bg='#2d2d2d')
        settings_line.pack()
        
        target_roi = coin['tp'] * coin['leverage']
        tk.Label(settings_line, text=f"Lev:{coin['leverage']}x", bg='#2d2d2d', fg='#aaaaaa', 
                font=('Arial', 9)).pack(side='left', padx=3)
        tk.Label(settings_line, text=f"TP:{coin['tp']}%→ROI +{target_roi:.1f}%", bg='#2d2d2d', fg='#00ff00',
                font=('Arial', 9, 'bold')).pack(side='left', padx=3)
        
        # SL은 항상 AUTO (스위칭)
        tk.Label(settings_line, text=f"SL:AUTO", bg='#2d2d2d', fg='#00ffff',
                font=('Arial', 9, 'bold')).pack(side='left', padx=3)
        
        tk.Label(settings_line, text=f"UT:{coin.get('ut_sens', 10)},{coin.get('ut_atr', 2)}", bg='#2d2d2d', fg='#aaaaaa',
                font=('Arial', 9)).pack(side='left', padx=3)
        tk.Label(settings_line, text=f"EMA:{coin.get('ema_fast', 34)},{coin.get('ema_slow', 55)}", bg='#2d2d2d', fg='#aaaaaa',
                font=('Arial', 9)).pack(side='left', padx=3)
        
        # 볼륨 필터 상태 표시
        vol_status = "ON" if coin.get('volume_filter_enabled', True) else "OFF"
        vol_color = '#00ff00' if coin.get('volume_filter_enabled', True) else '#ff4444'
        tk.Label(settings_line, text=f"Vol:{vol_status}", bg='#2d2d2d', fg=vol_color,
                font=('Arial', 9, 'bold')).pack(side='left', padx=3)
        
        tk.Button(settings_line, text="⚙️", command=lambda: self.edit_settings(coin), bg='#3d3d3d', fg='#ffffff',
                 font=('Arial', 9), padx=5, pady=0).pack(side='left', padx=5)
        
        # 코인별 통계 - 컴팩트
        stats_frame = tk.LabelFrame(self.main_frame, text="📊 통계", bg='#2d2d2d', fg='#ffffff',
                                    font=('Arial', 10, 'bold'))
        stats_frame.pack(fill='x', padx=10, pady=5)
        
        coin['stats_label'] = tk.Label(stats_frame, 
                                       text="수익: $0.00 | L:0회 (익0/손0) $0.00 | S:0회 (익0/손0) $0.00",
                                       bg='#2d2d2d', fg='#00ff00', font=('Arial', 9))
        coin['stats_label'].pack(pady=5)
        
        # 롱/숏 패널
        panels = tk.Frame(self.main_frame, bg='#1e1e1e')
        panels.pack(fill='both', expand=True, padx=10, pady=5)
        
        self.create_panel(panels, coin, 'LONG')
        self.create_panel(panels, coin, 'SHORT')
        
        # UI 생성 완료 후 통계 업데이트
        self.update_stats(coin)
        
        # 신호 업데이트 시작
        self.update_signals(coin)
    
    def edit_settings(self, coin):
        dialog = tk.Toplevel(self.root)
        dialog.title("설정 수정")
        dialog.geometry("450x900")  # 높이 증가
        dialog.configure(bg='#2d2d2d')
        
        # 🔥 스크롤 가능한 프레임 추가
        canvas = tk.Canvas(dialog, bg='#2d2d2d', highlightthickness=0)
        scrollbar = tk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2d2d2d')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 마우스 휠 스크롤
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # 🔥 모든 위젯을 scrollable_frame에 추가
        # 진입금
        tk.Label(scrollable_frame, text="진입금 (USDT):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        amount_var = tk.DoubleVar(value=coin['amount'])
        tk.Entry(scrollable_frame, textvariable=amount_var, font=('Arial', 11)).pack(pady=5)
        
        # 레버리지
        tk.Label(scrollable_frame, text="레버리지:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        lev_var = tk.IntVar(value=coin['leverage'])
        tk.Spinbox(scrollable_frame, from_=1, to=20, textvariable=lev_var, font=('Arial', 11)).pack(pady=5)
        
        # 익절
        tk.Label(scrollable_frame, text="익절 (%) - 기본:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 11)).pack(pady=5)
        tp_var = tk.DoubleVar(value=coin['tp'])
        tk.Entry(scrollable_frame, textvariable=tp_var, font=('Arial', 11)).pack(pady=5)
        
        # 🔥 동적 TP 설정
        tk.Label(scrollable_frame, text="━━━━━━━━━━━━━━━━━━━━━", bg='#2d2d2d', fg='#666666', font=('Arial', 10)).pack(pady=5)
        tk.Label(scrollable_frame, text="📊 동적 TP (ADX 기반)", bg='#2d2d2d', fg='#00ffff', font=('Arial', 11, 'bold')).pack(pady=5)
        
        tk.Label(scrollable_frame, text="추세장 TP (%) - ADX >= 21:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        tp_trend_var = tk.DoubleVar(value=coin.get('tp_trend', 1.2))  # ✅ 기본값 1.2%
        tk.Entry(scrollable_frame, textvariable=tp_trend_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="횡보장 TP (%) - ADX < 21:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        tp_sideways_var = tk.DoubleVar(value=coin.get('tp_sideways', 1.0))  # ✅ 기본값 1.0%
        tk.Entry(scrollable_frame, textvariable=tp_sideways_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="ADX 기간 (코인 최적: 10):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        adx_period_var = tk.IntVar(value=coin.get('adx_period', 10))
        tk.Spinbox(scrollable_frame, from_=5, to=20, textvariable=adx_period_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="💡 팁: 동적 TP 사용 안 하려면 추세장/횡보장 같은 값으로 설정", bg='#2d2d2d', fg='#ffff00', font=('Arial', 9, 'italic')).pack(pady=3)
        
        # SL은 항상 AUTO (고정)
        tk.Label(scrollable_frame, text="손절: AUTO (스위칭) - 고정", bg='#2d2d2d', fg='#00ffff', 
                font=('Arial', 11, 'bold')).pack(pady=10)
        
        # --- UT Bot 설정 ---
        tk.Label(scrollable_frame, text="━━━━━━━━━━━━━━━━━━━━━", bg='#2d2d2d', fg='#666666', font=('Arial', 10)).pack(pady=10)
        tk.Label(scrollable_frame, text="📊 UT Bot 설정", bg='#2d2d2d', fg='#00ffff', font=('Arial', 11, 'bold')).pack(pady=5)
        
        tk.Label(scrollable_frame, text="Key Value (Sensitivity):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        ut_sens_var = tk.DoubleVar(value=coin.get('ut_sens', 10))
        tk.Entry(scrollable_frame, textvariable=ut_sens_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="ATR Period:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        ut_atr_var = tk.IntVar(value=coin.get('ut_atr', 2))
        tk.Entry(scrollable_frame, textvariable=ut_atr_var, font=('Arial', 11)).pack(pady=3)
        
        # --- EMA 설정 ---
        tk.Label(scrollable_frame, text="━━━━━━━━━━━━━━━━━━━━━", bg='#2d2d2d', fg='#666666', font=('Arial', 10)).pack(pady=10)
        tk.Label(scrollable_frame, text="📈 EMA 설정", bg='#2d2d2d', fg='#00ff00', font=('Arial', 11, 'bold')).pack(pady=5)
        
        tk.Label(scrollable_frame, text="Fast EMA:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        ema_fast_var = tk.IntVar(value=coin.get('ema_fast', 34))
        tk.Entry(scrollable_frame, textvariable=ema_fast_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="Slow EMA:", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        ema_slow_var = tk.IntVar(value=coin.get('ema_slow', 55))
        tk.Entry(scrollable_frame, textvariable=ema_slow_var, font=('Arial', 11)).pack(pady=3)
        
        # --- 거래량 필터 설정 ---
        tk.Label(scrollable_frame, text="━━━━━━━━━━━━━━━━━━━━━", bg='#2d2d2d', fg='#666666', font=('Arial', 10)).pack(pady=10)
        tk.Label(scrollable_frame, text="📊 거래량 필터", bg='#2d2d2d', fg='#ffaa00', font=('Arial', 11, 'bold')).pack(pady=5)
        
        # 거래량 필터 활성화
        volume_filter_var = tk.BooleanVar(value=coin.get('volume_filter_enabled', True))
        volume_check = tk.Checkbutton(scrollable_frame, text="거래량 필터 활성화", variable=volume_filter_var,
                                      bg='#2d2d2d', fg='#ffffff', selectcolor='#1e1e1e',
                                      font=('Arial', 10), activebackground='#2d2d2d')
        volume_check.pack(pady=5)
        
        tk.Label(scrollable_frame, text="거래량 배수 (평균 × ?):", bg='#2d2d2d', fg='#ffffff', font=('Arial', 10)).pack(pady=3)
        volume_mult_var = tk.DoubleVar(value=coin.get('volume_multiplier', 1.0))
        tk.Entry(scrollable_frame, textvariable=volume_mult_var, font=('Arial', 11)).pack(pady=3)
        
        tk.Label(scrollable_frame, text="💡 거래량이 평균의 N배 이상일 때만 진입", bg='#2d2d2d', fg='#ffff00', font=('Arial', 9, 'italic')).pack(pady=3)
        tk.Label(scrollable_frame, text="(권장: 1.5 | 보수적: 2.0 | 공격적: 1.2)", bg='#2d2d2d', fg='#aaaaaa', font=('Arial', 8, 'italic')).pack(pady=1)
        
        def save():
            coin['amount'] = amount_var.get()  # 진입금 저장!
            coin['leverage'] = lev_var.get()
            coin['tp'] = tp_var.get()
            coin['sl'] = 0  # AUTO 고정
            
            # 🔥 동적 TP 저장
            coin['tp_trend'] = tp_trend_var.get()
            coin['tp_sideways'] = tp_sideways_var.get()
            coin['adx_period'] = adx_period_var.get()
            
            # UT Bot & EMA 설정 저장
            coin['ut_sens'] = ut_sens_var.get()
            coin['ut_atr'] = ut_atr_var.get()
            coin['ema_fast'] = ema_fast_var.get()
            coin['ema_slow'] = ema_slow_var.get()
            
            # 🔥 거래량 필터 저장
            coin['volume_filter_enabled'] = volume_filter_var.get()
            coin['volume_multiplier'] = volume_mult_var.get()
            
            # 마우스 휠 이벤트 해제
            canvas.unbind_all("<MouseWheel>")
            
            CustomMessageBox.showinfo("저장 완료", "설정이 저장되었습니다!\n\nSL은 AUTO (스위칭)로 고정됩니다.")
            dialog.destroy()
            self.show_coin(coin)
        
        tk.Button(scrollable_frame, text="💾 저장", command=save, bg='#00aa00', fg='#ffffff',
                 font=('Arial', 12, 'bold'), padx=30, pady=10).pack(pady=20)

    
    def create_panel(self, parent, coin, side):
        color = '#1a3d1a' if side == 'LONG' else '#3d1a1a'
        fg_color = '#00ff00' if side == 'LONG' else '#ff0000'
        
        panel = tk.LabelFrame(parent, text=f"{'🟢' if side == 'LONG' else '🔴'} {side}",
                             bg=color, fg=fg_color, font=('Arial', 12, 'bold'))
        panel.pack(side='left', fill='both', expand=True, padx=5)
        
        content = tk.Frame(panel, bg=color)
        content.pack(padx=10, pady=10, fill='both', expand=True)
        
        # 신호 상태
        sig_frame = tk.LabelFrame(content, text="진입 신호", bg=color, fg='#ffffff', font=('Arial', 10))
        sig_frame.pack(fill='x', pady=5)
        
        ut_label = tk.Label(sig_frame, text="UT: ⚪ OFF", bg=color, fg='#888888', font=('Arial', 10, 'bold'))
        ut_label.pack(pady=2)
        
        ema_label = tk.Label(sig_frame, text="EMA: ⚪ OFF", bg=color, fg='#888888', font=('Arial', 10, 'bold'))
        ema_label.pack(pady=2)
        
        status_label = tk.Label(sig_frame, text="⏸️ 대기중", bg=color, fg='#ffaa00', font=('Arial', 11, 'bold'))
        status_label.pack(pady=5)
        
        key = side.lower()
        if 'labels' not in coin:
            coin['labels'] = {}
        coin['labels'][f'{key}_ut'] = ut_label
        coin['labels'][f'{key}_ema'] = ema_label
        coin['labels'][f'{key}_status'] = status_label
        
        # 📊 실시간 ROI 표시
        roi_frame = tk.LabelFrame(content, text="📊 실시간 ROI", bg=color, fg='#ffffff', font=('Arial', 10))
        roi_frame.pack(fill='x', pady=5)
        
        current_roi_label = tk.Label(roi_frame, text="현재: -", bg=color, fg='#ffffff', font=('Arial', 11, 'bold'))
        current_roi_label.pack(pady=2)
        
        max_roi_label = tk.Label(roi_frame, text="최고: -", bg=color, fg='#00ff00', font=('Arial', 9))
        max_roi_label.pack(pady=1)
        
        min_roi_label = tk.Label(roi_frame, text="최저: -", bg=color, fg='#ff0000', font=('Arial', 9))
        min_roi_label.pack(pady=1)
        
        coin['labels'][f'{key}_current_roi'] = current_roi_label
        coin['labels'][f'{key}_max_roi'] = max_roi_label
        coin['labels'][f'{key}_min_roi'] = min_roi_label
        
        # 버튼
        btn_frame = tk.Frame(content, bg=color)
        btn_frame.pack(fill='x', pady=10)
        
        def start():
            def _start_thread():
                
                coin[f'{key}_active'] = True
                
                # 🔄 완전 초기화 (이전 데이터 제거!)
                coin['last_close_time'] = None
                coin['is_entering_long'] = False; coin['is_entering_short'] = False
                coin['is_closing'] = False
                coin['warned_position_exists'] = False
                coin['prev_signals'] = {
                    'ut_long': None,
                    'ut_short': None,
                    'ema_long': None,
                    'ema_short': None,
                    'long_signal': None,
                    'short_signal': None
                }
                
                # 🔥 완전히 새로 시작! (이전 포지션 정보 무시!)
                position = self.api.get_position(coin['symbol'])
                
                if key == 'long':
                    coin['roi']['long_entry'] = None
                    coin['roi']['long_current'] = 0
                    coin['roi']['long_max'] = 0
                    coin['roi']['long_min'] = 0
                    coin['restart_entry_long'] = None  # 무조건 None!
                    coin['chart_entry_long'] = None  # 🔥🔥🔥 차트 진입가도 초기화!
                    coin['target_usdt_long'] = 0
                    coin['entry_tp_long'] = None
                    coin['tp_reached_long'] = False
                    coin['entry_fee_long'] = 0  # 🔥 수수료도 초기화!
                    
                    # 🔥 UI 즉시 업데이트 (화면에 - 표시!)
                    try:
                        if 'labels' in coin and 'long_current_roi' in coin['labels']:
                            coin['labels']['long_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels']['long_max_roi'].config(text="최고: -")
                            coin['labels']['long_min_roi'].config(text="최저: -")
                    except Exception as e:
                        print(f"⚠️ UI 업데이트 실패: {e}")
                    
                    if position and position['side'] == 'long':
                        print(f"[{coin['symbol']}] LONG 시작 - 포지션 있음 (진입가 무시, ROI 0%부터)")
                    else:
                        print(f"[{coin['symbol']}] LONG 시작 - ROI 0%부터")
                        
                else:  # short
                    coin['roi']['short_entry'] = None
                    coin['roi']['short_current'] = 0
                    coin['roi']['short_max'] = 0
                    coin['roi']['short_min'] = 0
                    coin['restart_entry_short'] = None  # 무조건 None!
                    coin['chart_entry_short'] = None  # 🔥🔥🔥 차트 진입가도 초기화!
                    coin['target_usdt_short'] = 0
                    coin['entry_tp_short'] = None
                    coin['tp_reached_short'] = False
                    coin['entry_fee_short'] = 0  # 🔥 수수료도 초기화!
                    
                    # 🔥 UI 즉시 업데이트 (화면에 - 표시!)
                    try:
                        if 'labels' in coin and 'short_current_roi' in coin['labels']:
                            coin['labels']['short_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels']['short_max_roi'].config(text="최고: -")
                            coin['labels']['short_min_roi'].config(text="최저: -")
                    except Exception as e:
                        print(f"⚠️ UI 업데이트 실패: {e}")
                    
                    if position and position['side'] == 'short':
                        print(f"[{coin['symbol']}] SHORT 시작 - 포지션 있음 (진입가 무시, ROI 0%부터)")
                    else:
                        print(f"[{coin['symbol']}] SHORT 시작 - ROI 0%부터")
                
                if id(coin) not in self.bots:
                    self.bots[id(coin)] = {}
                
                # 🔥🔥🔥 기존 봇 완전 정지하고 삭제! (중요!)
                if key in self.bots[id(coin)]:
                    print(f"[{coin['symbol']}] {key.upper()} 기존 봇 정지 중...")
                    self.bots[id(coin)][key].stop()
                    time.sleep(0.5)  # 완전히 정지될 때까지 대기
                    del self.bots[id(coin)][key]
                    print(f"[{coin['symbol']}] {key.upper()} 기존 봇 삭제 완료")
                
                # 🔥 항상 완전히 새 봇 생성! (excel_callback 추가)
                print(f"[{coin['symbol']}] {key.upper()} 새 봇 생성 중...")
                self.bots[id(coin)][key] = TradingBot(
                    self.api, coin,
                    lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                    lambda: self.update_stats(coin),
                    self.save_trade_to_excel,  # 🔥 엑셀 콜백!
                    bot_type=key,
                    app=self  # 🔥 App 참조 전달
                )
                self.bots[id(coin)][key].start()
                print(f"[{coin['symbol']}] {key.upper()} 새 봇 시작 완료!")
                
                # 즉시 신호 확인 알림
                df = self.api.get_klines(coin['symbol'], coin['timeframe'])
                if df is not None and len(df) >= 60:
                    signals = Indicators.get_signals(df, coin['ut_sens'], coin['ut_atr'],
                                                    coin['ema_fast'], coin['ema_slow'])
                    
                    if side == 'LONG':
                        ut_on = signals.get('ut_position_long', False)
                        if ut_on and signals['ema_long']:
                            msg = f"✅ {side} 봇 시작!\n\n현재 진입 조건 충족 상태입니다.\n곧 자동으로 진입됩니다."
                        else:
                            ut_status = "✅" if ut_on else "❌"
                            ema_status = "✅" if signals['ema_long'] else "❌"
                            msg = f"✅ {side} 봇 시작!\n\nUT Bot: {ut_status}\nEMA 골든크로스: {ema_status}\n\n조건 충족 시 자동 진입됩니다."
                    else:  # SHORT
                        ut_on = signals.get('ut_position_short', False)
                        if ut_on and signals['ema_short']:
                            msg = f"✅ {side} 봇 시작!\n\n현재 진입 조건 충족 상태입니다.\n곧 자동으로 진입됩니다."
                        else:
                            ut_status = "✅" if ut_on else "❌"
                            ema_status = "✅" if signals['ema_short'] else "❌"
                            msg = f"✅ {side} 봇 시작!\n\nUT Bot: {ut_status}\nEMA 데드크로스: {ema_status}\n\n조건 충족 시 자동 진입됩니다."
                    
                    self.root.after(0, lambda: CustomMessageBox.showinfo("봇 시작", msg))
                else:
                    self.root.after(0, lambda: CustomMessageBox.showinfo("시작", f"{side} 봇을 시작합니다!"))
            
            # 스레드로 실행 (UI 프리즈 방지)
            threading.Thread(target=_start_thread, daemon=True).start()
        
        def pause():
            coin[f'{key}_active'] = False
            if id(coin) in self.bots and key in self.bots[id(coin)]:
                self.bots[id(coin)][key].stop()
            
            # 🔄 플래그 초기화 (재시작 가능하게!)
            coin['last_close_time'] = None
            coin['is_entering_long'] = False; coin['is_entering_short'] = False
            coin['is_closing'] = False
            coin['warned_position_exists'] = False
            coin['prev_signals'] = {
                'ut_long': None,
                'ut_short': None,
                'ema_long': None,
                'ema_short': None,
                'long_signal': None,
                'short_signal': None
            }
            
            CustomMessageBox.showinfo("정지", f"{side} 봇을 정지합니다!\n\n📊 통계는 유지됨\n💡 재시작하면 0%부터 시작!")
        
        def force_stop():
            if CustomMessageBox.askyesno("강제 청산", f"{side} 포지션을 청산하고 진입을 금지하시겠습니까?\n\n※ 봇 정지 + 포지션 청산\n※ 재시작하려면 '▶️ 시작' 버튼을 눌러주세요."):
                # 수동 청산 전 포지션 정보 가져오기
                position = self.api.get_position(coin['symbol'])
                if position:
                    pnl_usd = position['pnl']
                    entry_price = position['entry_price']
                    
                    # 🔥 현재가 조회
                    df = self.api.get_klines(coin['symbol'], coin['timeframe'], limit=1)
                    current_price = float(df['close'].iloc[-1]) if df is not None and len(df) > 0 else entry_price
                    
                    # 🔥 ROI 계산
                    if position['side'] == 'long':
                        roi_pct = ((current_price - entry_price) / entry_price) * 100 * coin['leverage']
                    else:
                        roi_pct = ((entry_price - current_price) / entry_price) * 100 * coin['leverage']
                    
                    # 🔥 수수료 계산
                    position_size = coin['amount'] * coin['leverage']
                    entry_fee = coin.get(f'entry_fee_{key}', position_size * FEE_RATE)
                    close_fee = position_size * FEE_RATE
                    total_fee = entry_fee + close_fee
                    net_profit = pnl_usd - total_fee
                    
                    # 통계 업데이트 (순수익 기준!)
                    if side == 'LONG':
                        coin['stats']['long_count'] += 1
                        if net_profit > 0:
                            coin['stats']['long_win'] += 1
                        else:
                            coin['stats']['long_loss'] += 1
                        coin['stats']['long_profit'] += net_profit
                        coin['stats']['total_pnl'] += net_profit
                        coin['stats']['long_fee'] = coin['stats'].get('long_fee', 0) + close_fee
                        coin['stats']['total_fee'] = coin['stats'].get('total_fee', 0) + close_fee
                    else:  # SHORT
                        coin['stats']['short_count'] += 1
                        if net_profit > 0:
                            coin['stats']['short_win'] += 1
                        else:
                            coin['stats']['short_loss'] += 1
                        coin['stats']['short_profit'] += net_profit
                        coin['stats']['total_pnl'] += net_profit
                        coin['stats']['short_fee'] = coin['stats'].get('short_fee', 0) + close_fee
                        coin['stats']['total_fee'] = coin['stats'].get('total_fee', 0) + close_fee
                    
                    # 청산
                    self.api.close_position(coin['symbol'])
                    
                    # 🔥 엑셀 저장
                    self.save_trade_to_excel(
                        coin=coin,
                        position_type=side,
                        entry_amount=coin['amount'],
                        leverage=coin['leverage'],
                        tp_pct=coin.get(f'entry_tp_{key}', 0),
                        sl_pct=0,
                        roi_pct=roi_pct,
                        profit_usdt=pnl_usd if pnl_usd > 0 else 0,
                        loss_usdt=abs(pnl_usd) if pnl_usd < 0 else 0,
                        entry_fee=entry_fee,
                        close_fee=close_fee,
                        total_fee=total_fee,
                        net_profit=net_profit,
                        trade_type='강제청산'
                    )
                    
                    # 로그
                    profit_sign = '+' if net_profit >= 0 else ''
                    self.add_log(coin, side, f"🛑 강제 청산!")
                    self.add_log(coin, side, f"   수익: ${pnl_usd:.2f} | 청산수수료: -${close_fee:.2f}")
                    self.add_log(coin, side, f"   순수익: {profit_sign}${net_profit:.2f}")
                    
                    # 통계 업데이트
                    self.update_stats(coin)
                else:
                    # 포지션 없으면 그냥 청산 시도
                    self.api.close_position(coin['symbol'])
                    self.add_log(coin, side, f"🛑 강제 청산 (포지션 없음)")
                
                # 🔥 진입 금지 (active = False)
                coin[f'{key}_active'] = False
                
                # 🔥 봇 정지
                if id(coin) in self.bots and key in self.bots[id(coin)]:
                    self.bots[id(coin)][key].stop()
                
                # 🔥 청산 플래그 설정 (포지션이 사라질 때까지)
                coin['is_closing'] = True
                
                # 🔥 ROI 초기화 ("-"로 표시)
                if 'labels' in coin and f'{key}_current_roi' in coin['labels']:
                    coin['labels'][f'{key}_current_roi'].config(text="현재: -", fg='#ffffff')
                    coin['labels'][f'{key}_max_roi'].config(text="최고: -")
                    coin['labels'][f'{key}_min_roi'].config(text="최저: -")
                
                # 🔥 ROI 데이터 초기화 (다음 진입을 위해)
                if key == 'long':
                    coin['roi']['long_entry'] = None
                    coin['roi']['long_current'] = 0
                    coin['roi']['long_max'] = 0
                    coin['roi']['long_min'] = 0
                    coin['restart_entry_long'] = None
                    coin['target_usdt_long'] = 0
                    coin['entry_tp_long'] = None
                    coin['tp_reached_long'] = False
                    coin['entry_fee_long'] = 0
                else:  # short
                    coin['roi']['short_entry'] = None
                    coin['roi']['short_current'] = 0
                    coin['roi']['short_max'] = 0
                    coin['roi']['short_min'] = 0
                    coin['restart_entry_short'] = None
                    coin['target_usdt_short'] = 0
                    coin['entry_tp_short'] = None
                    coin['tp_reached_short'] = False
                    coin['entry_fee_short'] = 0
                
                # 플래그 초기화
                coin['last_close_time'] = None
                coin['is_entering_long'] = False; coin['is_entering_short'] = False
                coin['warned_position_exists'] = False
                coin['prev_signals'] = {
                    'ut_long': None,
                    'ut_short': None,
                    'ema_long': None,
                    'ema_short': None,
                    'long_signal': None,
                    'short_signal': None
                }
                
                CustomMessageBox.showinfo("강제 청산", f"{side} 포지션 청산 완료!\n\n진입이 금지되었습니다.\n📝 거래 기록 저장됨\n\n재시작하려면 '▶️ 시작' 버튼을 눌러주세요.")
        
        def force_entry():
            if CustomMessageBox.askyesno("강제 진입", 
                f"{side} 포지션을 강제로 진입하시겠습니까?\n\n※ 기존 포지션이 있으면 청산 후 진입합니다.\n※ 신호 무시하고 즉시 진입합니다.\n※ 통계에 포함됩니다."):
                
                # 🔥 0단계: 잔여 TP 주문 정리
                try:
                    self.api.cancel_all_orders(coin['symbol'])
                except Exception:
                    pass

                # 🔥 1단계: 기존 포지션 청산!
                position = self.api.get_position(coin['symbol'])
                if position:
                    print(f"[{coin['symbol']}] 기존 포지션 발견 → 청산 후 진입")

                    try:
                        # 청산 주문
                        close_side = 'SELL' if position['side'] == 'long' else 'BUY'
                        close_qty = abs(position['amount'])
                        
                        close_order = self.api.create_order(
                            coin['symbol'],
                            close_side,
                            close_qty
                        )
                        
                        if close_order:
                            print(f"[{coin['symbol']}] 기존 포지션 청산 완료")
                            time.sleep(2)  # 청산 완료 대기
                        else:
                            print(f"[{coin['symbol']}] 청산 실패 - 강제 진입 취소")
                            CustomMessageBox.showwarning("강제 진입 실패", "기존 포지션 청산 실패!\n\n다시 시도하세요.")
                            return
                    except Exception as e:
                        print(f"[{coin['symbol']}] 청산 오류: {e}")
                        CustomMessageBox.showwarning("강제 진입 실패", f"기존 포지션 청산 오류!\n\n{str(e)}")
                        return
                
                # 🔥 2단계: ROI 완전 초기화 (청산 여부 무관)
                if side == 'LONG':
                    coin['roi']['long_entry'] = None
                    coin['roi']['long_current'] = 0
                    coin['roi']['long_max'] = 0
                    coin['roi']['long_min'] = 0
                    coin['restart_entry_long'] = None
                    coin['target_usdt_long'] = 0
                    coin['entry_tp_long'] = None
                    coin['tp_reached_long'] = False
                    print(f"[{coin['symbol']}] LONG ROI 초기화 완료 (0%부터)")
                else:  # SHORT
                    coin['roi']['short_entry'] = None
                    coin['roi']['short_current'] = 0
                    coin['roi']['short_max'] = 0
                    coin['roi']['short_min'] = 0
                    coin['restart_entry_short'] = None
                    coin['target_usdt_short'] = 0
                    coin['entry_tp_short'] = None
                    coin['tp_reached_short'] = False
                    print(f"[{coin['symbol']}] SHORT ROI 초기화 완료 (0%부터)")
                
                # 🔥 3단계: 격리 모드 설정 (고정)
                self.api.set_margin_type(coin['symbol'], 'ISOLATED')
                
                # 레버리지 설정
                self.api.set_leverage(coin['symbol'], coin['leverage'])
                
                # 현재가 가져오기
                df = self.api.get_klines(coin['symbol'], coin['timeframe'], limit=1)
                if df is not None and len(df) > 0:
                    current_price = float(df['close'].iloc[-1])
                    
                    # 수량 계산 (정확하게!)
                    qty = (coin['amount'] * coin['leverage']) / current_price
                    
                    # 🔥 코인별 정밀도 정확하게 조정!
                    symbol_clean = coin['symbol'].replace('/', '')
                    
                    if symbol_clean in ['BTCUSDT', 'ETHUSDT']:
                        qty = round(qty, 3)
                        min_qty = 0.001
                    elif symbol_clean in ['SOLUSDT', 'AVAXUSDT', 'DOTUSDT', 'XRPUSDT', 'ADAUSDT', 'DOGEUSDT']:
                        qty = int(qty)
                        if qty < 1:
                            qty = 1
                        min_qty = 1
                    elif symbol_clean in ['BNBUSDT', 'LTCUSDT']:
                        qty = round(qty, 2)
                        min_qty = 0.01
                    else:
                        qty = round(qty, 2)
                        min_qty = 0.01
                    
                    # 최소 주문량 체크
                    if qty < min_qty:
                        CustomMessageBox.showwarning("주문 불가", 
                            f"주문 금액이 너무 작습니다!\n\n"
                            f"현재 수량: {qty}\n"
                            f"최소 수량: {min_qty}\n\n"
                            f"진입금을 늘리거나 레버리지를 높이세요.")
                        return
                    
                    # 🔥 4단계: 주문 (강제 진입!)
                    order_side = 'BUY' if side == 'LONG' else 'SELL'
                    order = self.api.create_order(coin['symbol'], order_side, qty)
                    
                    if order:
                        # 진입 시간 기록
                        coin['entry_time'] = time.time()
                        
                        # 🔥 동적 TP 계산 (강제 진입도 동적 TP 사용, ADX는 1시간봉!)
                        try:
                            adx_interval = coin.get('adx_interval', '1h')
                            df_full = self.api.get_klines(coin['symbol'], adx_interval, limit=100)
                            if df_full is None or len(df_full) < 50:
                                df_full = self.api.get_klines(coin['symbol'], coin['timeframe'], limit=100)
                            if df_full is not None and len(df_full) >= 50:
                                adx_period = coin.get('adx_period', 10)
                                adx_value = Indicators.calculate_adx(df_full[:-1], period=adx_period)
                                
                                if adx_value >= 21:
                                    dynamic_tp = coin.get('tp_trend', 1.2)  # ✅ 1.2%
                                    market_type = '추세장'
                                else:
                                    dynamic_tp = coin.get('tp_sideways', 1.0)  # ✅ 1.0%
                                    market_type = '횡보장'
                                
                                # 진입 시점 TP 저장 (LONG/SHORT 분리!)
                                if side == 'LONG':
                                    coin['entry_tp_long'] = dynamic_tp
                                else:
                                    coin['entry_tp_short'] = dynamic_tp
                            else:
                                dynamic_tp = coin.get('tp', 0.3)
                                market_type = '기본'
                                adx_value = 0
                                if side == 'LONG':
                                    coin['entry_tp_long'] = dynamic_tp
                                else:
                                    coin['entry_tp_short'] = dynamic_tp
                        except:
                            dynamic_tp = coin.get('tp', 0.3)
                            market_type = '기본'
                            adx_value = 0
                            if side == 'LONG':
                                coin['entry_tp_long'] = dynamic_tp
                            else:
                                coin['entry_tp_short'] = dynamic_tp
                        
                        # 로그 (동적 TP 사용)
                        target_roi = dynamic_tp * coin['leverage']
                        
                        # 🔥 목표 USDT 계산 (LONG/SHORT 분리!)
                        target_usdt = coin['amount'] * (target_roi / 100)
                        if side == 'LONG':
                            coin['target_usdt_long'] = target_usdt  # 저장!
                        else:
                            coin['target_usdt_short'] = target_usdt  # 저장!
                        
                        # 🔥 order response에서 평균 체결가 가져오기!
                        filled_price = None
                        
                        # 1차 시도: order response에서 직접 가져오기
                        if order and 'avgPrice' in order and order['avgPrice']:
                            filled_price = float(order['avgPrice'])
                            print(f"[{coin['symbol']}] 강제 진입 체결가: ${filled_price:.2f}")
                        
                        # 2차 시도: order 정보 재조회
                        if not filled_price and order and 'orderId' in order:
                            try:
                                time.sleep(0.5)  # 0.5초 대기
                                order_info = self.api.get_order(coin['symbol'], order['orderId'])
                                if order_info and 'avgPrice' in order_info and order_info['avgPrice']:
                                    filled_price = float(order_info['avgPrice'])
                                    print(f"[{coin['symbol']}] 강제 진입 체결가: ${filled_price:.2f} (재조회)")
                            except Exception as e:
                                print(f"[{coin['symbol']}] order 재조회 실패: {e}")
                        
                        # 3차 시도: 체결 내역(fills)에서 계산
                        if not filled_price and order and 'fills' in order and order['fills']:
                            try:
                                total_qty = sum(float(fill['qty']) for fill in order['fills'])
                                total_cost = sum(float(fill['price']) * float(fill['qty']) for fill in order['fills'])
                                if total_qty > 0:
                                    filled_price = total_cost / total_qty
                                    print(f"[{coin['symbol']}] 강제 진입 체결가: ${filled_price:.2f} (fills 계산)")
                            except Exception as e:
                                print(f"[{coin['symbol']}] fills 계산 실패: {e}")
                        
                        # 출력 (체결가를 못 가져왔을 경우)
                        if not filled_price:
                            print(f"[{coin['symbol']}] 강제 진입 완료 (체결가 조회 실패 - position에서 확인 예정)")
                        
                        # 🔥 수수료 계산 (진입만)
                        position_size = coin['amount'] * coin['leverage']
                        entry_fee = position_size * FEE_RATE  # 진입 수수료
                        
                        # 🔥 has_position + 캐시 설정 (봇 루프 연동!)
                        coin['has_position'] = True
                        

                        # 🔥 수수료를 통계에 반영!
                        if side == 'LONG':
                            coin['entry_fee_long'] = entry_fee
                            coin['stats']['long_fee'] = coin['stats'].get('long_fee', 0) + entry_fee
                        else:
                            coin['entry_fee_short'] = entry_fee
                            coin['stats']['short_fee'] = coin['stats'].get('short_fee', 0) + entry_fee
                        coin['stats']['total_fee'] = coin['stats'].get('total_fee', 0) + entry_fee
                        coin['stats']['total_pnl'] = coin['stats'].get('total_pnl', 0) - entry_fee
                        
                        # 🔥🔥🔥 거래소 TP 주문 등록 (설정가 도달 시 거래소가 즉시 청산!)
                        try:
                            tp_base = filled_price or current_price
                            if side == 'LONG':
                                tp_price = tp_base * (1 + dynamic_tp / 100)
                            else:
                                tp_price = tp_base * (1 - dynamic_tp / 100)
                            tp_result = self.api.create_tp_order(coin['symbol'], side.lower(), tp_price)
                            if tp_result:
                                self.add_log(coin, side, f"   📌 거래소 TP 주문 등록: ${self.api.round_price(coin['symbol'], tp_price)}")
                            else:
                                self.add_log(coin, side, f"   ⚠️ TP 주문 등록 실패 → 봇 폴링으로 익절 처리")
                        except Exception as e:
                            print(f"[{coin['symbol']}] 강제 진입 TP 주문 오류: {e}")

                        self.add_log(coin, side, f"🔥 강제 진입! ${current_price:.2f} | {qty}개 | {coin['timeframe']} | {coin['leverage']}x")
                        if market_type != '기본':
                            self.add_log(coin, side, f"   📊 ADX: {adx_value:.1f} ({market_type}) → TP {dynamic_tp}%")
                        self.add_log(coin, side, f"   🎯 TP: 가격+{dynamic_tp:.2f}% → ROI +{target_roi:.2f}% → 💰 ${target_usdt:.2f}")
                        self.add_log(coin, side, f"   💸 진입 수수료: ${entry_fee:.2f} ({FEE_RATE*100:.2f}%) (청산 시 합산)")
                        self.add_log(coin, side, f"   🛡️ SL: AUTO (스위칭)")
                        
                        # 신호 초기화 (자동 청산 가능하게)
                        coin['prev_signals'] = {
                            'ut_long': False,
                            'ut_short': False,
                            'ema_long': False,
                            'ema_short': False,
                            'long_signal': False,
                            'short_signal': False
                        }
                        
                        # 🔥 청산 관련 플래그 초기화
                        coin['is_closing'] = False
                        coin['is_entering_long'] = False; coin['is_entering_short'] = False
                        coin['last_close_time'] = None
                        
                        # 🔥 해당 side의 봇이 실행 중이 아니면 자동 시작
                        key = side.lower()
                        if not coin.get(f'{key}_active'):
                            coin[f'{key}_active'] = True
                            
                            # 봇 생성 및 시작
                            if id(coin) not in self.bots:
                                self.bots[id(coin)] = {}
                            
                            if key not in self.bots[id(coin)]:
                                self.bots[id(coin)][key] = TradingBot(
                                    self.api, coin,
                                    lambda msg, pos_type, c=coin: self.add_log(c, pos_type, msg),
                                    lambda: self.update_stats(coin),
                                    self.save_trade_to_excel,  # 🔥 엑셀 콜백!
                                    bot_type=key,
                                    app=self  # 🔥 App 참조 전달
                                )
                            
                            # 봇 시작 (ROI 업데이트를 위해 필수!)
                            self.bots[id(coin)][key].start()
                            print(f"[{coin['symbol']}] 강제 진입 후 {side} 봇 자동 시작 (ROI 0%부터)")
                        
                        CustomMessageBox.showinfo("강제 진입", "강제 진입 완료!\n\n자동 익절/손절이 작동합니다.")
                    else:
                        CustomMessageBox.showwarning("진입 실패", "주문 실패!\n\nAPI 오류를 확인하세요.")
                else:
                    CustomMessageBox.showwarning("진입 실패", "가격 정보를 가져올 수 없습니다!")
        
        # 버튼 가로 1줄 배치
        tk.Button(btn_frame, text="▶️시작", command=start, bg='#00aa00' if side == 'LONG' else '#aa0000',
                 fg='#ffffff', font=('Arial', 9, 'bold'), padx=8, pady=5).pack(side='left', padx=2)
        tk.Button(btn_frame, text="⏸️정지", command=pause, bg='#666666', fg='#ffffff',
                 font=('Arial', 9, 'bold'), padx=8, pady=5).pack(side='left', padx=2)
        tk.Button(btn_frame, text="🔥강제진입", command=force_entry, bg='#0066cc', fg='#ffffff',
                 font=('Arial', 9, 'bold'), padx=8, pady=5).pack(side='left', padx=2)
        tk.Button(btn_frame, text="🛑강제청산", command=force_stop, bg='#ff6600', fg='#ffffff',
                 font=('Arial', 9, 'bold'), padx=8, pady=5).pack(side='left', padx=2)
        
        # 로그
        log_frame = tk.LabelFrame(content, text="📋 로그", bg=color, fg='#ffffff', font=('Arial', 10))
        log_frame.pack(fill='both', expand=True, pady=5)
        
        # 스크롤바 추가
        scroll_frame = tk.Frame(log_frame, bg='#1e1e1e')
        scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        scrollbar = tk.Scrollbar(scroll_frame)
        scrollbar.pack(side='right', fill='y')
        
        log_text = tk.Text(scroll_frame, height=15, bg='#1e1e1e', fg='#ffffff', font=('Arial', 9),
                          yscrollcommand=scrollbar.set, wrap='word')
        log_text.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=log_text.yview)
        
        coin['labels'][f'{key}_log'] = log_text
        
        # 저장된 로그 복원 (전부 복원!)
        log_list = coin.get(f'{key}_logs', [])
        for log_msg in log_list:  # 전체 로그
            log_text.insert('end', log_msg + '\n')
        
        # 자동 스크롤 (맨 아래로)
        log_text.see('end')
    
    def update_stats(self, coin):
        """코인별 통계 업데이트 (스레드 안전!)"""
        def _do_update():
            try:
                stats = coin['stats']
                if 'stats_label' in coin and coin['stats_label'].winfo_exists():
                    total_pnl = stats['total_pnl']
                    total_fee = stats.get('total_fee', 0)
                    pnl_color = '#00ff00' if total_pnl >= 0 else '#ff0000'
                    pnl_sign = '+' if total_pnl >= 0 else ''
                    long_total = stats['long_count']
                    long_win = stats['long_win']
                    long_loss = stats['long_loss']
                    long_fee = stats.get('long_fee', 0)
                    long_pnl_sign = '+' if stats['long_profit'] >= 0 else ''
                    short_total = stats['short_count']
                    short_win = stats['short_win']
                    short_loss = stats['short_loss']
                    short_fee = stats.get('short_fee', 0)
                    short_pnl_sign = '+' if stats['short_profit'] >= 0 else ''
                    stats_text = (f"순수익: {pnl_sign}${total_pnl:.2f} | 수수료: -${total_fee:.2f}\n"
                                f"L:{long_total}회 (익{long_win}/손{long_loss}) {long_pnl_sign}${stats['long_profit']:.2f} (수수료 -${long_fee:.2f})\n"
                                f"S:{short_total}회 (익{short_win}/손{short_loss}) {short_pnl_sign}${stats['short_profit']:.2f} (수수료 -${short_fee:.2f})")
                    coin['stats_label'].config(text=stats_text, fg=pnl_color)
                total_pnl = sum(c['stats']['total_pnl'] for c in self.coins)
                total_fee = sum(c['stats'].get('total_fee', 0) for c in self.coins)
                total_trades = sum(c['stats']['long_count'] + c['stats']['short_count'] for c in self.coins)
                total_wins = sum(c['stats']['long_win'] + c['stats']['short_win'] for c in self.coins)
                self.global_stats['total_pnl'] = total_pnl
                self.global_stats['total_fee'] = total_fee
                self.global_stats['total_trades'] = total_trades
                self.global_stats['total_wins'] = total_wins
                pnl_color = '#00ff00' if total_pnl >= 0 else '#ff0000'
                pnl_sign = '+' if total_pnl >= 0 else ''
                win_rate = (total_wins / total_trades * 100) if total_trades > 0 else 0
                self.global_pnl_label.config(text=f"📊 순수익: {pnl_sign}${total_pnl:.2f}", fg=pnl_color)
                self.global_fee_label.config(text=f"💸 수수료: -${total_fee:.2f}", fg='#ff6666')
                self.global_trades_label.config(text=f"거래: {total_trades}회 | 승률: {win_rate:.1f}%")
            except Exception as e:
                print(f"통계 업데이트 오류: {e}")
        try:
            self.root.after(0, _do_update)
        except:
            pass
    
    def add_log(self, coin, side, message):
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        log_line = f"[{timestamp}] {message}"
        
        key = side.lower()
        
        # 🔥 무조건 리스트에 저장!
        log_key = f'{key}_logs'
        if log_key not in coin:
            coin[log_key] = []
        coin[log_key].append(log_line)
        
        # 🔥 UI가 있으면 실시간 업데이트
        def update_ui():
            if 'labels' in coin and f'{key}_log' in coin['labels']:
                try:
                    log = coin['labels'][f'{key}_log']
                    log.insert('end', log_line + '\n')
                    log.see('end')
                except:
                    pass
        
        # 메인 스레드에서 실행
        try:
            self.root.after(0, update_ui)
        except:
            pass
    
    def update_signals(self, coin):
        try:
            # 위젯이 삭제되었는지 확인
            if 'usdt_label' in coin:
                try:
                    coin['usdt_label'].winfo_exists()
                except:
                    # 위젯이 삭제됨 - 업데이트 중지
                    return
            
            # 🔥 잔고 조회 최적화 (5초마다만)
            current_time = time.time()
            if not hasattr(self, '_last_balance_update'):
                self._last_balance_update = {}
            if coin['symbol'] not in self._last_balance_update or current_time - self._last_balance_update[coin['symbol']] > 5:
                balance = self.api.get_balance()
                coin['_cached_balance'] = balance
                self._last_balance_update[coin['symbol']] = current_time
            else:
                balance = coin.get('_cached_balance', 0)
            
            if 'usdt_label' in coin:
                try:
                    coin['usdt_label'].config(text=f"💰 ${balance:,.0f}")
                except:
                    return
            
            df = self.api.get_klines(coin['symbol'], coin['timeframe'])
            if df is not None and len(df) >= 60:
                # 🔥 완성된 봉만 사용 (copy 제거 - 성능 향상!)
                df_closed = df[:-1]
                
                signals = Indicators.get_signals(df_closed, coin['ut_sens'], coin['ut_atr'],
                                                coin['ema_fast'], coin['ema_slow'])
                
                # 🔥 현재가는 최신 데이터 사용
                current_price = df['close'].iloc[-1]
                
                # 현재가 표시 (실시간 가격)
                if 'price_label' in coin:
                    try:
                        coin['price_label'].config(text=f"💵 ${current_price:,.0f}")
                    except:
                        return
                
                # 디버깅 정보는 제거됨
                
                # UT Bot 상태 표시 (짧게)
                if 'ut_status_label' in coin:
                    try:
                        # 🔥 ut_position 사용 (현재 UT Bot 상태)
                        if signals.get('ut_position_long', False):
                            ut_text = "🤖 🟢L"
                            ut_color = '#00ff00'
                        elif signals.get('ut_position_short', False):
                            ut_text = "🤖 🔴S"
                            ut_color = '#ff0000'
                        else:
                            ut_text = "🤖 ⚪"
                            ut_color = '#aaaaaa'
                        
                        coin['ut_status_label'].config(text=ut_text, fg=ut_color)
                    except:
                        return
                
                # EMA 상태 표시 (짧게)
                if 'ema_status_label' in coin:
                    try:
                        ema_diff = signals['ema_fast_val'] - signals['ema_slow_val']
                        
                        if signals['ema_long']:
                            ema_text = f"📈 🟢G"
                            ema_color = '#00ff00'
                        elif signals['ema_short']:
                            ema_text = f"📉 🔴D"
                            ema_color = '#ff0000'
                        else:
                            ema_text = "📊 ⚪"
                            ema_color = '#aaaaaa'
                        
                        coin['ema_status_label'].config(text=ema_text, fg=ema_color)
                    except:
                        return
                
                # 🔥 ADX 막대바 업데이트 (1시간봉 기준, 60초 캐시 - 봇과 공유)
                if 'adx_canvas' in coin:
                    try:
                        adx_period = coin.get('adx_period', 10)
                        adx_interval = coin.get('adx_interval', '1h')
                        _cache = coin.get('_adx_htf_cache')
                        if _cache and (time.time() - _cache[1]) < 60:
                            adx_value = _cache[0]
                        else:
                            df_adx = df_closed
                            if adx_interval and adx_interval != coin.get('timeframe'):
                                df_h = self.api.get_klines(coin['symbol'], adx_interval)
                                if df_h is not None and len(df_h) > adx_period + 5:
                                    df_adx = df_h[:-1]
                            adx_value = Indicators.calculate_adx(df_adx, period=adx_period)
                            coin['_adx_htf_cache'] = (adx_value, time.time())
                        
                        # 시장 타입 결정
                        if adx_value >= 21:
                            market_type = '추세장'
                            market_color = '#00ff00'
                            current_tp = coin.get('tp_trend', 1.2)  # ✅ 1.2%
                        else:
                            market_type = '횡보장'
                            market_color = '#ffaa00'
                            current_tp = coin.get('tp_sideways', 1.0)  # ✅ 1.0%
                        
                        # ADX 값 라벨 업데이트
                        coin['adx_value_label'].config(text=f"{adx_value:.1f}", fg=market_color)
                        
                        # 시장 타입 라벨 업데이트
                        coin['market_type_label'].config(text=f"({market_type})", fg=market_color)
                        
                        # 🔥 TP 라벨 업데이트 (동적 TP 값 표시)
                        coin['dynamic_tp_label'].config(text=f"TP: {current_tp}%")
                        
                        # 막대바 그리기
                        canvas = coin['adx_canvas']
                        canvas.delete('all')
                        
                        # 배경
                        canvas.create_rectangle(0, 0, 200, 20, fill='#1e1e1e', outline='')
                        
                        # ADX 범위: 0~50 (50 이상은 50으로 표시)
                        bar_width = min(adx_value / 50.0 * 200, 200)
                        
                        # 색상 결정
                        if adx_value >= 21:
                            bar_color = '#00ff00'  # 추세장 (초록)
                        else:
                            bar_color = '#ffaa00'  # 횡보장 (주황)
                        
                        # 막대 그리기
                        if bar_width > 0:
                            canvas.create_rectangle(0, 0, bar_width, 20, fill=bar_color, outline='')
                        
                        # 경계선 (25 표시)
                        boundary_x = 25 / 50.0 * 200  # 25는 전체의 50%
                        canvas.create_line(boundary_x, 0, boundary_x, 20, fill='#ffffff', width=1, dash=(2, 2))
                        
                    except Exception as e:
                        print(f"ADX 업데이트 오류: {e}")
                
                # 🔥 Volume 막대바 업데이트 (로컬 함수 사용)
                if 'volume_canvas' in coin:
                    try:
                        # Volume 정보 가져오기 (완성된 봉 기준!)
                        volume_info = get_volume_info_local(df_closed)
                        current_volume = volume_info['current']
                        avg_volume = volume_info['average']
                        ratio = volume_info['ratio']
                        
                        # Volume 배수 설정 가져오기
                        volume_multiplier = coin.get('volume_multiplier', 1.0)
                        
                        # 데이터가 없거나 비정상일 때
                        if ratio == 0 or avg_volume == 0:
                            coin['volume_ratio_label'].config(text="--", fg='#888888')
                            coin['volume_status_label'].config(text="(로딩중)", fg='#888888')
                            
                            # 막대바 배경만
                            canvas = coin['volume_canvas']
                            canvas.delete('all')
                            canvas.create_rectangle(0, 0, 120, 20, fill='#1e1e1e', outline='')
                            canvas.create_text(60, 10, text="로딩중...", fill='#888888', font=('Arial', 8))
                        else:
                            # Volume 상태 결정
                            if ratio >= volume_multiplier:
                                volume_status = '충분'
                                status_color = '#00ff00'
                            elif ratio >= (volume_multiplier * 0.8):
                                volume_status = '보통'
                                status_color = '#ffaa00'
                            else:
                                volume_status = '부족'
                                status_color = '#ff4444'
                            
                            # Volume 배수 라벨 업데이트
                            coin['volume_ratio_label'].config(text=f"×{ratio:.2f}", fg=status_color)
                            
                            # Volume 상태 라벨 업데이트
                            coin['volume_status_label'].config(text=f"({volume_status})", fg=status_color)
                            
                            # 막대바 그리기
                            canvas = coin['volume_canvas']
                            canvas.delete('all')
                            
                            # 배경
                            canvas.create_rectangle(0, 0, 200, 20, fill='#1e1e1e', outline='')
                            
                            # Volume 범위: 0~3.0배 (3배 이상은 200으로 표시)
                            bar_width = min(ratio / 3.0 * 200, 200)
                            
                            # 색상 결정
                            if ratio >= volume_multiplier:
                                bar_color = '#00ff00'  # 충분 (초록)
                            elif ratio >= (volume_multiplier * 0.8):
                                bar_color = '#ffaa00'  # 보통 (주황)
                            else:
                                bar_color = '#ff4444'  # 부족 (빨강)
                            
                            # 막대 그리기
                            if bar_width > 0:
                                canvas.create_rectangle(0, 0, bar_width, 20, fill=bar_color, outline='')
                            
                            # 경계선 (목표 배수 표시)
                            boundary_x = volume_multiplier / 3.0 * 200
                            if boundary_x > 0 and boundary_x < 200:
                                canvas.create_line(boundary_x, 0, boundary_x, 20, fill='#ffffff', width=1, dash=(2, 2))
                        
                    except Exception as e:
                        print(f"[Volume 로컬] 업데이트 오류: {e}")
                        import traceback
                        traceback.print_exc()
                
                
                # 신호 변경 감지를 위한 이전 상태 저장
                prev_ut_long = coin.get('prev_ut_long')
                prev_ut_short = coin.get('prev_ut_short')
                prev_ema_long = coin.get('prev_ema_long')
                prev_ema_short = coin.get('prev_ema_short')
                
                # 신호가 변경되었는지 확인
                signal_changed = (
                    prev_ut_long != signals['ut_long'] or
                    prev_ut_short != signals['ut_short'] or
                    prev_ema_long != signals['ema_long'] or
                    prev_ema_short != signals['ema_short']
                )
                
                # 신호 변경 시에만 콘솔 출력
                if signal_changed or prev_ut_long is None:
                    print(f"\n[{coin['symbol']} {coin['timeframe']}] 🔔 신호 변경!")
                    print(f"  현재가: ${signals['price']:,.2f}")
                    print(f"  🔥 UT 교차: Long={signals['ut_long']} | Short={signals['ut_short']}")
                    print(f"  EMA: Long={signals['ema_long']} | Short={signals['ema_short']}")
                    
                    # UT Bot 현재 포지션 상태 (참고용)
                    print(f"  📊 UT 포지션 상태: Long={signals.get('ut_position_long', False)} | Short={signals.get('ut_position_short', False)} (스위칭용)")
                
                # 현재 상태 저장
                coin['prev_ut_long'] = signals['ut_long']
                coin['prev_ut_short'] = signals['ut_short']
                coin['prev_ema_long'] = signals['ema_long']
                coin['prev_ema_short'] = signals['ema_short']
                
                # 롱 신호
                if 'labels' in coin:
                    if 'long_ut' in coin['labels']:
                        try:
                            # 🔥 ut_position 사용 (현재 상태)
                            ut_on = signals.get('ut_position_long', False)
                            ut_text = "UT: 🟢 ON" if ut_on else "UT: ⚪ OFF"
                            ut_color = '#00ff00' if ut_on else '#888888'
                            coin['labels']['long_ut'].config(text=ut_text, fg=ut_color)
                        except:
                            pass
                    
                    if 'long_ema' in coin['labels']:
                        try:
                            ema_text = "EMA: 🟢 골든크로스" if signals['ema_long'] else "EMA: ⚪ OFF"
                            ema_color = '#00ff00' if signals['ema_long'] else '#888888'
                            coin['labels']['long_ema'].config(text=ema_text, fg=ema_color)
                        except:
                            pass
                    
                    if 'long_status' in coin['labels']:
                        try:
                            # 🔥 진입 조건: ut_position + ema (둘 다 현재 상태)
                            ut_on = signals.get('ut_position_long', False)
                            if ut_on and signals['ema_long']:
                                coin['labels']['long_status'].config(text="✅ 진입 조건 충족!", fg='#00ff00')
                            else:
                                status = "🔄 실행중..." if coin.get('long_active') else "⏸️ 대기중"
                                coin['labels']['long_status'].config(text=status, fg='#ffaa00')
                        except:
                            pass
                    
                    # 숏 신호
                    if 'short_ut' in coin['labels']:
                        try:
                            # 🔥 ut_position 사용 (현재 상태)
                            ut_on = signals.get('ut_position_short', False)
                            ut_text = "UT: 🔴 ON" if ut_on else "UT: ⚪ OFF"
                            ut_color = '#ff0000' if ut_on else '#888888'
                            coin['labels']['short_ut'].config(text=ut_text, fg=ut_color)
                        except:
                            pass
                    
                    if 'short_ema' in coin['labels']:
                        try:
                            ema_text = "EMA: 🔴 데드크로스" if signals['ema_short'] else "EMA: ⚪ OFF"
                            ema_color = '#ff0000' if signals['ema_short'] else '#888888'
                            coin['labels']['short_ema'].config(text=ema_text, fg=ema_color)
                        except:
                            pass
                    
                    if 'short_status' in coin['labels']:
                        try:
                            # 🔥 진입 조건: ut_position + ema (둘 다 현재 상태)
                            ut_on = signals.get('ut_position_short', False)
                            if ut_on and signals['ema_short']:
                                coin['labels']['short_status'].config(text="✅ 진입 조건 충족!", fg='#ff0000')
                            else:
                                status = "🔄 실행중..." if coin.get('short_active') else "⏸️ 대기중"
                                coin['labels']['short_status'].config(text=status, fg='#ffaa00')
                        except:
                            pass
                
                # 📊 ROI 업데이트는 별도 함수에서 처리 (_update_roi_display)
                
        except Exception as e:
            print(f"[⚠️ GUI 업데이트 오류] {coin.get('symbol', 'Unknown')}: {e}")
            # 오류 발생해도 계속 실행
        
        # 🔥🔥🔥 ROI 업데이트는 봇이 실행 중일 때만!
        # 봇이 실행 중이 아니면 ROI 표시 안 함
        if coin.get('long_active') or coin.get('short_active'):
            try:
                self._update_roi_display(coin)
            except Exception as e:
                print(f"[⚠️ ROI 업데이트 오류] {coin.get('symbol', 'Unknown')}: {e}")
        
        # 1초 후 재실행 (빠른 업데이트)
        try:
            self.root.after(1000, lambda: self.update_signals(coin))
        except:
            # 위젯이 삭제되었으면 재실행 안 함
            print(f"[⚠️ GUI 종료] {coin.get('symbol', 'Unknown')}: 업데이트 중단")
            pass
    
    def _update_roi_display(self, coin):
        """🔥 바이낸스 실시간 ROI 업데이트 (봇 루프 캐시 사용 — API 호출 없음!)"""
        position = coin.get('_cached_position')
        self._apply_roi_update(coin, position)
    
    def _apply_roi_update(self, coin, position):
        """ROI 데이터를 UI에 적용 (메인 스레드에서 호출)"""
        try:
        
            # 청산 중이면 "-" 표시
            if coin.get('is_closing'):
                coin['roi']['long_entry'] = None
                coin['roi']['short_entry'] = None
                coin['roi']['long_current'] = 0
                coin['roi']['short_current'] = 0
                coin['roi']['long_max'] = 0
                coin['roi']['short_max'] = 0
                coin['roi']['long_min'] = 0
                coin['roi']['short_min'] = 0
            
                for key in ['long', 'short']:
                    if f'{key}_current_roi' in coin.get('labels', {}):
                        try:
                            coin['labels'][f'{key}_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels'][f'{key}_max_roi'].config(text="최고: -")
                            coin['labels'][f'{key}_min_roi'].config(text="최저: -")
                        except:
                            pass
                return
        
            # 🔥 포지션이 있으면 ROI 계산!
            if position:
                entry_price = position['entry_price']
                pnl = position.get('pnl', 0)
                side = position['side']
                amount = position.get('amount', 0)
                leverage = position.get('leverage', coin.get('leverage', 5))
            
                # 🔥🔥🔥 ROI 직접 계산! (API에 roi_pct 없을 경우 대비)
                if 'roi_pct' in position:
                    roi_pct = position['roi_pct']
                else:
                    # 직접 계산: ROI = (PNL / 초기마진) * 100
                    initial_margin = (entry_price * amount) / leverage
                    if initial_margin > 0:
                        roi_pct = (pnl / initial_margin) * 100
                    else:
                        roi_pct = 0
            
                if side == 'long':
                    coin['roi']['long_entry'] = entry_price
                    coin['roi']['long_current'] = roi_pct
                
                    # 최고/최저 업데이트
                    if roi_pct > 0 and roi_pct > coin['roi'].get('long_max', 0):
                        coin['roi']['long_max'] = roi_pct
                    if roi_pct < 0 and roi_pct < coin['roi'].get('long_min', 0):
                        coin['roi']['long_min'] = roi_pct
                
                    # UI 업데이트
                    if 'long_current_roi' in coin.get('labels', {}):
                        try:
                            roi_color = '#00ff00' if roi_pct >= 0 else '#ff0000'
                            roi_sign = '+' if roi_pct >= 0 else ''
                            coin['labels']['long_current_roi'].config(
                                text=f"현재: {roi_sign}{roi_pct:.2f}%", fg=roi_color
                            )
                            max_val = max(0, coin['roi'].get('long_max', 0))
                            min_val = min(0, coin['roi'].get('long_min', 0))
                            coin['labels']['long_max_roi'].config(text=f"최고: +{max_val:.2f}%")
                            coin['labels']['long_min_roi'].config(text=f"최저: {min_val:.2f}%")
                        except:
                            pass
                
                    # SHORT는 "-" 표시
                    if 'short_current_roi' in coin.get('labels', {}):
                        try:
                            coin['labels']['short_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels']['short_max_roi'].config(text="최고: -")
                            coin['labels']['short_min_roi'].config(text="최저: -")
                        except:
                            pass
            
                else:  # short
                    coin['roi']['short_entry'] = entry_price
                    coin['roi']['short_current'] = roi_pct
                
                    # 최고/최저 업데이트
                    if roi_pct > 0 and roi_pct > coin['roi'].get('short_max', 0):
                        coin['roi']['short_max'] = roi_pct
                    if roi_pct < 0 and roi_pct < coin['roi'].get('short_min', 0):
                        coin['roi']['short_min'] = roi_pct
                
                    # UI 업데이트
                    if 'short_current_roi' in coin.get('labels', {}):
                        try:
                            roi_color = '#00ff00' if roi_pct >= 0 else '#ff0000'
                            roi_sign = '+' if roi_pct >= 0 else ''
                            coin['labels']['short_current_roi'].config(
                                text=f"현재: {roi_sign}{roi_pct:.2f}%", fg=roi_color
                            )
                            max_val = max(0, coin['roi'].get('short_max', 0))
                            min_val = min(0, coin['roi'].get('short_min', 0))
                            coin['labels']['short_max_roi'].config(text=f"최고: +{max_val:.2f}%")
                            coin['labels']['short_min_roi'].config(text=f"최저: {min_val:.2f}%")
                        except:
                            pass
                
                    # LONG은 "-" 표시
                    if 'long_current_roi' in coin.get('labels', {}):
                        try:
                            coin['labels']['long_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels']['long_max_roi'].config(text="최고: -")
                            coin['labels']['long_min_roi'].config(text="최저: -")
                        except:
                            pass
        
            else:
                # 🔥 포지션 없으면 완전 초기화!
                coin['roi']['long_entry'] = None
                coin['roi']['short_entry'] = None
                coin['roi']['long_current'] = 0
                coin['roi']['short_current'] = 0
                coin['roi']['long_max'] = 0
                coin['roi']['short_max'] = 0
                coin['roi']['long_min'] = 0
                coin['roi']['short_min'] = 0
                coin['chart_entry_long'] = None
                coin['chart_entry_short'] = None
                coin['restart_entry_long'] = None
                coin['restart_entry_short'] = None
            
                # is_closing 해제
                if coin.get('is_closing'):
                    coin['is_closing'] = False
            
                # UI는 "-"로 표시
                for key in ['long', 'short']:
                    if f'{key}_current_roi' in coin.get('labels', {}):
                        try:
                            coin['labels'][f'{key}_current_roi'].config(text="현재: -", fg='#ffffff')
                            coin['labels'][f'{key}_max_roi'].config(text="최고: -")
                            coin['labels'][f'{key}_min_roi'].config(text="최저: -")
                        except:
                            pass

        except Exception as e:
            pass

def main():
    try:
        root = tk.Tk()
        app = App(root)
        root.mainloop()
    except Exception as e:
        print("\n" + "="*60)
        print("❌ 프로그램 실행 중 오류 발생!")
        print("="*60)
        print(f"\n오류 내용: {e}\n")
        import traceback
        traceback.print_exc()
        print("\n" + "="*60)
        print("위 오류 메시지를 캡처해서 개발자에게 보내주세요!")
        print("="*60)
        input("\n아무 키나 눌러 종료...")

if __name__ == "__main__":
    main()

